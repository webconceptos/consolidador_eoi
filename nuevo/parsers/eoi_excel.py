
# parsers/eoi_excel.py
# -*- coding: utf-8 -*-
"""
Parser EOI Excel (EDI) - Consolidado

Este módulo concentra la lógica de parseo desde Excel.
Se asume openpyxl como engine.
"""

from __future__ import annotations

from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional

from openpyxl import load_workbook

from .utils import norm, _parse_date_any, _days_between


def _is_desc_label_row(ws, r: int) -> bool:
    """
    Detecta fila que contiene el rótulo "Descripción del trabajo Realizado:"
    (suele estar en una celda combinada)
    """
    for c in range(1, 15):
        v = ws.cell(r, c).value
        if v is None:
            continue
        t = norm(str(v)).lower()
        if "descripción del trabajo" in t or "descripcion del trabajo" in t:
            return True
    return False


def _get_desc_detail(ws, r: int) -> str:
    """
    Obtiene el detalle de la descripción. Suele estar en una celda combinada.
    Tomamos el texto más largo de la fila.
    """
    best = ""
    for c in range(1, 25):
        v = ws.cell(r, c).value
        if v is None:
            continue
        t = norm(str(v))
        if len(t) > len(best):
            best = t
    return best


def parse_experiencia_general_excel(
    ws,
    start_row: int = 1,
    end_row: int = 500,
    debug: bool = False
) -> Dict[str, Any]:
    """
    Experiencia General (EDI) - FORMATO REAL (bloques con descripción):
    - Fila datos: C..J
      C: nro
      D/E: entidad
      F: proyecto
      G: cargo/servicio
      H: fecha_inicio
      I: fecha_fin
      J: tiempo_en_cargo (texto)
    - Luego:
      + fila con texto "Descripción del trabajo Realizado:"
      + fila siguiente con el detalle (celda combinada)
    (El encabezado se repite por cada experiencia)
    """
    COL_C = 3  # C
    items: List[Dict[str, Any]] = []
    lines: List[str] = []
    seen = set()

    r = start_row
    while r <= end_row:
        nro = norm(str(ws.cell(r, COL_C + 0).value or ""))

        entidad = " ".join([
            norm(str(ws.cell(r, COL_C + 1).value or "")),  # D
            norm(str(ws.cell(r, COL_C + 2).value or "")),  # E
        ]).strip()

        proyecto = norm(str(ws.cell(r, COL_C + 3).value or ""))  # F
        cargo = norm(str(ws.cell(r, COL_C + 4).value or ""))     # G
        fi = ws.cell(r, COL_C + 5).value                         # H
        ff = ws.cell(r, COL_C + 6).value                         # I
        tiempo = norm(str(ws.cell(r, COL_C + 7).value or ""))    # J

        base_has_data = any([entidad, proyecto, cargo, fi, ff, tiempo, nro])

        if _is_desc_label_row(ws, r):
            r += 1
            continue

        if not base_has_data:
            r += 1
            continue

        base_text = " ".join([nro, entidad, proyecto, cargo, str(fi), str(ff), tiempo]).upper()
        if ("NRO" in base_text and "ENTIDAD" in base_text and "PROYECTO" in base_text) or ("CARGO" in base_text and "FECHA" in base_text):
            r += 1
            continue

        # descripción asociada
        descripcion = ""
        if r + 1 <= end_row and _is_desc_label_row(ws, r + 1):
            if r + 2 <= end_row:
                descripcion = _get_desc_detail(ws, r + 2)
            next_r = r + 3
        else:
            next_r = r + 1

        # normaliza fecha fin tipo "Actualidad"
        ff_str = norm(str(ff or ""))
        ff_up = ff_str.upper()
        if ff_up in ("ACTUALIDAD", "ACTUAL", "A LA FECHA", "HASTA LA FECHA"):
            ff_str = "ACTUALIDAD"

        d_fi = _parse_date_any(fi)
        d_ff = datetime.now() if ff_str == "ACTUALIDAD" else _parse_date_any(ff)
        dias = _days_between(d_fi, d_ff)

        key = (
            entidad.lower(), proyecto.lower(), cargo.lower(),
            norm(str(fi or "")), ff_str, tiempo.lower(),
            descripcion.lower()
        )
        if key in seen:
            r = next_r
            continue
        seen.add(key)

        it = {
            "row": r,
            "nro": nro,
            "entidad": entidad,
            "proyecto": proyecto,
            "cargo": cargo,
            "fecha_inicio": fi,
            "fecha_fin": ff_str if ff_str == "ACTUALIDAD" else ff,
            "tiempo_en_cargo": tiempo,
            "dias_calc": dias,
            "descripcion": descripcion,
        }
        items.append(it)

        head = " | ".join([p for p in [entidad, proyecto, cargo] if p]).strip()
        fechas = " - ".join([p for p in [norm(str(fi or "")), ff_str] if p]).strip(" -")
        tail = " | ".join([p for p in [fechas, tiempo] if p]).strip()
        line = " | ".join([p for p in [head, tail] if p]).strip()
        if descripcion:
            line += f"\n  Desc: {descripcion}"
        lines.append(line)

        r = next_r

    total_dias = sum(int(x.get("dias_calc") or 0) for x in items)
    total_anios = round(total_dias / 365.0, 2) if total_dias else 0.0
    resumen = "\n\n".join([x for x in lines if x]).strip()

    out: Dict[str, Any] = {
        "items": items,
        "total_dias_calc": total_dias,
        "total_anios_calc": total_anios,
        "resumen": resumen,
    }

    if debug:
        print("[DEBUG EG] range:", start_row, "->", end_row, "| items:", len(items))
        print("  total_dias_calc:", total_dias, "| total_anios_calc:", total_anios)

    return out


def parse_eoi_excel(path: Path, sheet_name: Optional[str] = None, debug: bool = False) -> Dict[str, Any]:
    """
    Parseador principal de EOI en Excel.
    Por ahora retorna:
      - exp_general (items + totales)
      - exp_general_dias
    Se puede ampliar con datos personales, formación, etc.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    # Búsqueda dinámica del bloque de experiencia general:
    start_row = 1
    end_row = ws.max_row

    for r in range(1, min(ws.max_row, 800) + 1):
        row_text = " ".join([norm(str(ws.cell(r, c).value or "")) for c in range(1, 15)])
        if "experiencia general" in row_text.lower():
            start_row = r
            break

    exp_general = parse_experiencia_general_excel(ws, start_row=start_row, end_row=end_row, debug=debug)

    out = {
        "exp_general": exp_general,
        "exp_general_items": exp_general.get("items", []),
        "exp_general_resumen": exp_general.get("resumen", ""),
        "exp_general_dias": exp_general.get("total_dias_calc", 0),
    }
    return out
