# parsers/eoi_excel.py
# -*- coding: utf-8 -*-
"""
Parser EOI (Excel) - CONSOLIDADOR EOI

(Archivo base del usuario; SOLO se corrigió la parte de EXPERIENCIA para:
 - no considerar filas de encabezado repetidas / "basura"
 - capturar correctamente el bloque "Descripción del Trabajo Realizado" + su detalle
 - evitar que la descripción se contamine con texto de otras secciones)
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime, date

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ============================================================
# Utils
# ============================================================
def norm(x: Any) -> str:
    if x is None:
        return ""
    s = str(x).replace("\u00a0", " ").strip()
    return re.sub(r"\s+", " ", s)


def cell_raw(ws: Worksheet, r: int, c: int) -> Any:
    return ws.cell(row=r, column=c).value


def cell_str(ws: Worksheet, r: int, c: int) -> str:
    return norm(cell_raw(ws, r, c))


def row_text(ws: Worksheet, r: int, c1: int = 1, c2: int = 15) -> str:
    parts: List[str] = []
    for c in range(c1, c2 + 1):
        v = ws.cell(row=r, column=c).value
        if v is None:
            continue
        s = norm(v)
        if s:
            parts.append(s)
    return " | ".join(parts)


def normalize_email(x: str) -> str:
    x = norm(x).lower().replace(" ", "")
    return x


def normalize_phone(x: str) -> str:
    x = norm(x)
    d = re.sub(r"\D+", "", x)
    if len(d) >= 9:
        return d[-9:]
    return d


def normalize_dni(x: str) -> str:
    d = re.sub(r"\D+", "", norm(x))
    if len(d) >= 8:
        return d[-8:]
    return d


def as_date_str(x: Any) -> str:
    if x is None:
        return ""
    if isinstance(x, datetime):
        return x.strftime("%d/%m/%Y")
    if isinstance(x, date):
        return x.strftime("%d/%m/%Y")
    s = norm(x)
    if not s:
        return ""
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.strftime("%d/%m/%Y")
        except Exception:
            pass
    try:
        dt = datetime.fromisoformat(s.replace("Z", ""))
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return s


def parse_date_any(x: Any) -> Optional[datetime]:
    if x is None:
        return None
    if isinstance(x, datetime):
        return x
    if isinstance(x, date):
        return datetime(x.year, x.month, x.day)
    s = norm(x)
    if not s:
        return None
    s = s.replace(".", "/").replace("-", "/")
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    try:
        return datetime.fromisoformat(norm(x).replace("Z", ""))
    except Exception:
        return None


def days_between(d1: Optional[datetime], d2: Optional[datetime]) -> int:
    if not d1 or not d2:
        return 0
    if d2 < d1:
        return 0
    return int((d2 - d1).days) + 1


def find_best_sheet(wb) -> Worksheet:
    best = wb.worksheets[0]
    best_score = -1
    for ws in wb.worksheets:
        score = 0
        for r in range(1, min(ws.max_row, 80) + 1):
            t = row_text(ws, r, 1, 12).upper()
            if "DATOS PERSONALES" in t:
                score += 5
            if "FORMACIÓN ACADÉMICA" in t or "FORMACION ACADEMICA" in t:
                score += 3
            if "ESTUDIOS COMPLEMENTARIOS" in t:
                score += 3
            if "EXPERIENCIA" in t and "IV." in t:
                score += 3
        if score > best_score:
            best_score = score
            best = ws
    return best


# ============================================================
# 1) Datos Personales (igual)
# ============================================================
def _find_label_cell(ws: Worksheet, label_regex: str, r1: int, r2: int, c1: int, c2: int) -> Optional[Tuple[int, int]]:
    rgx = re.compile(label_regex, re.IGNORECASE)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            v = cell_str(ws, r, c)
            if v and rgx.search(v):
                return (r, c)
    return None


def _value_below(ws: Worksheet, r: int, c: int, max_down: int = 2) -> str:
    for k in range(1, max_down + 1):
        v = cell_raw(ws, r + k, c)
        s = norm(v)
        if s:
            return s
    return ""


def _value_right(ws: Worksheet, r: int, c: int, max_right: int = 3) -> str:
    for k in range(1, max_right + 1):
        s = cell_str(ws, r, c + k)
        if s:
            return s
    return ""


def parse_datos_personales(ws: Worksheet, debug: bool = False) -> Dict[str, Any]:
    out = {
        "dni": "",
        "apellido_paterno": "",
        "apellido_materno": "",
        "nombres": "",
        "nombre_full": "",
        "email": "",
        "celular": "",
    }

    R1, R2 = 1, 40
    C1, C2 = 1, 12

    pos = _find_label_cell(ws, r"\bApellido\s*Paterno\b", R1, R2, C1, C2)
    if pos:
        out["apellido_paterno"] = _value_below(ws, *pos) or _value_right(ws, *pos)

    pos = _find_label_cell(ws, r"\bApellido\s*Materno\b", R1, R2, C1, C2)
    if pos:
        out["apellido_materno"] = _value_below(ws, *pos) or _value_right(ws, *pos)

    pos = _find_label_cell(ws, r"\bNombres\b", R1, R2, C1, C2)
    if pos:
        out["nombres"] = _value_below(ws, *pos) or _value_right(ws, *pos)

    pos = _find_label_cell(ws, r"Documento\s+de\s+identidad|DNI", R1, R2, C1, C2)
    if pos:
        out["dni"] = normalize_dni(_value_below(ws, *pos) or _value_right(ws, *pos))

    pos = _find_label_cell(ws, r"\bCelular\b", R1, R2, C1, C2)
    if pos:
        out["celular"] = normalize_phone(_value_below(ws, *pos) or _value_right(ws, *pos))
    else:
        pos = _find_label_cell(ws, r"\bTel[ée]fono\b", R1, R2, C1, C2)
        if pos:
            out["celular"] = normalize_phone(_value_below(ws, *pos) or _value_right(ws, *pos))

    pos = _find_label_cell(ws, r"\bemail\b|correo", R1, R2, C1, C2)
    if pos:
        out["email"] = normalize_email(_value_below(ws, *pos) or _value_right(ws, *pos))

    ap = " ".join([out["apellido_paterno"], out["apellido_materno"]]).strip()
    nm = out["nombres"].strip()
    out["nombre_full"] = norm(" ".join([ap, nm]))

    if debug:
        print("[DP]", out)

    return out


# ============================================================
# 2) Formación Académica (igual)
# ============================================================
def _find_row_contains(ws: Worksheet, text_regex: str, r1: int, r2: int, c1: int = 1, c2: int = 12) -> Optional[int]:
    rgx = re.compile(text_regex, re.IGNORECASE)
    for r in range(r1, r2 + 1):
        if rgx.search(row_text(ws, r, c1, c2)):
            return r
    return None


def parse_formacion_obligatoria(ws: Worksheet, debug: bool = False) -> Dict[str, Any]:
    header_like = _find_row_contains(ws, r"COLEGIATURA|MAESTRIA|TITULO|BACHILLER|EGRESADO\s+UNIVERSITARIO", 40, 80, 1, 12)
    if not header_like:
        start_row, end_row = 51, 56
    else:
        start_row = 51
        er = _find_row_contains(ws, r"EGRESADO\s+UNIVERSITARIO", start_row, start_row + 25, 1, 12)
        end_row = er if er else (start_row + 5)

    colmap = {"titulo_item": 3, "especialidad": 6, "fecha": 7, "centro": 8, "ciudad": 10}

    items: List[Dict[str, Any]] = []
    resumen_parts: List[str] = []

    for r in range(start_row, end_row + 1):
        titulo = cell_str(ws, r, colmap["titulo_item"])
        esp = cell_str(ws, r, colmap["especialidad"])
        fec = as_date_str(cell_raw(ws, r, colmap["fecha"]))
        cen = cell_str(ws, r, colmap["centro"])
        ciu = cell_str(ws, r, colmap["ciudad"])

        has_data = any([esp, fec, cen, ciu])
        
        if titulo and has_data:
            it = {
                "row": r,
                "titulo_item": titulo,
                "especialidad": esp,
                "fecha": fec,
                "centro": cen,
                "ciudad": ciu,
                "has_data": bool(has_data),
            }
            items.append(it)

            parts = [p for p in [fec, cen, ciu] if p]
            resumen_parts.append(f"{titulo}: {esp} ({' | '.join(parts)})" if esp else f"{titulo}: ({' | '.join(parts)})")

    resumen = " ; ".join(resumen_parts).strip()

    out = {
        "items": items,
        "resumen": resumen,
        "meta": {"start_row": start_row, "end_row": end_row, "colmap": colmap},
    }

    if debug:
        print("[FA]", out["meta"], "items=", len(items))

    return out


# ============================================================
# 3) Estudios Complementarios (igual)
# ============================================================
def _is_stop_row_for_blocks(ws: Worksheet, r: int) -> bool:
    t = row_text(ws, r, 1, 12).upper()
    return ("IV." in t and "EXPERIENCIA" in t) or ("IV." in t and " EXPERIENCIA" in t)


def _parse_block_table(ws: Worksheet, title_row: int, debug: bool = False) -> Dict[str, Any]:
    title = cell_str(ws, title_row, 2) or row_text(ws, title_row, 1, 12)
    header_row = None
    for r in range(title_row, min(title_row + 8, ws.max_row) + 1):
        t = row_text(ws, r, 1, 12).upper()
        if "NO." in t and ("CENTRO" in t or "CAPACIT" in t) and ("FECHA" in t or "HORAS" in t):
            header_row = r
            break
    if not header_row:
        return {"row": title_row, "title": title, "items": [], "total_horas": 0, "resumen": ""}

    col_nro = 3
    col_centro = 4
    col_cap = 6
    col_ini = 8
    col_fin = 9
    col_horas = 10

    items: List[Dict[str, Any]] = []
    resumen_lines: List[str] = []
    total_horas = 0

    r = header_row + 2
    while r <= ws.max_row:
        if _is_stop_row_for_blocks(ws, r):
            break

        nro = cell_str(ws, r, col_nro)
        centro = cell_str(ws, r, col_centro)
        cap = cell_str(ws, r, col_cap)
        fi = as_date_str(cell_raw(ws, r, col_ini))
        ff = as_date_str(cell_raw(ws, r, col_fin))
        horas_raw = cell_raw(ws, r, col_horas)

        tt = row_text(ws, r, 1, 12)
        if re.search(r"Puede\s+adicionar", tt, re.IGNORECASE):
            break

        if not any([nro, centro, cap, fi, ff, horas_raw]):
            r += 1
            continue

        horas = 0
        try:
            horas = int(float(horas_raw)) if horas_raw not in (None, "") else 0
        except Exception:
            horas = 0

        it = {
            "row": r,
            "nro": nro,
            "centro": centro,
            "capacitacion": cap,
            "fecha_inicio": fi,
            "fecha_fin": ff,
            "horas": horas,
        }
        items.append(it)
        total_horas += horas

        if centro or cap:
            resumen_lines.append(f"{centro} - {cap} ({fi} | {ff} | {horas}h)".strip())

        r += 1

    return {
        "row": title_row,
        "title": title,
        "items": items,
        "total_horas": total_horas,
        "resumen": "\n".join(resumen_lines).strip(),
    }


def parse_estudios_complementarios(ws: Worksheet, debug: bool = False) -> Dict[str, Any]:
    blocks: List[Dict[str, Any]] = []
    for r in range(1, min(ws.max_row, 200) + 1):
        if _is_stop_row_for_blocks(ws, r):
            break
        t = row_text(ws, r, 1, 12)
        if re.search(r"\bb\.\d\)", t, re.IGNORECASE):
            b = _parse_block_table(ws, r, debug=debug)
            m = re.search(r"\b(b\.\d)\)", t, re.IGNORECASE)
            bid = m.group(1).lower() if m else f"b.{len(blocks)+1}"
            b["id"] = bid
            blocks.append(b)

    total_horas = sum(int(b.get("total_horas") or 0) for b in blocks)
    resumen_parts: List[str] = []
    for b in blocks:
        bid = (b.get("id") or "").upper()
        body = b.get("resumen") or ""
        resumen_parts.append(f"{bid}:\n{body}" if body else f"{bid}:\n(sin cursos declarados)")

    return {
        "blocks": blocks,
        "total_horas": total_horas,
        "resumen": "\n\n".join(resumen_parts).strip(),
    }


# ============================================================
# 4) Experiencia (GENERAL / ESPECIFICA) - ✅ CORREGIDO
# ============================================================
def _is_desc_label_row(ws: Worksheet, r: int) -> bool:
    t = row_text(ws, r, 1, 12).upper()
    return ("DESCRIPCIÓN DEL TRABAJO" in t) or ("DESCRIPCION DEL TRABAJO" in t)


def _looks_like_exp_header_row_text(t: str) -> bool:
    tu = (t or "").upper()
    # header típico: "No." + "Entidad" + "Fecha"
    return ("NO." in tu or "N°" in tu) and ("ENTIDAD" in tu or "EMPRESA" in tu) and ("FECHA" in tu)


def _looks_like_section_start(t: str) -> bool:
    tu = (t or "").upper()
    return bool(
        re.search(r"^\s*(IV|V)\.", tu)
        or re.search(r"\bA\)\s*EXPERIENCIA\b", tu)
        or re.search(r"\bB\)\s*EXPERIENCIA\b", tu)
        or re.search(r"\bEXPERIENCIA\s+GENERAL\b", tu)
        or re.search(r"\bEXPERIENCIA\s+ESPECIFICA\b", tu)
    )


def _looks_like_day_month_year_row(t: str) -> bool:
    tu = (t or "").upper().replace("Í", "I")
    return "DIA/MES/ANO" in tu or "DÍA/MES/AÑO" in tu


def _clean_desc(desc: str) -> str:
    s = (desc or "").strip()
    if not s:
        return ""

    # corta contaminación típica dentro de la descripción
    cut_patterns = [
        r"\bb\)\s*EXPERIENCIA\s+ESPECIFICA\b.*",
        r"\ba\)\s*EXPERIENCIA\s+GENERAL\b.*",
        r"\bIV\.\s*EXPERIENCIA\b.*",
        r"\bV\.\b.*",
        r"\bTiempo\s+en\s+el\s+Cargo\b.*",
    ]
    for pat in cut_patterns:
        s = re.sub(pat, "", s, flags=re.IGNORECASE | re.DOTALL).strip()

    # normaliza separadores
    s = s.replace(" | ", " ").strip()
    return s


def _read_desc_block(ws: Worksheet, start_row: int, max_lines: int = 25) -> Tuple[str, int]:
    """
    Lee el bloque de descripción desde start_row (primer row con texto del detalle)
    hasta encontrar un "corte" (nuevo registro/header/sección/puede adicionar/vacío).
    Retorna (descripcion_limpia, next_row).
    """
    lines: List[str] = []
    r = start_row

    for _ in range(max_lines):
        if r > ws.max_row:
            break

        trow = row_text(ws, r, 1, 12)

        # cortes fuertes
        if not norm(trow):
            break
        if re.search(r"Puede\s+adicionar", trow, re.IGNORECASE):
            break
        if _looks_like_exp_header_row_text(trow):
            break
        if _looks_like_section_start(trow):
            break

        # si aparece "Descripción del Trabajo Realizado" otra vez, no la dupliques
        if _is_desc_label_row(ws, r):
            r += 1
            continue

        # preferir celda C (3) si existe, sino toda la fila
        line = cell_str(ws, r, 3) or trow
        line = _clean_desc(line)
        if line:
            lines.append(line)

        r += 1

    # Une líneas: si excel ya trae bullets en una sola celda, aquí no lo rompes.
    desc = "\n".join(lines).strip()
    desc = _clean_desc(desc)

    return desc, r


def _find_section_anchor(ws: Worksheet, anchor_regex: str, r1: int = 1, r2: Optional[int] = None) -> Optional[int]:
    if r2 is None:
        r2 = ws.max_row
    rgx = re.compile(anchor_regex, re.IGNORECASE)
    for r in range(r1, r2 + 1):
        if rgx.search(row_text(ws, r, 1, 12)):
            return r
    return None


def _find_exp_header_row(ws: Worksheet, anchor_row: int) -> Optional[int]:
    for r in range(anchor_row, min(anchor_row + 14, ws.max_row) + 1):
        t = row_text(ws, r, 1, 12)
        if _looks_like_exp_header_row_text(t):
            return r
    return None


def _parse_experiencia_from_header(ws: Worksheet, anchor_row: int, debug: bool = False) -> Dict[str, Any]:
    header_row = _find_exp_header_row(ws, anchor_row)
    if not header_row:
        return {"items": [], "total_dias_calc": 0, "resumen": "", "_meta": {"anchor_row": anchor_row, "header_row": None}}

    # Layout observado (constante en tu formato):
    COL_C = 3
    col_nro = COL_C + 0
    col_ent1 = COL_C + 1
    col_ent2 = COL_C + 2
    col_proy = COL_C + 3
    col_cargo = COL_C + 4
    col_ini = COL_C + 5
    col_fin = COL_C + 6
    col_tiempo = COL_C + 7

    def nro_ok(v: str) -> bool:
        v = norm(v)
        return bool(re.fullmatch(r"\d+", v))

    items: List[Dict[str, Any]] = []
    resumen_lines: List[str] = []

    r = header_row + 1
    # salta filas "Día/Mes/Año" si existen
    while r <= ws.max_row and _looks_like_day_month_year_row(row_text(ws, r, 1, 12)):
        r += 1

    while r <= ws.max_row:
        trow = row_text(ws, r, 1, 12)

        # cortes
        if re.search(r"Puede\s+adicionar", trow, re.IGNORECASE):
            break
        if r > anchor_row and re.search(r"^\s*b\)\s+EXPERIENCIA", trow, re.IGNORECASE):
            break
        if _looks_like_section_start(trow) and r > header_row + 1:
            # OJO: evita cortar en el propio ancla/header
            break
        if _looks_like_exp_header_row_text(trow):
            # encabezado repetido dentro del bloque
            r += 1
            continue
        if _is_desc_label_row(ws, r):
            # fila basura (título de descripción sin estar asociada a un registro)
            r += 1
            continue
        if _looks_like_day_month_year_row(trow):
            r += 1
            continue

        nro = norm(cell_raw(ws, r, col_nro))
        # fila basura si "No." o vacío
        if not nro_ok(str(nro)):
            # si está completamente vacía, avanzar; si es texto, también (basura)
            if not norm(trow):
                r += 1
                continue
            r += 1
            continue

        entidad = " ".join([cell_str(ws, r, col_ent1), cell_str(ws, r, col_ent2)]).strip()
        proyecto = cell_str(ws, r, col_proy)
        cargo = cell_str(ws, r, col_cargo)
        fi_raw = cell_raw(ws, r, col_ini)
        ff_raw = cell_raw(ws, r, col_fin)
        tiempo = cell_str(ws, r, col_tiempo)

        fi = as_date_str(fi_raw)
        ff = as_date_str(ff_raw)

        descripcion = ""
        next_r = r + 1

        # buscar etiqueta de descripción dentro de las siguientes 5 filas
        desc_label_row = None
        for rr in range(r + 1, min(r + 6, ws.max_row) + 1):
            if _is_desc_label_row(ws, rr):
                desc_label_row = rr
                break
            # si aparece un nuevo nro antes, no hay descripción para este registro
            if nro_ok(cell_str(ws, rr, col_nro)):
                break
            # si aparece header/sección, cortar
            if _looks_like_exp_header_row_text(row_text(ws, rr, 1, 12)) or _looks_like_section_start(row_text(ws, rr, 1, 12)):
                break

        if desc_label_row:
            descripcion, next_r = _read_desc_block(ws, desc_label_row + 1)
        else:
            descripcion = ""

        d_fi = parse_date_any(fi_raw)
        d_ff = parse_date_any(ff_raw)
        dias = days_between(d_fi, d_ff)

        it = {
            "row": r,
            "nro": str(nro),
            "entidad": entidad,
            "proyecto": proyecto,
            "cargo": cargo,
            "fecha_inicio": fi,
            "fecha_fin": ff,
            "tiempo_en_cargo": tiempo,
            "dias_calc": dias,
            "descripcion": descripcion,
        }
        items.append(it)

        head = " - ".join([p for p in [entidad, cargo] if p]).strip(" -")
        fechas = " a ".join([p for p in [fi, ff] if p]).strip(" a")
        line = " | ".join([p for p in [head, fechas] if p]).strip()
        if descripcion:
            line += f"\n  Desc: {descripcion}"
        resumen_lines.append(line)

        r = max(next_r, r + 1)

    total_dias = sum(int(x.get("dias_calc") or 0) for x in items)
    resumen = "\n\n".join([x for x in resumen_lines if x]).strip()

    if debug:
        print(f"[EXP] anchor={anchor_row} header={header_row} items={len(items)} total_dias={total_dias}")

    return {
        "items": items,
        "total_dias_calc": total_dias,
        "resumen": resumen,
        "_meta": {"anchor_row": anchor_row, "header_row": header_row, "start_row": header_row + 1},
    }


def parse_experiencia_general(ws: Worksheet, debug: bool = False) -> Dict[str, Any]:
    anchor = _find_section_anchor(ws, r"a\)\s*EXPERIENCIA\s+GENERAL", 1, ws.max_row)
    if not anchor:
        anchor = _find_section_anchor(ws, r"EXPERIENCIA\s+GENERAL", 1, ws.max_row)
    if not anchor:
        return {"items": [], "total_dias_calc": 0, "resumen": "", "_meta": {"anchor_row": None}}
    return _parse_experiencia_from_header(ws, anchor, debug=debug)


def parse_experiencia_especifica(ws: Worksheet, debug: bool = False) -> Dict[str, Any]:
    anchor = _find_section_anchor(ws, r"b\)\s*EXPERIENCIA\s+ESPECIFICA", 1, ws.max_row)
    if not anchor:
        anchor = _find_section_anchor(ws, r"EXPERIENCIA\s+ESPECIFICA", 1, ws.max_row)
    if not anchor:
        return {"items": [], "total_dias_calc": 0, "resumen": "", "_meta": {"anchor_row": None}}
    return _parse_experiencia_from_header(ws, anchor, debug=debug)


# ============================================================
# API principal (igual)
# ============================================================
def parse_eoi_excel(
    xlsx_path: Path,
    debug: bool = False,
    layout: Optional[Dict[str, Any]] = None
) -> Dict[str, Any]:
    xlsx_path = Path(xlsx_path)
    wb = load_workbook(xlsx_path, data_only=True)
    ws = find_best_sheet(wb)

    dp = parse_datos_personales(ws, debug=debug)
    fa = parse_formacion_obligatoria(ws, debug=debug)
    ec = parse_estudios_complementarios(ws, debug=debug)
    eg = parse_experiencia_general(ws, debug=debug)
    ee = parse_experiencia_especifica(ws, debug=debug)

    out = {
        **dp,
        "formacion_obligatoria": fa,
        "estudios_complementarios": ec,
        "exp_general": eg,
        "exp_especifica": ee,
        "exp_general_dias": int(eg.get("total_dias_calc") or 0),
        "exp_especifica_dias": int(ee.get("total_dias_calc") or 0),
        "source_file": str(xlsx_path),
    }

    out["exp_general_resumen_text"] = (eg.get("resumen") or "").strip()
    out["exp_especifica_resumen_text"] = (ee.get("resumen") or "").strip()

    def to_ymd(dias: int) -> str:
        if dias <= 0:
            return "0 año(s), 0 mes(es), 0 día(s)"
        anios = dias // 365
        rem = dias % 365
        meses = rem // 30
        dd = rem % 30
        return f"{anios} año(s), {meses} mes(es), {dd} día(s)"

    out["exp_general_total_text"] = to_ymd(out["exp_general_dias"])
    out["exp_especifica_total_text"] = to_ymd(out["exp_especifica_dias"])

    out["_fill_payload"] = {
        "dni": out.get("dni", ""),
        "nombre_full": out.get("nombre_full", ""),
        "email": out.get("email", ""),
        "celular": out.get("celular", ""),
        "formacion_obligatoria_resumen": out.get("formacion_obligatoria_resumen", ""),
        "estudios_complementarios_resumen": (ec.get("resumen") or ""),
        "exp_general_resumen_text": out.get("exp_general_resumen_text", ""),
        "exp_general_dias": out.get("exp_general_dias", 0),
        "exp_especifica_resumen_text": out.get("exp_especifica_resumen_text", ""),
        "exp_especifica_dias": out.get("exp_especifica_dias", 0),
    }

    if debug:
        print("[EOI_EXCEL] OK ->", out["_fill_payload"])

    return out
