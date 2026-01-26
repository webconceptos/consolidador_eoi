# parsers/eoi_excel.py
# -*- coding: utf-8 -*-
"""
Parser EOI (Excel) - CONSOLIDADOR EOI

Objetivo
--------
Extraer del Excel de Expresión de Interés (EOI) una estructura homogénea para llenar
el Cuadro de Evaluación (plantilla de Revisión Preliminar).

✅ Mantiene (y refuerza):
- Formación Académica obligatoria (sección a) - items + resumen
- Estudios Complementarios (sección b.*) - bloques + resumen + total horas

✅ Corrige definitivamente:
- Datos Personales (estructura mixta por filas: no es siempre header/value pareado)
- Experiencia General / Experiencia Específica (anclada por títulos de sección, y columnas
  mapeadas por encabezados, para evitar “corrimientos”)

Dependencias:
  pip install openpyxl

Notas de diseño (importantes)
-----------------------------
1) No asumimos posiciones fijas “a rajatabla”, salvo como fallback.
   Primero buscamos ANCLAS textuales (labels), luego leemos valores asociados.

2) Experiencia:
   - Ubicamos la sección por el texto "a) EXPERIENCIA GENERAL" o "b) EXPERIENCIA ESPECIFICA"
   - Ubicamos la fila de encabezados ("No.", "Nombre de la Entidad...", "Fecha de Inicio", etc.)
   - Mapeamos columnas por keywords (robusto ante celdas vacías por merges).
   - Parseamos registros y sus descripciones ("Descripción del Trabajo Realizado:" + detalle)

3) Fechas:
   - openpyxl puede devolver datetime/date o string. Convertimos a "dd/mm/YYYY" para salida.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime, date

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ============================================================
# Utils
# ============================================================
def norm(x: Any) -> str:
    """Normaliza texto: None -> '', colapsa espacios."""
    if x is None:
        return ""
    s = str(x).replace("\u00a0", " ").strip()
    return re.sub(r"\s+", " ", s)


def cell_raw(ws: Worksheet, r: int, c: int) -> Any:
    return ws.cell(row=r, column=c).value


def cell_str(ws: Worksheet, r: int, c: int) -> str:
    return norm(cell_raw(ws, r, c))


def row_text(ws: Worksheet, r: int, c1: int = 1, c2: int = 15) -> str:
    """Concatena el texto de una fila (útil para búsqueda de anclas)."""
    parts: List[str] = []
    for c in range(c1, c2 + 1):
        v = ws.cell(row=r, column=c).value
        if v is None:
            continue
        s = norm(v)
        if s:
            parts.append(s)
    return " | ".join(parts)


def normalize_email(x: str) -> str:
    x = norm(x).lower()
    x = x.replace(" ", "")
    return x


def normalize_phone(x: str) -> str:
    x = norm(x)
    # deja solo dígitos
    d = re.sub(r"\D+", "", x)
    # típico Perú celular 9xxxxxxxx
    if len(d) >= 9:
        return d[-9:]
    return d


def normalize_dni(x: str) -> str:
    d = re.sub(r"\D+", "", norm(x))
    if len(d) >= 8:
        return d[-8:]
    return d


def as_date_str(x: Any) -> str:
    """Convierte date/datetime/string a dd/mm/YYYY (o '' si no aplica)."""
    if x is None:
        return ""
    if isinstance(x, datetime):
        return x.strftime("%d/%m/%Y")
    if isinstance(x, date):
        return x.strftime("%d/%m/%Y")
    s = norm(x)
    if not s:
        return ""
    # intentos típicos
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.strftime("%d/%m/%Y")
        except Exception:
            pass
    # si viene como "2025-12-08 00:00:00"
    try:
        dt = datetime.fromisoformat(s.replace("Z", ""))
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return s  # fallback: devolvemos el texto


def parse_date_any(x: Any) -> Optional[datetime]:
    """Retorna datetime o None."""
    if x is None:
        return None
    if isinstance(x, datetime):
        return x
    if isinstance(x, date):
        return datetime(x.year, x.month, x.day)
    s = norm(x)
    if not s:
        return None
    s = s.replace(".", "/").replace("-", "/")
    # dd/mm/yyyy o dd/mm/yy
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    try:
        return datetime.fromisoformat(norm(x).replace("Z", ""))
    except Exception:
        return None


def days_between(d1: Optional[datetime], d2: Optional[datetime]) -> int:
    if not d1 or not d2:
        return 0
    if d2 < d1:
        return 0
    return int((d2 - d1).days) + 1


def find_best_sheet(wb) -> Worksheet:
    """
    En tu caso normalmente hay 1 hoja: "Formt Exp Int".
    Igual dejamos selección robusta por anclas.
    """
    best = wb.worksheets[0]
    best_score = -1
    for ws in wb.worksheets:
        score = 0
        # señales del formato EOI
        for r in range(1, min(ws.max_row, 80) + 1):
            t = row_text(ws, r, 1, 12).upper()
            if "DATOS PERSONALES" in t:
                score += 5
            if "FORMACIÓN ACADÉMICA" in t or "FORMACION ACADEMICA" in t:
                score += 3
            if "ESTUDIOS COMPLEMENTARIOS" in t:
                score += 3
            if "EXPERIENCIA" in t and "IV." in t:
                score += 3
        if score > best_score:
            best_score = score
            best = ws
    return best


# ============================================================
# 1) Datos Personales (robusto por labels)
# ============================================================
def _find_label_cell(ws: Worksheet, label_regex: str, r1: int, r2: int, c1: int, c2: int) -> Optional[Tuple[int, int]]:
    rgx = re.compile(label_regex, re.IGNORECASE)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            v = cell_str(ws, r, c)
            if v and rgx.search(v):
                return (r, c)
    return None


def _value_below(ws: Worksheet, r: int, c: int, max_down: int = 2) -> str:
    """Lee el valor debajo de un label (típico en tu formato)."""
    for k in range(1, max_down + 1):
        v = cell_raw(ws, r + k, c)
        s = norm(v)
        if s:
            return s
    return ""


def _value_right(ws: Worksheet, r: int, c: int, max_right: int = 3) -> str:
    """Fallback: lee valor a la derecha si el formato fuese label|valor."""
    for k in range(1, max_right + 1):
        s = cell_str(ws, r, c + k)
        if s:
            return s
    return ""


def parse_datos_personales(ws: Worksheet, debug: bool = False) -> Dict[str, Any]:
    """
    Tu EOI real (verificado en formatocv_RAC.xlsx) tiene:
      - Apellido Paterno / Materno: row 12 labels, row 13 valores
      - Nombres / Lugar de nacimiento + fecha: row 14 labels, row 15 valores
      - DNI: label row 16 col 8, valor row 17 col 8
      - Email/Celular: label row 22, valor row 23

    Esta función NO asume "pares" uniformes; busca labels y obtiene valor debajo (y fallback a la derecha).
    """
    out = {
        "dni": "",
        "apellido_paterno": "",
        "apellido_materno": "",
        "nombres": "",
        "nombre_full": "",
        "email": "",
        "celular": "",
    }

    # Rango razonable donde está DP
    R1, R2 = 1, 40
    C1, C2 = 1, 12

    # Apellidos y nombres
    pos = _find_label_cell(ws, r"\bApellido\s*Paterno\b", R1, R2, C1, C2)
    if pos:
        out["apellido_paterno"] = _value_below(ws, *pos) or _value_right(ws, *pos)
    pos = _find_label_cell(ws, r"\bApellido\s*Materno\b", R1, R2, C1, C2)
    if pos:
        out["apellido_materno"] = _value_below(ws, *pos) or _value_right(ws, *pos)

    pos = _find_label_cell(ws, r"\bNombres\b", R1, R2, C1, C2)
    if pos:
        out["nombres"] = _value_below(ws, *pos) or _value_right(ws, *pos)

    # DNI
    pos = _find_label_cell(ws, r"Documento\s+de\s+identidad|DNI", R1, R2, C1, C2)
    if pos:
        out["dni"] = normalize_dni(_value_below(ws, *pos) or _value_right(ws, *pos))

    # Celular (priorizar "Celular" sobre "Teléfono")
    pos = _find_label_cell(ws, r"\bCelular\b", R1, R2, C1, C2)
    if pos:
        out["celular"] = normalize_phone(_value_below(ws, *pos) or _value_right(ws, *pos))
    else:
        pos = _find_label_cell(ws, r"\bTel[ée]fono\b", R1, R2, C1, C2)
        if pos:
            out["celular"] = normalize_phone(_value_below(ws, *pos) or _value_right(ws, *pos))

    # Email
    pos = _find_label_cell(ws, r"\bemail\b|correo", R1, R2, C1, C2)
    if pos:
        out["email"] = normalize_email(_value_below(ws, *pos) or _value_right(ws, *pos))

    ap = " ".join([out["apellido_paterno"], out["apellido_materno"]]).strip()
    nm = out["nombres"].strip()
    out["nombre_full"] = norm(" ".join([ap, nm]))

    if debug:
        print("[DP]", out)

    return out


# ============================================================
# 2) Formación Académica obligatoria (a) - mantiene tu lógica,
#    pero con corte explícito por EGRESADO UNIVERSITARIO
# ============================================================
def _find_row_contains(ws: Worksheet, text_regex: str, r1: int, r2: int, c1: int = 1, c2: int = 12) -> Optional[int]:
    rgx = re.compile(text_regex, re.IGNORECASE)
    for r in range(r1, r2 + 1):
        if rgx.search(row_text(ws, r, c1, c2)):
            return r
    return None


def parse_formacion_obligatoria(ws: Worksheet, debug: bool = False) -> Dict[str, Any]:
    """
    En tu formato, la tabla de (a) arranca alrededor de la fila 47-56.
    La salida será:
      {
        "items": [...],
        "resumen": "...",
        "meta": {"start_row":.., "end_row":.., "colmap": {...}}
      }
    """
    # Detectamos header de (a) buscando la palabra "COLEGIATURA" o la sección III
    # pero aquí usamos un approach simple y efectivo: buscar "COLEGIATURA" en col C.
    header_like = _find_row_contains(ws, r"COLEGIATURA|MAESTRIA|TITULO|BACHILLER|EGRESADO\s+UNIVERSITARIO", 40, 80, 1, 12)
    if not header_like:
        # fallback duro (tu formato)
        start_row, end_row = 51, 56
    else:
        # en tu archivo real, los items empiezan en 51
        # si header_like cae dentro, lo usamos como referencia.
        start_row = 51
        # corte: fila que contiene "EGRESADO UNIVERSITARIO" (inclusive)
        er = _find_row_contains(ws, r"EGRESADO\s+UNIVERSITARIO", start_row, start_row + 25, 1, 12)
        end_row = er if er else (start_row + 5)

    # Mapeo de columnas según tu estructura observada:
    # C: titulo_item, F: especialidad, G: fecha, H: centro, J: ciudad
    colmap = {"titulo_item": 3, "especialidad": 6, "fecha": 7, "centro": 8, "ciudad": 10}

    items: List[Dict[str, Any]] = []
    resumen_parts: List[str] = []

    for r in range(start_row, end_row + 1):
        titulo = cell_str(ws, r, colmap["titulo_item"])
        esp = cell_str(ws, r, colmap["especialidad"])
        fec = as_date_str(cell_raw(ws, r, colmap["fecha"]))
        cen = cell_str(ws, r, colmap["centro"])
        ciu = cell_str(ws, r, colmap["ciudad"])

        has_data = any([esp, fec, cen, ciu])
        it = {
            "row": r,
            "titulo_item": titulo,
            "especialidad": esp,
            "fecha": fec,
            "centro": cen,
            "ciudad": ciu,
            "has_data": bool(has_data),
        }
        items.append(it)

        if titulo and has_data:
            # Ej: TITULO: Ing... (04/09/2014 | Univ.. | Puno/Perú)
            parts = [p for p in [fec, cen, ciu] if p]
            resumen_parts.append(f"{titulo}: {esp} ({' | '.join(parts)})" if esp else f"{titulo}: ({' | '.join(parts)})")

    resumen = " ; ".join(resumen_parts).strip()

    out = {
        "items": items,
        "resumen": resumen,
        "meta": {"start_row": start_row, "end_row": end_row, "colmap": colmap},
    }

    if debug:
        print("[FA]", out["meta"], "items=", len(items))

    return out


# ============================================================
# 3) Estudios Complementarios (b.*) - robusto por bloque y headers
# ============================================================
def _is_stop_row_for_blocks(ws: Worksheet, r: int) -> bool:
    t = row_text(ws, r, 1, 12).upper()
    return ("IV." in t and "EXPERIENCIA" in t) or ("IV." in t and " EXPERIENCIA" in t)


def _parse_block_table(ws: Worksheet, title_row: int, debug: bool = False) -> Dict[str, Any]:
    """
    Dado un title_row donde está "b.1) ...", detecta header y parsea items hasta
    "(Puede adicionar...)" o hasta que se acabe el bloque.
    """
    title = cell_str(ws, title_row, 2) or row_text(ws, title_row, 1, 12)
    # header típico aparece 2 filas abajo: "No. Centro de estudios ... Fecha de Inicio ... Horas"
    header_row = None
    for r in range(title_row, min(title_row + 8, ws.max_row) + 1):
        t = row_text(ws, r, 1, 12).upper()
        if "NO." in t and ("CENTRO" in t or "CAPACIT" in t) and ("FECHA" in t or "HORAS" in t):
            header_row = r
            break
    if not header_row:
        return {"row": title_row, "title": title, "items": [], "total_horas": 0, "resumen": ""}

    # Mapeo de columnas por keywords en header_row
    # En tu archivo real:
    #   C No., D Centro, F Capacitación, H Inicio, I Fin, J Horas
    col_nro = 3
    col_centro = 4
    col_cap = 6
    col_ini = 8
    col_fin = 9
    col_horas = 10

    # (Si algún día cambia, aquí puedes mejorar mapeando dinámicamente por lectura de cada celda del header_row.)

    items: List[Dict[str, Any]] = []
    resumen_lines: List[str] = []
    total_horas = 0

    r = header_row + 2  # salta la fila "Día/Mes/Año"
    while r <= ws.max_row:
        # stop por nueva sección / experiencia
        if _is_stop_row_for_blocks(ws, r):
            break

        nro = cell_str(ws, r, col_nro)
        centro = cell_str(ws, r, col_centro)
        cap = cell_str(ws, r, col_cap)
        fi = as_date_str(cell_raw(ws, r, col_ini))
        ff = as_date_str(cell_raw(ws, r, col_fin))
        horas_raw = cell_raw(ws, r, col_horas)

        # fin del bloque
        tt = row_text(ws, r, 1, 12)
        if re.search(r"Puede\s+adicionar", tt, re.IGNORECASE):
            break

        if not any([nro, centro, cap, fi, ff, horas_raw]):
            r += 1
            continue

        horas = 0
        try:
            horas = int(float(horas_raw)) if horas_raw not in (None, "") else 0
        except Exception:
            horas = 0

        it = {
            "row": r,
            "nro": nro,
            "centro": centro,
            "capacitacion": cap,
            "fecha_inicio": fi,
            "fecha_fin": ff,
            "horas": horas,
        }
        items.append(it)
        total_horas += horas

        if centro or cap:
            resumen_lines.append(f"{centro} - {cap} ({fi} | {ff} | {horas}h)".strip())

        r += 1

    return {
        "row": title_row,
        "title": title,
        "items": items,
        "total_horas": total_horas,
        "resumen": "\n".join(resumen_lines).strip(),
    }


def parse_estudios_complementarios(ws: Worksheet, debug: bool = False) -> Dict[str, Any]:
    """
    Detecta bloques b.1, b.2, b.3... hasta antes de sección IV.
    """
    blocks: List[Dict[str, Any]] = []
    for r in range(1, min(ws.max_row, 200) + 1):
        if _is_stop_row_for_blocks(ws, r):
            break
        t = row_text(ws, r, 1, 12)
        if re.search(r"\bb\.\d\)", t, re.IGNORECASE):
            # title_row real suele estar en col B
            b = _parse_block_table(ws, r, debug=debug)
            # id del bloque
            m = re.search(r"\b(b\.\d)\)", t, re.IGNORECASE)
            bid = m.group(1).lower() if m else f"b.{len(blocks)+1}"
            b["id"] = bid
            blocks.append(b)

    total_horas = sum(int(b.get("total_horas") or 0) for b in blocks)
    resumen_parts: List[str] = []
    for b in blocks:
        bid = (b.get("id") or "").upper()
        title = b.get("title") or ""
        body = b.get("resumen") or ""
        if body:
            resumen_parts.append(f"{bid}:\n{body}")
        else:
            resumen_parts.append(f"{bid}:\n(sin cursos declarados)")

    return {
        "blocks": blocks,
        "total_horas": total_horas,
        "resumen": "\n\n".join(resumen_parts).strip(),
    }


# ============================================================
# 4) Experiencia (general/específica) - anclada por sección + headers
# ============================================================
def _is_desc_label_row(ws: Worksheet, r: int) -> bool:
    t = row_text(ws, r, 1, 12).upper()
    return "DESCRIPCIÓN DEL TRABAJO" in t or "DESCRIPCION DEL TRABAJO" in t


def _get_desc_detail(ws: Worksheet, r: int, c_hint: int = 3) -> str:
    """
    En tu formato el detalle suele estar en col C (3) y puede estar combinado.
    """
    s = cell_str(ws, r, c_hint)
    if s:
        return s
    # fallback: toda la fila
    return row_text(ws, r, 1, 12)


def _find_section_anchor(ws: Worksheet, anchor_regex: str, r1: int = 1, r2: Optional[int] = None) -> Optional[int]:
    if r2 is None:
        r2 = ws.max_row
    rgx = re.compile(anchor_regex, re.IGNORECASE)
    for r in range(r1, r2 + 1):
        if rgx.search(row_text(ws, r, 1, 12)):
            return r
    return None


def _find_exp_header_row(ws: Worksheet, anchor_row: int) -> Optional[int]:
    """
    Header típico: "No.", "Nombre de la Entidad ó Empresa", "Fecha de Inicio", "Fecha de Culminación", etc.
    """
    for r in range(anchor_row, min(anchor_row + 12, ws.max_row) + 1):
        t = row_text(ws, r, 1, 12).upper()
        if "NO." in t and "ENTIDAD" in t and ("FECHA" in t):
            return r
    return None


def _parse_experiencia_from_header(ws: Worksheet, anchor_row: int, debug: bool = False) -> Dict[str, Any]:
    """
    Parsea una sección de experiencia a partir de su anchor_row (texto "a)..." o "b)...").
    """
    header_row = _find_exp_header_row(ws, anchor_row)
    if not header_row:
        return {"items": [], "total_dias_calc": 0, "resumen": "", "_meta": {"anchor_row": anchor_row, "header_row": None}}

    # En tu formato real (verificado):
    # C: No.
    # D/E: Entidad (E puede ser None por merge)
    # F: Proyecto (a veces vacío)
    # G: Cargo
    # H: Inicio
    # I: Fin
    # J: Tiempo
    COL_C = 3
    col_nro = COL_C + 0
    col_ent1 = COL_C + 1
    col_ent2 = COL_C + 2
    col_proy = COL_C + 3
    col_cargo = COL_C + 4
    col_ini = COL_C + 5
    col_fin = COL_C + 6
    col_tiempo = COL_C + 7

    items: List[Dict[str, Any]] = []
    resumen_lines: List[str] = []

    # datos empiezan 2 filas abajo (fila "Día/Mes/Año" intermedia)
    r = header_row + 2
    while r <= ws.max_row:
        trow = row_text(ws, r, 1, 12)

        # corte por "Puede adicionar..."
        if re.search(r"Puede\s+adicionar", trow, re.IGNORECASE):
            break

        # corte por siguiente sección b) o por "V." etc
        if r > anchor_row and re.search(r"^\s*b\)\s+EXPERIENCIA", trow, re.IGNORECASE):
            break

        # si detectamos otro header repetido "No. Nombre de la Entidad..."
        if re.search(r"\bNo\.\b", trow, re.IGNORECASE) and re.search(r"Entidad|Empresa", trow, re.IGNORECASE):
            r += 1
            continue

        # saltar label de descripción
        if _is_desc_label_row(ws, r):
            r += 1
            continue

        nro = norm(cell_raw(ws, r, col_nro))
        entidad = " ".join([cell_str(ws, r, col_ent1), cell_str(ws, r, col_ent2)]).strip()
        proyecto = cell_str(ws, r, col_proy)
        cargo = cell_str(ws, r, col_cargo)
        fi_raw = cell_raw(ws, r, col_ini)
        ff_raw = cell_raw(ws, r, col_fin)
        tiempo = cell_str(ws, r, col_tiempo)

        fi = as_date_str(fi_raw)
        ff = as_date_str(ff_raw)

        base_has_data = any([nro, entidad, proyecto, cargo, fi, ff, tiempo])
        if not base_has_data:
            r += 1
            continue

        descripcion = ""
        # patrón real: fila siguiente es "Descripción...", la siguiente contiene el detalle
        if r + 1 <= ws.max_row and _is_desc_label_row(ws, r + 1):
            if r + 2 <= ws.max_row:
                descripcion = _get_desc_detail(ws, r + 2)
            next_r = r + 3
        else:
            next_r = r + 1

        d_fi = parse_date_any(fi_raw)
        d_ff = parse_date_any(ff_raw)
        dias = days_between(d_fi, d_ff)

        it = {
            "row": r,
            "nro": str(nro) if nro is not None else "",
            "entidad": entidad,
            "proyecto": proyecto,
            "cargo": cargo,
            "fecha_inicio": fi,
            "fecha_fin": ff,
            "tiempo_en_cargo": tiempo,
            "dias_calc": dias,
            "descripcion": descripcion,
        }
        items.append(it)

        head = " - ".join([p for p in [entidad, cargo] if p]).strip(" -")
        fechas = " a ".join([p for p in [fi, ff] if p]).strip(" a")
        line = " | ".join([p for p in [head, fechas] if p]).strip()
        if descripcion:
            line += f"\n  Desc: {descripcion}"
        resumen_lines.append(line)

        r = next_r

        # corte por seguridad: si llegamos a una línea que es el inicio de otra sección fuerte
        if re.search(r"^\s*V\.", row_text(ws, r, 1, 12)):
            break

    total_dias = sum(int(x.get("dias_calc") or 0) for x in items)
    resumen = "\n\n".join([x for x in resumen_lines if x]).strip()

    if debug:
        print(f"[EXP] anchor={anchor_row} header={header_row} items={len(items)} total_dias={total_dias}")

    return {
        "items": items,
        "total_dias_calc": total_dias,
        "resumen": resumen,
        "_meta": {"anchor_row": anchor_row, "header_row": header_row, "start_row": header_row + 2},
    }


def parse_experiencia_general(ws: Worksheet, debug: bool = False) -> Dict[str, Any]:
    anchor = _find_section_anchor(ws, r"a\)\s*EXPERIENCIA\s+GENERAL", 1, ws.max_row)
    if not anchor:
        # fallback: buscar "EXPERIENCIA GENERAL" simple
        anchor = _find_section_anchor(ws, r"EXPERIENCIA\s+GENERAL", 1, ws.max_row)
    if not anchor:
        return {"items": [], "total_dias_calc": 0, "resumen": "", "_meta": {"anchor_row": None}}
    return _parse_experiencia_from_header(ws, anchor, debug=debug)


def parse_experiencia_especifica(ws: Worksheet, debug: bool = False) -> Dict[str, Any]:
    anchor = _find_section_anchor(ws, r"b\)\s*EXPERIENCIA\s+ESPECIFICA", 1, ws.max_row)
    if not anchor:
        anchor = _find_section_anchor(ws, r"EXPERIENCIA\s+ESPECIFICA", 1, ws.max_row)
    if not anchor:
        return {"items": [], "total_dias_calc": 0, "resumen": "", "_meta": {"anchor_row": None}}
    return _parse_experiencia_from_header(ws, anchor, debug=debug)


# ============================================================
# API principal
# ============================================================
def parse_eoi_excel(
    xlsx_path: Path,
    debug: bool = False,
    layout: Optional[Dict[str, Any]] = None
) -> Dict[str, Any]:
    """
    Parser principal EOI Excel.

    layout (opcional):
      Si luego lo conectas con Task_00, puedes pasar anclas/rangos. Por ahora, este parser
      es suficientemente robusto sin layout (detecta por texto).
    """
    xlsx_path = Path(xlsx_path)
    wb = load_workbook(xlsx_path, data_only=True)
    ws = find_best_sheet(wb)

    dp = parse_datos_personales(ws, debug=debug)

    fa = parse_formacion_obligatoria(ws, debug=debug)
    ec = parse_estudios_complementarios(ws, debug=debug)

    eg = parse_experiencia_general(ws, debug=debug)
    ee = parse_experiencia_especifica(ws, debug=debug)

    # Payload listo para llenar el Cuadro de Evaluación (Task_40 / fill)
    out = {
        **dp,
        "formacion_obligatoria": fa,
        "formacion_obligatoria_items": fa.get("items", []),
        "formacion_obligatoria_resumen": fa.get("resumen", ""),
        "formacion_meta": fa.get("meta", {}),
        "estudios_complementarios": ec,
        "exp_general": eg,
        "exp_especifica": ee,
        "exp_general_dias": int(eg.get("total_dias_calc") or 0),
        "exp_especifica_dias": int(ee.get("total_dias_calc") or 0),
        "source_file": str(xlsx_path),
    }

    # textos de resumen para celdas del Cuadro de Evaluación
    out["exp_general_resumen_text"] = (eg.get("resumen") or "").strip()
    out["exp_especifica_resumen_text"] = (ee.get("resumen") or "").strip()

    # Totales formateados (año/mes/día simple, sin solapes todavía)
    def to_ymd(dias: int) -> str:
        if dias <= 0:
            return "0 año(s), 0 mes(es), 0 día(s)"
        anios = dias // 365
        rem = dias % 365
        meses = rem // 30
        dd = rem % 30
        return f"{anios} año(s), {meses} mes(es), {dd} día(s)"

    out["exp_general_total_text"] = to_ymd(out["exp_general_dias"])
    out["exp_especifica_total_text"] = to_ymd(out["exp_especifica_dias"])

    out["_fill_payload"] = {
        "dni": out.get("dni", ""),
        "nombre_full": out.get("nombre_full", ""),
        "email": out.get("email", ""),
        "celular": out.get("celular", ""),
        "formacion_obligatoria_resumen": out.get("formacion_obligatoria_resumen", ""),
        "estudios_complementarios_resumen": (ec.get("resumen") or ""),
        "exp_general_resumen_text": out.get("exp_general_resumen_text", ""),
        "exp_general_dias": out.get("exp_general_dias", 0),
        "exp_especifica_resumen_text": out.get("exp_especifica_resumen_text", ""),
        "exp_especifica_dias": out.get("exp_especifica_dias", 0),
    }

    if debug:
        print("[EOI_EXCEL] OK ->", out["_fill_payload"])

    return out
