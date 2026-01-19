# run.py
import re
import json
import csv
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

from parsers.eoi_excel import parse_eoi_excel
from parsers.eoi_pdf import parse_eoi_pdf
from utils.experience import ymd_from_days

from datetime import datetime

def ts():
    return datetime.now().isoformat(timespec="seconds")

def log_file_append(path, msg: str):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        f.write(f"[{ts()}] {msg}\n")


# -------------------------
# Helpers
# -------------------------
def norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def safe_filename(s: str, max_len: int = 90) -> str:
    s = norm(s)
    s = re.sub(r"[^A-Za-z0-9._-]+", "_", s)
    return s[:max_len] if len(s) > max_len else s


def slot_columns(slot_index: int):
    # slot 0 -> F/G, slot 1 -> H/I, etc.
    base_col = 6 + slot_index * 2
    score_col = base_col + 1
    return base_col, score_col


def next_free_slot(ws, max_slots: int = 8):
    # revisa fila 3 en columnas F, H, J...
    for i in range(max_slots):
        base_col, _ = slot_columns(i)
        v = ws.cell(row=3, column=base_col).value
        if v is None or str(v).strip() == "" or "NOMBRE DEL CONSULTOR" in str(v):
            return i
    return None


def ensure_dir(p: Path):
    p.mkdir(parents=True, exist_ok=True)


# -------------------------
# Estructura de carpetas (tu estándar)
# -------------------------
IN_FOLDER_NAME = "009. EDI RECIBIDA"
OUT_FOLDER_NAME = "011. INSTALACIÓN DE COMITÉ"


def infer_proceso_from_path(root: Path, file_path: Path) -> str:
    """
    root: .../Procesos de Selección
    file_path: .../Procesos de Selección/<Proceso>/009. EDI RECIBIDA/<Postulante>/archivo
    """
    rel = file_path.relative_to(root)
    return rel.parts[0]  # <Proceso>


def get_in_out_dirs(root: Path, proceso: str):
    in_dir = root / proceso / IN_FOLDER_NAME
    out_dir = root / proceso / OUT_FOLDER_NAME
    procesados_dir = out_dir / "procesados"
    procesados_dir.mkdir(parents=True, exist_ok=True)

    return in_dir, out_dir , procesados_dir


def find_eoi_files_old(process_root: Path):
    """
    Recorre SOLO dentro de .../<Proceso>/009. EDI RECIBIDA/
    y devuelve archivos .xlsx/.xlsm/.xls/.pdf
    """
    if not process_root.exists():
        return

    for p in process_root.rglob("*"):
        if not p.is_file():
            continue
        if p.name.startswith("~$"):
            continue
        if p.suffix.lower() in (".xlsx", ".xlsm", ".xls", ".pdf"):
            yield p


def find_eoi_files(in_dir: Path):
    """
    Devuelve UN archivo por carpeta de postulante.
    Regla:
      - Si hay Excel, usa Excel (prioridad: xlsx > xlsm > xls)
      - Si NO hay Excel, usa PDF
    """
    if not in_dir.exists():
        return []

    # agrupamos por carpeta "postulante"
    by_folder = {}
    for p in in_dir.rglob("*"):
        if not p.is_file():
            continue
        if p.name.startswith("~$"):
            continue
        ext = p.suffix.lower()
        if ext not in (".xlsx", ".xlsm", ".xls", ".pdf"):
            continue
        by_folder.setdefault(p.parent, []).append(p)

    chosen = []
    for folder, files in by_folder.items():
        excels = [f for f in files if f.suffix.lower() in (".xlsx", ".xlsm", ".xls")]
        pdfs   = [f for f in files if f.suffix.lower() == ".pdf"]

        def score_excel(f: Path):
            name = f.name.lower()
            ext = f.suffix.lower()
            # prioridad por extensión
            ext_score = {".xlsx": 3, ".xlsm": 2, ".xls": 1}.get(ext, 0)
            # bonus si parece ser el formato CV
            name_bonus = 2 if any(k in name for k in ("formatocv", "formato", "cv", "edi")) else 0
            return (ext_score, name_bonus, -len(name))  # estable

        if excels:
            excels.sort(key=score_excel, reverse=True)
            chosen.append(excels[0])
        elif pdfs:
            # si hay varios pdfs, toma el primero por nombre
            pdfs.sort(key=lambda x: x.name.lower())
            chosen.append(pdfs[0])
        # si no hay nada, no agrega

    # orden consistente
    chosen.sort(key=lambda x: str(x).lower())
    return chosen


# -------------------------
# Exporta calculadora por postulante
# -------------------------
def export_calculadora(calc_template: Path, out_dir: Path, data: dict):
    """
    Genera un archivo por postulante basado en tu 'CALCULADORA DE EXPERIENCIA',
    llenando únicamente el área de entrada (Desde/Hasta) con los periodos detectados.
    """
    if not calc_template.exists():
        return

    dni = norm(data.get("dni", "")) or "SIN_DNI"
    nombre = norm(data.get("nombre_full", "")) or "SIN_NOMBRE"
    out_path = out_dir / f"Calculadora_{safe_filename(dni + '_' + nombre)}.xlsx"

    wb = load_workbook(calc_template)
    ws_g = wb["Exp General"]
    ws_e = wb["Exp específica"]

    # Área de entrada (según tu plantilla): columnas B/C desde fila 7
    start_row = 7
    max_rows = 60

    # Limpia entradas previas
    for r in range(start_row, start_row + max_rows):
        ws_g.cell(r, 2).value = None
        ws_g.cell(r, 3).value = None
        ws_e.cell(r, 2).value = None
        ws_e.cell(r, 3).value = None

    exps = data.get("experiencias", [])

    # General: todos los intervalos válidos
    intervals_g = [(rec.get("fi"), rec.get("ff")) for rec in exps if rec.get("fi") and rec.get("ff")]
    for i, (fi, ff) in enumerate(intervals_g[:max_rows]):
        ws_g.cell(start_row + i, 2).value = fi
        ws_g.cell(start_row + i, 3).value = ff

    # Específica: heurística (ajústala si tu comité tiene regla exacta)
    intervals_e = []
    for rec in exps:
        txt = f"{rec.get('cargo','')} {rec.get('proyecto','')}".upper()
        if any(k in txt for k in ("DESARROL", "PROGRAM", "ANALISTA", "SISTEM", "SOFTWARE", "JAVA", "ORACLE")):
            if rec.get("fi") and rec.get("ff"):
                intervals_e.append((rec["fi"], rec["ff"]))

    for i, (fi, ff) in enumerate(intervals_e[:max_rows]):
        ws_e.cell(start_row + i, 2).value = fi
        ws_e.cell(start_row + i, 3).value = ff

    wb.save(out_path)


# -------------------------
# Escribe un postulante en el Cuadro (bloques F/G, H/I, ...)
# -------------------------
def write_postulante(ws, data: dict, cfg: dict):
    max_slots = int(cfg.get("max_postulantes_por_cuadro", 8))
    slot = next_free_slot(ws, max_slots=max_slots)
    if slot is None:
        return False, "NO_HAY_SLOT"

    base_col, score_col = slot_columns(slot)

    # Cabecera (fila 3)
    ws.cell(row=3, column=base_col).value = norm(data.get("nombre_full", ""))

    # FORMACION (fila 6): detalle en base_col, cumple/no en score_col
    crit = cfg["criterios"]["formacion"]
    r = int(crit["row"])
    detalle = crit["detalle_template"].format(
        titulo=norm(data.get("titulo", "")),
        bachiller=norm(data.get("bachiller", "")),
        egresado=norm(data.get("egresado", "")),
    )
    ws.cell(row=r, column=base_col).value = detalle
    cumple = "CUMPLE" if any(norm(data.get(k, "")).upper() for k in ("titulo", "bachiller", "egresado")) else "NO CUMPLE"
    ws.cell(row=r, column=score_col).value = cumple

    # CURSOS: filas (por config). Detalle en base_col y checks en score_col
    cursos_txt = " | ".join([norm(x) for x in data.get("cursos", []) if norm(x)])[:1000]
    for rr in cfg["criterios"]["cursos"]["rows"]:
        ws.cell(row=int(rr), column=base_col).value = cursos_txt

    c_up = (cursos_txt or "").upper()
    scrum_ok = any(k in c_up for k in cfg["criterios"]["cursos"]["scrum_keywords"])
    ti_ok = any(k in c_up for k in cfg["criterios"]["cursos"]["ti_keywords"])
    ws.cell(row=9, column=score_col).value = "CUMPLE" if scrum_ok else "NO CUMPLE"
    ws.cell(row=10, column=score_col).value = "CUMPLE" if ti_ok else "NO CUMPLE"
    ws.cell(row=11, column=score_col).value = 1 if len(cursos_txt) > 0 else 0

    # EXPERIENCIA GENERAL EFECTIVA (sin superposición)
    gen_days = int(data.get("exp_general_dias", 0) or 0)
    y, m, d = ymd_from_days(gen_days)
    ws.cell(row=13, column=base_col).value = f"Experiencia general efectiva: {gen_days} días ({y}a {m}m {d}d)"
    umbral = float(cfg["criterios"]["exp_general"]["umbral_anios_1"])
    ws.cell(row=13, column=score_col).value = "CUMPLE" if (gen_days / 365.25) >= umbral else "NO CUMPLE"

    # EXPERIENCIA ESPECIFICA EFECTIVA (sin superposición)
    spec_days = int(data.get("exp_especifica_dias", 0) or 0)
    y2, m2, d2 = ymd_from_days(spec_days)
    ws.cell(row=17, column=base_col).value = f"Experiencia específica efectiva: {spec_days} días ({y2}a {m2}m {d2}d)"
    umbral_e = float(cfg["criterios"]["exp_especifica"]["umbral_anios_1"])
    ws.cell(row=17, column=score_col).value = "CUMPLE" if (spec_days / 365.25) >= umbral_e else "NO CUMPLE"

    # criterios deseables (java/oracle)
    ws.cell(row=20, column=score_col).value = 1 if data.get("java_ok", False) else 0
    ws.cell(row=21, column=score_col).value = 1 if data.get("oracle_ok", False) else 0

    # Pie informativo
    ws.cell(row=22, column=base_col).value = f"DNI: {norm(data.get('dni',''))} | Email: {norm(data.get('email',''))} | Cel: {norm(data.get('celular',''))}"

    return True, f"SLOT_{slot}"


# -------------------------
# CSV helpers
# -------------------------
def save_csv(path: Path, rows: list, header: list):
    ensure_dir(path.parent)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)


# -------------------------
# Main: un Cuadro por PROCESO, con “rotación” cuando se llenan los 8 slots
# -------------------------
def main():
    cfg = json.loads(Path("configs/config.json").read_text(encoding="utf-8"))
    
    root = Path(cfg["input_root"])
    print(root)
    if str(root) == "D:\\Acciones_Inversion\\1.8.2-SINF\\ProcesoSelección\\":
        raise SystemExit("Edita configs/config.json y coloca tu ruta en input_root (carpeta 'Procesos de Selección').")
    if not root.exists():
        raise SystemExit(f"No existe input_root: {root}")

    template_path = Path(cfg["output_template"])
    if not template_path.exists():
        raise SystemExit(f"No encuentro la plantilla de salida: {template_path}")

    calc_template = Path(cfg.get("calc_template", "CALCULADORA_DE_EXPERIENCIA.xlsx"))
    export_calc = bool(cfg.get("export_calculadora", True))
    max_slots = int(cfg.get("max_postulantes_por_cuadro", 8))

    # Detecta procesos: subcarpetas directas dentro de "Procesos de Selección"
    procesos = [p.name for p in root.iterdir() if p.is_dir()]
    procesos.sort()

    for proceso in procesos:
        in_dir, out_dir, procesados_dir = get_in_out_dirs(root, proceso)
        debug_log = out_dir / "debug_run.log"
        
        if not in_dir.exists():
            # no hay recepción; saltar
            continue

        ensure_dir(out_dir)

        # Archivos del proceso
        files = list(find_eoi_files(in_dir))

        log_file_append(debug_log, f"[DEBUG] archivos elegidos (1 por postulante): {len(files)}")
        for f in files[:50]:
            log_file_append(debug_log, f"   CHOSEN: {f}")



        if not files:
            continue

        # Logs por proceso
        log_rows = []
        flat_rows = []
        flat_header = [
            "proceso", "archivo", "tipo", "dni", "nombre", "email", "celular",
            "exp_general_dias", "exp_especifica_dias", "java_ok", "oracle_ok",
            "cursos_count", "cuadro", "slot"
        ]

        # Inicia Cuadro_01
        cuadro_idx = 1
        wb_out = load_workbook(template_path)
        ws_out = wb_out["Evaluación CV"]

        def save_current_cuadro():
            nonlocal cuadro_idx, wb_out
            out_xlsx = out_dir / f"Cuadro_Evaluacion_{safe_filename(proceso)}_{cuadro_idx:02d}.xlsx"
            wb_out.save(out_xlsx)

        for fp in files:
            try:
                if fp.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
                    data = parse_eoi_excel(fp)
                    ftype = "EXCEL"
                else:
                    data = parse_eoi_pdf(fp, use_ocr=cfg.get("pdf", {}).get("use_ocr", False))
                    ftype = "PDF"

                data["proceso"] = proceso

                ok, where = write_postulante(ws_out, data, cfg)

                # Si se llenó el cuadro, guardamos y abrimos uno nuevo y reintentamos
                if not ok and where == "NO_HAY_SLOT":
                    save_current_cuadro()
                    cuadro_idx += 1
                    wb_out = load_workbook(template_path)
                    ws_out = wb_out["Evaluación CV"]

                    ok, where = write_postulante(ws_out, data, cfg)

                if not ok:
                    log_rows.append([datetime.now().isoformat(timespec="seconds"), str(fp), ftype, "ERROR", where])
                    continue

                if export_calc:
                    export_calculadora(calc_template,procesados_dir, data)
                    #export_calculadora(calc_template, out_dir, data)

                flat_rows.append([
                    proceso, str(fp), ftype,
                    data.get("dni", ""), data.get("nombre_full", ""),
                    data.get("email", ""), data.get("celular", ""),
                    int(data.get("exp_general_dias", 0) or 0),
                    int(data.get("exp_especifica_dias", 0) or 0),
                    int(bool(data.get("java_ok", False))),
                    int(bool(data.get("oracle_ok", False))),
                    len(data.get("cursos", [])),
                    f"{cuadro_idx:02d}",
                    where
                ])

                log_rows.append([datetime.now().isoformat(timespec="seconds"), str(fp), ftype, "OK", f"CUADRO_{cuadro_idx:02d}_{where}"])

            except Exception as e:
                log_rows.append([datetime.now().isoformat(timespec="seconds"), str(fp), fp.suffix.lower(), "ERROR", repr(e)])

        # guarda el último cuadro
        save_current_cuadro()

        # guarda CSVs del proceso
        save_csv(out_dir / "consolidado.csv", flat_rows, flat_header)
        save_csv(out_dir / "log.csv", log_rows, ["fecha", "archivo", "tipo", "estado", "detalle"])

        print(f"[OK] Proceso: {proceso} -> salida en: {out_dir}")

    print("Listo. Revisa cada '011. INSTALACIÓN DE COMITÉ' por proceso.")


if __name__ == "__main__":
    main()
