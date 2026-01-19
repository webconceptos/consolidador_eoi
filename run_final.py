# run.py
import re
import json
import csv
import shutil
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

from parsers.eoi_excel import parse_eoi_excel
from parsers.eoi_pdf import parse_eoi_pdf
from utils.experience import ymd_from_days

# === NOMBRES EXACTOS DE TUS CARPETAS ===
IN_FOLDER_NAME = "009. EDI RECIBIDA"
OUT_FOLDER_NAME = "011. INSTALACIÓN DE COMITÉ"  # <- debe coincidir EXACTO con tu carpeta real
PROCESS_OUTPUT_TEMPLATE_NAME = "Formato_Salida_Expresion_Interes.xlsx"


# Dentro de OUT_FOLDER_NAME:
PROCESADOS_SUBFOLDER = "procesados"
CALCULADORAS_SUBFOLDER = "calculadoras"

# Hoja plantilla que se replica cuando se llenan los slots
EVAL_SHEET_TEMPLATE = "Evaluación CV"


# -------------------------
# Logging (a archivo)
# -------------------------
def ts():
    return datetime.now().isoformat(timespec="seconds")


def log_file_append(path: Path, msg: str):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        f.write(f"[{ts()}] {msg}\n")


# -------------------------
# Helpers
# -------------------------
def norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def safe_filename(s: str, max_len: int = 120) -> str:
    s = norm(s)
    s = re.sub(r"[^A-Za-z0-9._-]+", "_", s)
    return s[:max_len] if len(s) > max_len else s


def ensure_dir(p: Path):
    p.mkdir(parents=True, exist_ok=True)


def slot_columns(slot_index: int):
    # slot 0 -> F/G, slot 1 -> H/I, etc.
    base_col = 6 + slot_index * 2
    score_col = base_col + 1
    return base_col, score_col


def next_free_slot(ws, max_slots: int):
    # revisa fila 3 en columnas F, H, J...
    for i in range(max_slots):
        base_col, _ = slot_columns(i)
        v = ws.cell(row=3, column=base_col).value
        if v is None or str(v).strip() == "" or "NOMBRE DEL CONSULTOR" in str(v).upper():
            return i
    return None


# -------------------------
# Estructura por proceso
# -------------------------
def get_in_out_dirs(root: Path, proceso: str):
    in_dir = root / proceso / IN_FOLDER_NAME
    out_dir = root / proceso / OUT_FOLDER_NAME
    return in_dir, out_dir

def resolve_template_for_proceso(root: Path, proceso: str, cfg_global: dict) -> Path:
    """
    Busca plantilla por proceso dentro de:
      <proceso>/011. INSTALACIÓN DE COMITÉ/Formato_Salida_Expresion_Interes.xlsx
    Si no existe, usa cfg_global["output_template"].
    """
    _, out_dir = get_in_out_dirs(root, proceso)
    tpl_local = out_dir / PROCESS_OUTPUT_TEMPLATE_NAME
    if tpl_local.exists():
        return tpl_local

    return Path(cfg_global["output_template"])


def choose_one_file_per_postulante_folder(in_dir: Path):
    """
    Devuelve UN archivo por carpeta de postulante.
    Regla:
      - Si hay Excel -> usa Excel
      - Si NO hay Excel -> usa PDF, PERO ignora PDFs tipo 'correo de presentación'
    """
    if not in_dir.exists():
        return []

    by_folder = {}
    for p in in_dir.rglob("*"):
        if not p.is_file():
            continue
        if p.name.startswith("~$"):
            continue
        ext = p.suffix.lower()
        if ext not in (".xlsx", ".xlsm", ".xls", ".pdf"):
            continue
        by_folder.setdefault(p.parent, []).append(p)

    def score_excel(f: Path):
        name = f.name.lower()
        ext = f.suffix.lower()
        ext_score = {".xlsx": 30, ".xlsm": 20, ".xls": 10}.get(ext, 0)
        bonus = 0
        if any(k in name for k in ("formatocv", "formato", "cv", "edi")):
            bonus += 8
        if any(k in name for k in ("plantilla", "template", "blank")):
            bonus -= 8
        return ext_score + bonus

    def is_bad_pdf_name(name: str) -> bool:
        name = name.lower()
        return any(k in name for k in ("correo", "presentacion", "presentación", "mail", "mensaje"))

    def score_pdf(f: Path):
        name = f.name.lower()
        bonus = 0
        if any(k in name for k in ("formatocv", "cv", "expresion", "expresión", "edi")):
            bonus += 10
        if is_bad_pdf_name(name):
            bonus -= 50
        return bonus

    chosen = []
    skipped_folders = []

    for folder, files in by_folder.items():
        excels = [f for f in files if f.suffix.lower() in (".xlsx", ".xlsm", ".xls")]
        pdfs = [f for f in files if f.suffix.lower() == ".pdf"]

        if excels:
            excels.sort(key=score_excel, reverse=True)
            chosen.append(excels[0])
            continue

        if pdfs:
            pdfs.sort(key=score_pdf, reverse=True)
            best = pdfs[0]
            # si el "mejor" sigue siendo correo, no lo procesamos
            if is_bad_pdf_name(best.name):
                skipped_folders.append(folder)
                continue
            chosen.append(best)
        else:
            skipped_folders.append(folder)

    # (opcional) si quieres loguear carpetas omitidas, puedes retornarlas también
    chosen.sort(key=lambda x: str(x).lower())
    return chosen



# -------------------------
# Exporta calculadora por postulante
# -------------------------
def export_calculadora(calc_template: Path, out_calc_dir: Path, data: dict, debug_log: Path):
    if not calc_template.exists():
        log_file_append(debug_log, f"[WARN] No existe calc_template: {calc_template}")
        return

    dni = norm(data.get("dni", "")) or "SIN_DNI"
    nombre = norm(data.get("nombre_full", "")) or "SIN_NOMBRE"
    out_path = out_calc_dir / f"Calculadora_{safe_filename(dni + '_' + nombre)}.xlsx"

    wb = load_workbook(calc_template)
    ws_g = wb["Exp General"]
    ws_e = wb["Exp específica"]

    # Área de entrada (según tu plantilla): columnas B/C desde fila 7
    start_row = 7
    max_rows = 60

    # Limpia entradas previas
    for r in range(start_row, start_row + max_rows):
        ws_g.cell(r, 2).value = None
        ws_g.cell(r, 3).value = None
        ws_e.cell(r, 2).value = None
        ws_e.cell(r, 3).value = None

    exps = data.get("experiencias", [])

    # General: todos los intervalos válidos
    intervals_g = [(rec.get("fi"), rec.get("ff")) for rec in exps if rec.get("fi") and rec.get("ff")]
    for i, (fi, ff) in enumerate(intervals_g[:max_rows]):
        ws_g.cell(start_row + i, 2).value = fi
        ws_g.cell(start_row + i, 3).value = ff

    # Específica: heurística (puedes afinar luego con regla exacta de tu comité)
    intervals_e = []
    for rec in exps:
        txt = f"{rec.get('cargo','')} {rec.get('proyecto','')}".upper()
        if any(k in txt for k in ("DESARROL", "PROGRAM", "ANALISTA", "SISTEM", "SOFTWARE", "JAVA", "ORACLE")):
            if rec.get("fi") and rec.get("ff"):
                intervals_e.append((rec["fi"], rec["ff"]))

    for i, (fi, ff) in enumerate(intervals_e[:max_rows]):
        ws_e.cell(start_row + i, 2).value = fi
        ws_e.cell(start_row + i, 3).value = ff

    wb.save(out_path)
    log_file_append(debug_log, f"[OK] Calculadora generada: {out_path.name}")


# -------------------------
# Escribe un postulante en el Cuadro (bloques F/G, H/I, ...)
# -------------------------
def write_postulante(ws, data: dict, cfg: dict):
    max_slots = int(cfg.get("max_postulantes_por_hoja", 8))
    slot = next_free_slot(ws, max_slots=max_slots)
    if slot is None:
        return False, "NO_HAY_SLOT"

    base_col, score_col = slot_columns(slot)

    # 🔒 PROTECCIÓN ANTI-SOBRESCRITURA (AQUÍ VA)
    prev = ws.cell(row=3, column=base_col).value
    if prev and str(prev).strip():
        return False, f"SLOT_OCUPADO_{slot}"
    
    # Cabecera (fila 3)
    #ws.cell(row=3, column=base_col).value = norm(data.get("nombre_full", ""))
    display_name = norm(data.get("nombre_full", ""))

    # Fallbacks: DNI -> nombre de carpeta -> nombre de archivo
    if not display_name:
        dni = norm(data.get("dni", ""))
        src = norm(data.get("source_label", ""))  # lo ponemos desde el loop
        if dni:
            display_name = f"SIN_NOMBRE_{dni}"
        elif src:
            display_name = f"SIN_NOMBRE_{src}"
        else:
            display_name = "SIN_NOMBRE"

    ws.cell(row=3, column=base_col).value = display_name


    # FORMACION (fila 6): detalle en base_col, cumple/no en score_col
    crit = cfg["criterios"]["formacion"]
    r = int(crit["row"])
    detalle = crit["detalle_template"].format(
        titulo=norm(data.get("titulo", "")),
        bachiller=norm(data.get("bachiller", "")),
        egresado=norm(data.get("egresado", "")),
    )
    ws.cell(row=r, column=base_col).value = detalle
    cumple = "CUMPLE" if any(norm(data.get(k, "")).upper() for k in ("titulo", "bachiller", "egresado")) else "NO CUMPLE"
    ws.cell(row=r, column=score_col).value = cumple

    # CURSOS
    cursos_txt = " | ".join([norm(x) for x in data.get("cursos", []) if norm(x)])[:1000]
    for rr in cfg["criterios"]["cursos"]["rows"]:
        ws.cell(row=int(rr), column=base_col).value = cursos_txt

    c_up = (cursos_txt or "").upper()
    scrum_ok = any(k in c_up for k in cfg["criterios"]["cursos"]["scrum_keywords"])
    ti_ok = any(k in c_up for k in cfg["criterios"]["cursos"]["ti_keywords"])
    ws.cell(row=9, column=score_col).value = "CUMPLE" if scrum_ok else "NO CUMPLE"
    ws.cell(row=10, column=score_col).value = "CUMPLE" if ti_ok else "NO CUMPLE"
    ws.cell(row=11, column=score_col).value = 1 if len(cursos_txt) > 0 else 0

    # EXPERIENCIA GENERAL EFECTIVA (sin superposición)
    gen_days = int(data.get("exp_general_dias", 0) or 0)
    print(gen_days)
    y, m, d = ymd_from_days(gen_days)
    #ws.cell(row=13, column=base_col).value = f"Experiencia general efectiva: {gen_days} días ({y}a {m}m {d}d)"
    ws.cell(row=13, column=base_col).value = f"{gen_days} días ({y}a {m}m {d}d)"    
    umbral = float(cfg["criterios"]["exp_general"]["umbral_anios_1"])
    ws.cell(row=13, column=score_col).value = "CUMPLE" if (gen_days / 365.25) >= umbral else "NO CUMPLE"

    # EXPERIENCIA ESPECIFICA EFECTIVA (sin superposición)
    spec_days = int(data.get("exp_especifica_dias", 0) or 0)
    print(spec_days)
    y2, m2, d2 = ymd_from_days(spec_days)
    #ws.cell(row=17, column=base_col).value = f"Experiencia específica efectiva: {spec_days} días ({y2}a {m2}m {d2}d)"
    ws.cell(row=17, column=base_col).value = f"{spec_days} días ({y2}a {m2}m {d2}d)"    
    umbral_e = float(cfg["criterios"]["exp_especifica"]["umbral_anios_1"])
    ws.cell(row=17, column=score_col).value = "CUMPLE" if (spec_days / 365.25) >= umbral_e else "NO CUMPLE"

    # Deseables (java/oracle)
    ws.cell(row=20, column=score_col).value = 1 if data.get("java_ok", False) else 0
    ws.cell(row=21, column=score_col).value = 1 if data.get("oracle_ok", False) else 0

    # Pie informativo
    ws.cell(row=22, column=base_col).value = f"DNI: {norm(data.get('dni',''))} | Email: {norm(data.get('email',''))} | Cel: {norm(data.get('celular',''))}"

    return True, f"SLOT_{slot}"


def get_or_create_eval_sheet(wb, template_name: str, sheet_index: int):
    """
    Hoja 1 = plantilla original.
    Hoja 2..N = copia de la plantilla, renombrada "Evaluación CV (2)", etc.
    """
    if sheet_index == 1:
        return wb[template_name]

    title = f"{template_name} ({sheet_index})"
    if title in wb.sheetnames:
        return wb[title]

    new_ws = wb.copy_worksheet(wb[template_name])
    new_ws.title = title
    return new_ws


def save_csv(path: Path, rows: list, header: list):
    ensure_dir(path.parent)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)


# -------------------------
# Main: 1 ARCHIVO POR PROCESO, con tantas hojas como haga falta
# -------------------------
def main():
    #cfg = json.loads(Path("configs/config.json").read_text(encoding="utf-8"))
    cfg_global = json.loads(Path("configs/config.json").read_text(encoding="utf-8"))
    cfg = cfg_global


    #root = Path(cfg["input_root"])
    root = Path(cfg_global["input_root"])
    print(root)  # para que veas rápidamente que está bien seteado

    if not root.exists():
        raise SystemExit(f"No existe input_root: {root}")



    calc_template = Path(cfg.get("calc_template", "CALCULADORA_DE_EXPERIENCIA.xlsx"))
    export_calc = bool(cfg.get("export_calculadora", True))
    use_ocr = bool(cfg.get("pdf", {}).get("use_ocr", False))

    # Para "procesados": copiar o mover (recomendado copiar)
    processed_mode = cfg.get("processed_mode", "copy")  # "copy" o "move"
    if processed_mode not in ("copy", "move"):
        processed_mode = "copy"

    max_slots = int(cfg.get("max_postulantes_por_hoja", 8))

    procesos = [p.name for p in root.iterdir() if p.is_dir()]
    procesos.sort()

    for proceso in procesos:
        in_dir, out_dir = get_in_out_dirs(root, proceso)
        
        template_path = resolve_template_for_proceso(root, proceso, cfg_global)
        #template_path = Path(cfg["output_template"])
        if not template_path.exists():
            log_file_append(debug_log, f"[WARN] No hay plantilla local; usando default: {template_path}")


        # Si no existe entrada, saltar
        if not in_dir.exists():
            continue

        ensure_dir(out_dir)
        procesados_dir = out_dir / PROCESADOS_SUBFOLDER
        calculadoras_dir = out_dir / CALCULADORAS_SUBFOLDER
        ensure_dir(procesados_dir)
        ensure_dir(calculadoras_dir)

        debug_log = out_dir / "debug_run.log"
        log_file_append(debug_log, f"[CFG] plantilla_proceso: {template_path}")

        if debug_log.exists():
            debug_log.unlink(missing_ok=True)

        log_file_append(debug_log, f"== PROCESO: {proceso} ==")
        log_file_append(debug_log, f"in_dir: {in_dir}")
        log_file_append(debug_log, f"out_dir: {out_dir}")
        log_file_append(debug_log, f"max_postulantes_por_hoja: {max_slots}")
        log_file_append(debug_log, f"processed_mode: {processed_mode}")

        files = choose_one_file_per_postulante_folder(in_dir)
        log_file_append(debug_log, f"[DEBUG] archivos elegidos (1 por postulante): {len(files)}")
        for f in files[:60]:
            log_file_append(debug_log, f"   CHOSEN: {f}")

        if not files:
            log_file_append(debug_log, "[SKIP] No hay archivos elegibles en 009 EDI RECIBIDA.")
            continue

        # Cargar una sola vez el workbook de salida del proceso
        wb_out = load_workbook(template_path)
        sheet_idx = 1
        ws_out = get_or_create_eval_sheet(wb_out, EVAL_SHEET_TEMPLATE, sheet_idx)

        log_rows = []
        flat_rows = []
        flat_header = [
            "proceso", "archivo_origen", "archivo_procesado", "tipo",
            "dni", "nombre", "email", "celular",
            "exp_general_dias", "exp_especifica_dias",
            "hoja", "slot"
        ]

        for fp in files:
            try:
                log_file_append(debug_log, f"[START] {fp}")

                # Parse según tipo (Excel preferido ya viene elegido)
                if fp.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
                    data = parse_eoi_excel(fp)
                    ftype = "EXCEL"
                else:
                    data = parse_eoi_pdf(fp, use_ocr=use_ocr)
                    ftype = "PDF"

                log_file_append(debug_log, f"[PARSE_OK] tipo={ftype} dni='{data.get('dni','')}' nombre='{data.get('nombre_full','')}'")

                data["source_label"] = fp.parent.name

                ok, where = write_postulante(ws_out, data, cfg)

                # Si ya no hay slot en la hoja, crear nueva hoja dentro del MISMO archivo
                if not ok and where == "NO_HAY_SLOT":
                    sheet_idx += 1
                    ws_out = get_or_create_eval_sheet(wb_out, EVAL_SHEET_TEMPLATE, sheet_idx)
                    log_file_append(debug_log, f"[INFO] Nueva hoja creada/seleccionada: {ws_out.title}")
                    ok, where = write_postulante(ws_out, data, cfg)

                if not ok:
                    log_file_append(debug_log, f"[ERROR] No se pudo escribir postulante: {where}")
                    log_rows.append([ts(), str(fp), ftype, "ERROR", where])
                    continue

                # Calculadora por postulante
                if export_calc:
                    export_calculadora(calc_template, calculadoras_dir, data, debug_log)

                # Copiar/Mover a procesados
                dest = procesados_dir / fp.name
                if dest.exists():
                    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    dest = procesados_dir / f"{fp.stem}_{stamp}{fp.suffix}"

                if processed_mode == "move":
                    shutil.move(str(fp), str(dest))
                else:
                    shutil.copy2(str(fp), str(dest))

                flat_rows.append([
                    proceso, str(fp), str(dest), ftype,
                    data.get("dni", ""), data.get("nombre_full", ""),
                    data.get("email", ""), data.get("celular", ""),
                    int(data.get("exp_general_dias", 0) or 0),
                    int(data.get("exp_especifica_dias", 0) or 0),
                    ws_out.title, where
                ])

                log_rows.append([ts(), str(fp), ftype, "OK", f"{ws_out.title}:{where}"])
                log_file_append(debug_log, f"[OK] Escrito en {ws_out.title} {where} | procesado -> {dest.name}")

            except Exception as e:
                log_file_append(debug_log, f"[EXCEPTION] {repr(e)}")
                log_rows.append([ts(), str(fp), fp.suffix.lower(), "ERROR", repr(e)])

        # Guardar UN SOLO ARCHIVO por proceso
        out_xlsx = out_dir / f"Cuadro_Evaluacion_{safe_filename(proceso)}.xlsx"
        wb_out.save(out_xlsx)
        log_file_append(debug_log, f"[DONE] Guardado cuadro final: {out_xlsx}")

        # CSVs
        save_csv(out_dir / "consolidado.csv", flat_rows, flat_header)
        save_csv(out_dir / "log.csv", log_rows, ["fecha", "archivo", "tipo", "estado", "detalle"])
        log_file_append(debug_log, f"[DONE] Total postulantes (archivos elegidos): {len(files)}")

    print("Listo. Revisa cada '011. INSTALACIÓN DE COMITÉ' por proceso.")


if __name__ == "__main__":
    main()
