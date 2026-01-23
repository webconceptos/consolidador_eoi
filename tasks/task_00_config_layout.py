
# tasks/task_00_config_layout.py
# -*- coding: utf-8 -*-
"""
Task 00: config_layout
"""

import argparse
import json
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional

import sys
from pathlib import Path

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))


from openpyxl import load_workbook

IN_FOLDER_NAME = "009. EDI RECIBIDAS"
OUT_FOLDER_NAME = "011. INSTALACIÓN DE COMITÉ"

OUT_LAYOUT = "config_layout.json"


def ts():
    return datetime.now().isoformat(timespec="seconds")


def norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def ensure_dir(p: Path):
    p.mkdir(parents=True, exist_ok=True)


def load_global_config() -> dict:
    p = Path("configs/config.json")
    if not p.exists():
        return {}
    return json.loads(p.read_text(encoding="utf-8"))


def find_row_contains(ws, needle: str, max_rows: int = 800, max_cols: int = 25) -> Optional[int]:
    needle = norm(needle).lower()
    for r in range(1, min(ws.max_row, max_rows) + 1):
        row_text = " ".join([norm(str(ws.cell(r, c).value or "")) for c in range(1, max_cols + 1)])
        if needle in row_text.lower():
            return r
    return None


def detect_layout_from_excel(xlsx_path: Path) -> Dict[str, Any]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    eg = find_row_contains(ws, "experiencia general")
    ee = find_row_contains(ws, "experiencia específica") or find_row_contains(ws, "experiencia especifica")

    layout = {
        "generated_at": ts(),
        "source_file": str(xlsx_path),
        "sheet": ws.title,
        "sections": {
            "experiencia_general": {
                "anchor_row": eg or 1,
                "start_row": (eg or 1),
                "end_row_hint": ws.max_row,
            },
            "experiencia_especifica": {
                "anchor_row": ee or None,
                "start_row": (ee or 1) if ee else None,
                "end_row_hint": ws.max_row,
            }
        }
    }
    return layout


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", type=str, default="", help="Carpeta raíz: contiene los procesos")
    ap.add_argument("--only-proc", type=str, default="", help="Procesar solo procesos cuyo nombre contenga este texto")
    args = ap.parse_args()

    cfg = load_global_config()
    root = Path(args.root) if args.root.strip() else Path(cfg.get("input_root", ""))

    if not str(root).strip():
        raise SystemExit("Debes pasar --root o definir input_root en configs/config.json")
    if not root.exists():
        raise SystemExit(f"No existe root: {root}")

    only_filter = norm(args.only_proc).lower()

    procesos = [p for p in root.iterdir() if p.is_dir()]
    procesos.sort(key=lambda x: x.name.lower())

    ok = 0
    skip = 0

    print(f"[task_00] root={root} procesos={len(procesos)}")

    for proc_dir in procesos:
        proceso = proc_dir.name
        if only_filter and only_filter not in proceso.lower():
            continue

        in_dir = proc_dir / IN_FOLDER_NAME
        out_dir = proc_dir / OUT_FOLDER_NAME

        if not in_dir.exists():
            print(f"  - SKIP: {proceso} (no existe 009)")
            skip += 1
            continue

        # elegimos el primer excel que encontremos (solo para detectar layout)
        excels = []
        for p in in_dir.rglob("*"):
            if p.is_file() and p.suffix.lower() in (".xlsx", ".xlsm", ".xls") and not p.name.startswith("~$"):
                excels.append(p)
        excels.sort(key=lambda x: x.name.lower())

        if not excels:
            print(f"  - SKIP: {proceso} (no hay Excel para layout)")
            skip += 1
            continue

        layout = detect_layout_from_excel(excels[0])

        ensure_dir(out_dir)
        (out_dir / OUT_LAYOUT).write_text(json.dumps(layout, ensure_ascii=False, indent=2), encoding="utf-8")

        ok += 1
        print(f"  - OK: {proceso} -> {OUT_LAYOUT}")

    print("")
    print(f"[task_00] resumen: OK={ok} SKIP={skip}")


if __name__ == "__main__":
    main()
