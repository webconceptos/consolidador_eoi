# tasks/task_00_layout_unificado.py
# -*- coding: utf-8 -*-
"""
TASK 00 — Layout Unificado (por proceso)

Objetivo
========
Generar un archivo único de configuración por proceso:

    <PROCESO>/011. INSTALACIÓN DE COMITÉ/config_layout.json

Este archivo es el "contrato" que usarán las tareas siguientes para:
- Crear el archivo "Cuadro_Evaluacion_<NOMBRE_PROCESO>.xlsx" desde la plantilla
  "Revision Preliminar ....xlsx" ubicada en la misma carpeta 011.
- Llenar datos por postulante sin depender de "datos de ejemplo" en la plantilla.
- Dejar trazabilidad clara: cuántos postulantes hay realmente y si hay inconsistencias
  entre el nombre del proceso y la plantilla utilizada.

Uso
===
python tasks/task_00_layout_unificado.py --root "C:\\IA_Investigacion\\ProcesoSelección" --slots-per-sheet 20

Opcionales:
--only-proc "SCI N° 068-2025"
--dry-run
"""

import argparse
import json
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional, List, Tuple

from openpyxl import load_workbook


# --------------------------------------------------------------------
# Convenciones de carpetas y nombres
# --------------------------------------------------------------------
OUT_FOLDER_NAME_DEFAULT = "011. INSTALACIÓN DE COMITÉ"

IN_FOLDER_CANDIDATES = [
    "009. EDI RECIBIDA",
    "009. EDI RECIBIDAS",
]

TEMPLATE_PREFIX = "Revision Preliminar"
TEMPLATE_EXTS = (".xlsx", ".xlsm")  # openpyxl NO soporta .xls
OUT_CONFIG_NAME = "config_layout.json"
PROCESADOS_SUBFOLDER = "procesados"
DEFAULT_HEADER_ROW = 3
DEFAULT_SLOT_START_COL = 6
DEFAULT_SLOT_STEP = 2


# --------------------------------------------------------------------
# Helpers generales
# --------------------------------------------------------------------
def ts() -> str:
    return datetime.now().isoformat(timespec="seconds")


def norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def col_letter(n: int) -> str:
    """Convierte 1->A, 2->B, ..., 27->AA, etc."""
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def is_int_like(s: str) -> bool:
    return bool(re.fullmatch(r"\d+", norm(s)))


def cell_text(ws, r: int, c: int) -> str:
    return norm(str(ws.cell(row=r, column=c).value or ""))


def load_global_config() -> dict:
    """
    Config global opcional:
    - configs/config.json

    Campos sugeridos (opcionales):
    {
      "input_root": "D:/RUTA/PROCESOS",
      "slots_per_sheet": 20,
      "header_row": 3,
      "slot_start_col": 6
    }
    """
    p = Path("configs/config.json")
    if not p.exists():
        return {}
    return json.loads(p.read_text(encoding="utf-8"))


# --------------------------------------------------------------------
# Identificadores auditables (mismatch proceso vs plantilla)
# --------------------------------------------------------------------
def extract_sci_code(text: str) -> Dict[str, Any]:
    """
    Extrae número y año de textos tipo:
    - "SCI N° 068-2025"
    - "SCI 00068-2026"
    Devuelve:
      {"raw": "...", "num": 68, "year": 2025} o {"raw": "...", "num": None, "year": None}
    """
    t = text or ""
    m = re.search(
        r"SCI\s*(?:N[°º]\s*)?(\d{1,6})\s*[-–]\s*(\d{4})",
        t,
        flags=re.IGNORECASE
    )
    if not m:
        return {"raw": norm(t), "num": None, "year": None}
    return {"raw": norm(t), "num": int(m.group(1)), "year": int(m.group(2))}


def compute_mismatch_warnings(process_name: str, template_stem: str) -> Tuple[Dict[str, Any], List[str]]:
    """
    Compara (si se puede) el código SCI de:
    - nombre de proceso (carpeta)
    - nombre de plantilla (sin extensión)
    Retorna: (identifiers, warnings[])
    """
    proc_code = extract_sci_code(process_name)
    tpl_code = extract_sci_code(template_stem)

    warnings: List[str] = []

    if proc_code["num"] and tpl_code["num"] and proc_code["num"] != tpl_code["num"]:
        warnings.append("template_num_mismatch")
    if proc_code["year"] and tpl_code["year"] and proc_code["year"] != tpl_code["year"]:
        warnings.append("template_year_mismatch")

    identifiers = {
        "process_code": proc_code,
        "template_code": tpl_code,
    }
    return identifiers, warnings


# --------------------------------------------------------------------
# Resolución de carpetas 009 / conteo de postulantes
# --------------------------------------------------------------------
def find_009_dir(proc_dir: Path) -> Optional[Path]:
    for name in IN_FOLDER_CANDIDATES:
        p = proc_dir / name
        if p.exists() and p.is_dir():
            return p
    return None


def count_postulantes_en_009(proc_dir: Path) -> int:
    """
    Postulante = carpeta directa dentro de 009.
    Ignora elementos ocultos y no-carpetas.
    """
    in_dir = find_009_dir(proc_dir)
    if not in_dir:
        return 0

    postulantes = [p for p in in_dir.iterdir() if p.is_dir() and not p.name.startswith(".")]
    return len(postulantes)


def ceil_div(a: int, b: int) -> int:
    return (a + b - 1) // b if b > 0 else 0


# --------------------------------------------------------------------
# Plantilla: selección y escaneo de layout
# --------------------------------------------------------------------
def find_process_template(out_dir_011: Path) -> Optional[Path]:
    """
    Busca la plantilla "Revision Preliminar*.xlsx/xlsm" dentro de 011.
    Devuelve el archivo más probable (heurística simple: mtime).
    """
    if not out_dir_011.exists():
        return None

    candidates = []
    for p in out_dir_011.iterdir():
        if not p.is_file():
            continue
        if p.name.startswith("~$"):
            continue
        if p.suffix.lower() not in TEMPLATE_EXTS:
            continue
        if not p.name.lower().startswith(TEMPLATE_PREFIX.lower()):
            continue
        candidates.append(p)

    if not candidates:
        return None

    candidates.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return candidates[0]


def find_base_sheet_name(wb) -> str:
    """
    Intenta usar una hoja "Evaluación CV" si existe,
    caso contrario usa la primera hoja del workbook.
    """
    preferred = "Evaluación CV"
    if preferred in wb.sheetnames:
        return preferred
    return wb.sheetnames[0]


def slot_columns(slot_index: int, start_col: int) -> Tuple[int, int]:
    """
    Cada postulante ocupa 2 columnas:
    - base_col  : información del postulante / campos
    - score_col : puntajes
    """
    base_col = start_col + slot_index * DEFAULT_SLOT_STEP
    score_col = base_col + 1
    return base_col, score_col


def detect_slots_fixed_count(slots_per_sheet: int, start_col: int) -> List[Dict[str, Any]]:
    """
    Genera slots de manera determinística.
    IMPORTANTE: No lee celdas de plantilla para decidir si el slot "existe".
    """
    slots = []
    for i in range(int(slots_per_sheet)):
        base_col, score_col = slot_columns(i, start_col=start_col)
        slots.append({
            "slot_index": i,
            "base_col": base_col,
            "score_col": score_col,
            "base_col_letter": col_letter(base_col),
            "score_col_letter": col_letter(score_col),
        })
    return slots


def find_label_rows(ws, max_rows: int = 80, max_cols: int = 8) -> Dict[str, int]:
    """
    Encuentra filas con títulos/labels de secciones.
    Heurístico, pero útil si la plantilla mantiene títulos similares.
    """
    patterns = {
        "formacion": [r"\bFORMACI[ÓO]N\b", r"\bFORMACION\b", r"\bFORMACI[ÓO]N\s+ACAD"],
        "complementarios": [r"\bESTUDIOS\b", r"\bCOMPLEMENT", r"\bCAPACIT", r"\bCURSOS\b"],
        "exp_general": [r"\bEXPERIENCIA\b.*\bGENERAL\b", r"\bEXP\.\b.*\bGENERAL\b"],
        "exp_especifica": [r"\bEXPERIENCIA\b.*\bESPECIFICA", r"\bEXP\.\b.*\bESPECIFICA"],
        "entrevista": [r"\bENTREVISTA\b"],
        "puntaje_total": [r"\bPUNTAJE\b.*\bTOTAL\b", r"\bTOTAL\b.*\bPUNTAJE\b"],
    }

    found: Dict[str, Optional[int]] = {k: None for k in patterns.keys()}

    for r in range(1, max_rows + 1):
        row_text = " ".join(
            norm(str(ws.cell(row=r, column=c).value or "")) for c in range(1, max_cols + 1)
        ).upper()

        if not row_text.strip():
            continue

        for key, pats in patterns.items():
            if found[key] is not None:
                continue
            for p in pats:
                if re.search(p, row_text, flags=re.IGNORECASE):
                    found[key] = r
                    break

    return {k: v for k, v in found.items() if v is not None}


def find_first_numeric_down(ws, start_row: int, col: int, max_rows: int = 350) -> Optional[int]:
    """
    Busca hacia abajo (col fija) una celda que parezca número (1,2,3,...)
    Útil para ubicar el primer ítem de una tabla numerada.
    """
    for r in range(start_row, min(ws.max_row, max_rows) + 1):
        v = cell_text(ws, r, col)
        if is_int_like(v):
            return r
    return None


def find_section_end_row(ws, start_row: int, stop_at_rows=None, max_rows: int = 350) -> int:
    """
    Estima el final de un bloque:
    - si hay un stop_at_row (inicio de otra sección), corta antes
    - si no, usa heurística de "vacíos consecutivos" en columna C.
    """
    stop_at_rows = [r for r in (stop_at_rows or []) if isinstance(r, int) and r > 0]
    stop_at = min(stop_at_rows) if stop_at_rows else None
    if stop_at and stop_at > start_row:
        return stop_at - 1

    empty_streak = 0
    last_good = start_row
    for r in range(start_row, min(ws.max_row, max_rows) + 1):
        v = cell_text(ws, r, 3)  # col C
        if v.strip() == "":
            empty_streak += 1
        else:
            empty_streak = 0
            last_good = r
        if empty_streak >= 8:
            return last_good
    return last_good


def scan_template_layout(
    template_path: Path,
    slots_per_sheet: int,
    header_row: int,
    slot_start_col: int,
) -> Dict[str, Any]:
    """
    Escanea la plantilla para identificar:
    - Hoja base
    - Labels de secciones (filas)
    - Rango aproximado de experiencia general / específica
    - Y además define explícitamente "targets" (summary_row / total_row) para escribir
      el resumen y el total en el cuadro (tu lógica real de 2 celdas por sección)
    """
    wb = load_workbook(template_path)
    sheet_name = find_base_sheet_name(wb)
    ws = wb[sheet_name]

    slots = detect_slots_fixed_count(slots_per_sheet=slots_per_sheet, start_col=slot_start_col)
    label_rows = find_label_rows(ws)

    # section_rows mantiene los rangos (por compatibilidad) y targets explícitos (lo recomendable)
    section_rows: Dict[str, Any] = {}

    # Formación Académica (fila aproximada siguiente a "FORMACIÓN")
    if "formacion" in label_rows:
        section_rows["fa_row"] = int(label_rows["formacion"]) + 1

    # Estudios Complementarios: base aproximada
    if "complementarios" in label_rows:
        section_rows["ec_row_base"] = int(label_rows["complementarios"]) + 1

    # Experiencia General: ubica primera fila numerada en col C
    if "exp_general" in label_rows:
        r_label = int(label_rows["exp_general"])
        exp_start = find_first_numeric_down(ws, start_row=r_label, col=3, max_rows=350)
        exp_start = exp_start if exp_start is not None else (r_label + 1)

        stop_candidates = []
        if "exp_especifica" in label_rows:
            stop_candidates.append(int(label_rows["exp_especifica"]))
        exp_end = find_section_end_row(ws, start_row=exp_start, stop_at_rows=stop_candidates, max_rows=350)

        section_rows["exp_general_start_row"] = int(exp_start)
        section_rows["exp_general_end_row"] = int(exp_end)

        # Targets explícitos (tu lógica: 2 celdas)
        section_rows["exp_general"] = {
            "summary_row": int(exp_start),
            "total_row": int(exp_end),
        }

    # Experiencia Específica
    if "exp_especifica" in label_rows:
        r_label = int(label_rows["exp_especifica"])
        exp_start = find_first_numeric_down(ws, start_row=r_label, col=3, max_rows=350)
        exp_start = exp_start if exp_start is not None else (r_label + 1)

        stop_candidates = []
        if "puntaje_total" in label_rows:
            stop_candidates.append(int(label_rows["puntaje_total"]))
        exp_end = find_section_end_row(ws, start_row=exp_start, stop_at_rows=stop_candidates, max_rows=350)

        section_rows["exp_especifica_start_row"] = int(exp_start)
        section_rows["exp_especifica_end_row"] = int(exp_end)

        # Targets explícitos (tu lógica: 2 celdas)
        section_rows["exp_especifica"] = {
            "summary_row": int(exp_start),
            "total_row": int(exp_end),
        }

    return {
        "generated_at": ts(),
        "template_file": template_path.name,
        "sheet_base": sheet_name,
        "header_row": int(header_row),
        "slot_start_col": int(slot_start_col),
        "slot_step": int(DEFAULT_SLOT_STEP),
        "slots_per_sheet": int(slots_per_sheet),
        "max_postulantes_por_hoja_detectado": int(slots_per_sheet),  # determinístico
        "slots": slots,
        "label_rows_detectados": label_rows,
        "section_rows": section_rows,
    }


# --------------------------------------------------------------------
# input_hints (opcionales) desde un excel de 009
# --------------------------------------------------------------------
def find_row_contains(ws, needle: str, max_rows: int = 800, max_cols: int = 25) -> Optional[int]:
    needle = norm(needle).lower()
    for r in range(1, min(ws.max_row, max_rows) + 1):
        row_t = " ".join([norm(str(ws.cell(r, c).value or "")) for c in range(1, max_cols + 1)])
        if needle in row_t.lower():
            return r
    return None


def detect_input_hints_from_excel(xlsx_path: Path) -> Dict[str, Any]:
    """
    Lee un Excel de postulante (si existe) y detecta filas donde aparece:
    - "experiencia general"
    - "experiencia específica"
    Es una pista para Task 20 (parser), no una regla dura.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    eg = find_row_contains(ws, "experiencia general")
    ee = find_row_contains(ws, "experiencia específica") or find_row_contains(ws, "experiencia especifica")

    return {
        "generated_at": ts(),
        "source_file": str(xlsx_path),
        "sheet": ws.title,
        "anchors": {
            "experiencia_general": eg,
            "experiencia_especifica": ee,
        }
    }


def pick_one_input_excel(in_dir_009: Path) -> Optional[Path]:
    """
    Busca un Excel cualquiera dentro de 009 (recursivo).
    Si no existe, retorna None (no falla la tarea).
    """
    excels = []
    for p in in_dir_009.rglob("*"):
        if not p.is_file():
            continue
        if p.name.startswith("~$"):
            continue
        if p.suffix.lower() in (".xlsx", ".xlsm"):
            excels.append(p)

    excels.sort(key=lambda x: x.name.lower())
    return excels[0] if excels else None


# --------------------------------------------------------------------
# MAIN
# --------------------------------------------------------------------
def main() -> None:
    cfg = load_global_config()

    ap = argparse.ArgumentParser()
    ap.add_argument("--root", type=str, default="", help="Carpeta raíz: contiene los procesos")
    ap.add_argument("--only-proc", type=str, default="", help="Procesar solo procesos cuyo nombre contenga este texto")
    ap.add_argument("--dry-run", action="store_true", help="No escribe archivos (solo simula)")
    ap.add_argument("--out-folder-name", type=str, default=OUT_FOLDER_NAME_DEFAULT, help="Nombre de carpeta 011")

    ap.add_argument("--slots-per-sheet", type=int, default=0,
                    help="Capacidad de postulantes por hoja (determinístico). 0=usar config/default")
    ap.add_argument("--default-slots-per-sheet", type=int,
                    default=int(cfg.get("slots_per_sheet", 20) or 20),
                    help="Default si no hay slots en config ni por parámetro (por defecto 20)")

    ap.add_argument("--header-row", type=int,
                    default=int(cfg.get("header_row", DEFAULT_HEADER_ROW) or DEFAULT_HEADER_ROW),
                    help="Fila header del cuadro (por defecto 3)")

    ap.add_argument("--slot-start-col", type=int,
                    default=int(cfg.get("slot_start_col", DEFAULT_SLOT_START_COL) or DEFAULT_SLOT_START_COL),
                    help="Columna inicial del primer slot (por defecto 6)")

    args = ap.parse_args()

    root = Path(args.root) if norm(args.root) else Path(cfg.get("input_root", ""))
    if not norm(str(root)):
        raise SystemExit("Debes pasar --root o definir input_root en configs/config.json")
    root = root.resolve()
    if not root.exists():
        raise SystemExit(f"No existe root: {root}")

    out_folder_name = norm(args.out_folder_name) or OUT_FOLDER_NAME_DEFAULT

    # slots_per_sheet determinístico (orden de prioridad: CLI -> config -> default)
    slots_per_sheet = int(args.slots_per_sheet or 0)
    if slots_per_sheet <= 0:
        slots_per_sheet = int(cfg.get("slots_per_sheet", 0) or 0)
    if slots_per_sheet <= 0:
        slots_per_sheet = int(args.default_slots_per_sheet)
    if slots_per_sheet <= 0:
        raise SystemExit("slots_per_sheet debe ser > 0 (usa --slots-per-sheet o config).")

    header_row = int(args.header_row)
    slot_start_col = int(args.slot_start_col)

    only_filter = norm(args.only_proc).lower()

    procesos = [p for p in root.iterdir() if p.is_dir()]
    procesos.sort(key=lambda x: x.name.lower())

    ok = 0
    skip = 0
    fail = 0

    print(f"[task_00_layout_unificado] root={root} procesos={len(procesos)} slots_per_sheet={slots_per_sheet}")

    for proc_dir in procesos:
        proceso = proc_dir.name
        if only_filter and only_filter not in proceso.lower():
            continue

        out_dir_011 = proc_dir / out_folder_name
        tpl = find_process_template(out_dir_011)

        if tpl is None:
            print(f"  - SKIP: {proceso} (no hay plantilla '{TEMPLATE_PREFIX}*.xlsx/xlsm' en {out_folder_name})")
            skip += 1
            continue

        try:
            in_dir_009 = find_009_dir(proc_dir)
            total_post = count_postulantes_en_009(proc_dir)
            sheets_required = ceil_div(total_post, slots_per_sheet) if total_post > 0 else 0

            template_layout = scan_template_layout(
                template_path=tpl,
                slots_per_sheet=slots_per_sheet,
                header_row=header_row,
                slot_start_col=slot_start_col,
            )

            # Warnings de mismatch (auditables)
            identifiers, warnings = compute_mismatch_warnings(proceso, tpl.stem)

            # input_hints opcional (no bloquea)
            hints = None
            if in_dir_009:
                one = pick_one_input_excel(in_dir_009)
                if one:
                    hints = detect_input_hints_from_excel(one)

            unified = {
                "generated_at": ts(),
                "process": proceso,
                "identifiers": identifiers,
                "warnings": warnings,
                "paths": {
                    "process_dir": str(proc_dir),
                    "in_dir_009": str(in_dir_009) if in_dir_009 else None,
                    "out_dir_011": str(out_dir_011),
                    "template_file": tpl.name,
                },
                "runtime": {
                    "total_postulantes": int(total_post),
                    "slots_per_sheet": int(slots_per_sheet),
                    "sheets_required": int(sheets_required),
                },
                "template_layout": template_layout,
                "input_hints": hints,  # puede ser None
            }

            if args.dry_run:
                print(
                    f"  - OK(dry): {proceso} | postulantes={total_post} | hojas_req={sheets_required} "
                    f"| tpl={tpl.name} | slots_per_sheet={slots_per_sheet} "
                    f"| warnings={warnings if warnings else '[]'} | hints={'YES' if hints else 'NO'}"
                )
            else:
                ensure_dir(out_dir_011)
                (out_dir_011 / OUT_CONFIG_NAME).write_text(
                    json.dumps(unified, ensure_ascii=False, indent=2),
                    encoding="utf-8"
                )
                print(
                    f"  - OK: {proceso} -> {OUT_CONFIG_NAME} | postulantes={total_post} | hojas_req={sheets_required} "
                    f"| slots_per_sheet={slots_per_sheet} | warnings={warnings if warnings else '[]'}"
                )

            ok += 1

        except Exception as e:
            fail += 1
            print(f"  - FAIL: {proceso} | {repr(e)}")

    print("")
    print(f"[task_00_layout_unificado] resumen: OK={ok} SKIP={skip} FAIL={fail}")


if __name__ == "__main__":
    main()
