# tasks/task_00_layout_unificado.py
# -*- coding: utf-8 -*-
"""
Task 00 (UNIFICADO): layout_unificado

Objetivo:
- Generar un único config_layout.json por proceso en:
    <Proceso>/011. INSTALACIÓN DE COMITÉ/config_layout.json

Sin perder datos:
A) "template_layout": escanea la PLANTILLA de salida (011) (lo importante para llenar el cuadro)
   - slots, columnas, filas de secciones, rangos de experiencia, etc.
B) "input_hints": opcional, escanea un Excel de 009 (si existe) para hints adicionales
   - anchors de "experiencia general/específica" (solo como pistas, no reemplazan el template)

"""

import argparse
import json
import re
import sys
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional

from openpyxl import load_workbook

# -------------------------
# Constantes
# -------------------------
OUT_FOLDER_NAME = "011. INSTALACIÓN DE COMITÉ"

# soporta ambas variantes del folder 009:
IN_FOLDER_CANDIDATES = [
    "009. EDI RECIBIDA",
    "009. EDI RECIBIDAS",
]

TEMPLATE_PREFIX = "Revision Preliminar"
TEMPLATE_EXTS = (".xlsx", ".xlsm", ".xls")
OUT_CONFIG_NAME = "config_layout.json"


# -------------------------
# Helpers base
# -------------------------
def ts() -> str:
    return datetime.now().isoformat(timespec="seconds")

def norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())

def ensure_dir(p: Path):
    p.mkdir(parents=True, exist_ok=True)

def col_letter(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

def is_int_like(s: str) -> bool:
    s = norm(s)
    return bool(re.fullmatch(r"\d+", s))

def cell_text(ws, r: int, c: int) -> str:
    return norm(str(ws.cell(row=r, column=c).value or ""))

def find_first_numeric_down(ws, start_row: int, col: int, max_rows: int = 350) -> Optional[int]:
    for r in range(start_row, min(ws.max_row, max_rows) + 1):
        v = cell_text(ws, r, col)
        if is_int_like(v):
            return r
    return None

def find_section_end_row(ws, start_row: int, stop_at_rows=None, max_rows: int = 350) -> int:
    stop_at_rows = [r for r in (stop_at_rows or []) if isinstance(r, int) and r > 0]
    stop_at = min(stop_at_rows) if stop_at_rows else None
    if stop_at and stop_at > start_row:
        return stop_at - 1

    empty_streak = 0
    last_good = start_row
    for r in range(start_row, min(ws.max_row, max_rows) + 1):
        v = cell_text(ws, r, 3)  # col C
        if v.strip() == "":
            empty_streak += 1
        else:
            empty_streak = 0
            last_good = r
        if empty_streak >= 8:
            return last_good
    return last_good


# -------------------------
# 1) Template layout (desde 011)
# -------------------------
def find_process_template(out_dir: Path) -> Optional[Path]:
    if not out_dir.exists():
        return None

    candidates = []
    for p in out_dir.iterdir():
        if not p.is_file():
            continue
        if p.name.startswith("~$"):
            continue
        if p.suffix.lower() not in TEMPLATE_EXTS:
            continue
        if not p.name.lower().startswith(TEMPLATE_PREFIX.lower()):
            continue
        candidates.append(p)

    if not candidates:
        return None

    def score(p: Path):
        n = p.name.lower()
        s = 0
        if "sci" in n:
            s += 20
        try:
            s += int(p.stat().st_mtime / 100000)
        except Exception:
            pass
        return s

    candidates.sort(key=score, reverse=True)
    return candidates[0]

def find_base_sheet_name(wb):
    preferred = "Evaluación CV"
    if preferred in wb.sheetnames:
        return preferred
    return wb.sheetnames[0]

def slot_columns(slot_index: int, start_col: int = 6):
    base_col = start_col + slot_index * 2
    score_col = base_col + 1
    return base_col, score_col

def detect_slots(ws, header_row: int = 3, start_col: int = 6, max_scan_slots: int = 50):
    slots = []
    for i in range(max_scan_slots):
        base_col, score_col = slot_columns(i, start_col=start_col)

        v3 = ws.cell(row=header_row, column=base_col).value
        v6 = ws.cell(row=6, column=base_col).value
        v13 = ws.cell(row=13, column=base_col).value

        def has(x):
            return x is not None and str(x).strip() != ""

        #exists = has(v3) or has(v6) or has(v13)
        exists = True

        if not exists:
            if len(slots) > 0:
                break
            continue

        slots.append({
            "slot_index": i,
            "base_col": base_col,
            "score_col": score_col,
            "base_col_letter": col_letter(base_col),
            "score_col_letter": col_letter(score_col),
        })

    return slots

def find_label_rows(ws, max_rows: int = 80, max_cols: int = 8):
    patterns = {
        "formacion": [r"\bFORMACI[ÓO]N\b", r"\bFORMACION\b", r"\bFORMACI[ÓO]N\s+ACAD"],
        "complementarios": [r"\bESTUDIOS\b", r"\bCOMPLEMENT", r"\bCAPACIT", r"\bCURSOS\b"],
        "exp_general": [r"\bEXPERIENCIA\b.*\bGENERAL\b", r"\bEXP\.\b.*\bGENERAL\b"],
        "exp_especifica": [r"\bEXPERIENCIA\b.*\bESPECIFICA", r"\bEXP\.\b.*\bESPECIFICA"],
        "entrevista": [r"\bENTREVISTA\b"],
        "puntaje_total": [r"\bPUNTAJE\b.*\bTOTAL\b", r"\bTOTAL\b.*\bPUNTAJE\b"],
    }

    found = {k: None for k in patterns.keys()}

    for r in range(1, max_rows + 1):
        text = " ".join(
            norm(str(ws.cell(row=r, column=c).value or "")) for c in range(1, max_cols + 1)
        ).upper()
        if not text.strip():
            continue

        for key, pats in patterns.items():
            if found[key] is not None:
                continue
            for p in pats:
                if re.search(p, text, flags=re.IGNORECASE):
                    found[key] = r
                    break

    return {k: v for k, v in found.items() if v is not None}

def row_text(ws, r: int, max_cols: int = 8) -> str:
    return " ".join(norm(str(ws.cell(row=r, column=c).value or "")) for c in range(1, max_cols + 1))

def find_first_row_down(ws, start_row: int, patterns, max_rows: int = 250, max_cols: int = 8) -> Optional[int]:
    for r in range(start_row, min(ws.max_row, max_rows) + 1):
        t = row_text(ws, r, max_cols=max_cols)
        if not t.strip():
            continue
        for p in patterns:
            if re.search(p, t, flags=re.IGNORECASE):
                return r
    return None

def scan_template_layout(template_path: Path) -> Dict[str, Any]:
    wb = load_workbook(template_path)
    sheet_name = find_base_sheet_name(wb)
    ws = wb[sheet_name]

    slots = detect_slots(ws)

    label_rows = find_label_rows(ws)
    section_rows = {}

    # Formación Académica
    if "formacion" in label_rows:
        section_rows["fa_row"] = int(label_rows["formacion"]) + 1

    # Estudios Complementarios
    if "complementarios" in label_rows:
        start = int(label_rows["complementarios"])
        ec_base = find_first_row_down(
            ws, start_row=start, patterns=[r"\bb\s*\.\s*1\b"], max_rows=250, max_cols=8
        )
        section_rows["ec_row_base"] = int(ec_base) if ec_base is not None else (start + 1)

    # Experiencia General (rango)
    if "exp_general" in label_rows:
        r_label = int(label_rows["exp_general"])
        exp_start = find_first_numeric_down(ws, start_row=r_label, col=3, max_rows=350)
        exp_start = exp_start if exp_start is not None else (r_label + 1)
        section_rows["exp_general_start_row"] = exp_start

        candidates_stop = []
        if "exp_especifica" in label_rows:
            candidates_stop.append(int(label_rows["exp_especifica"]))

        section_rows["exp_general_end_row"] = find_section_end_row(
            ws, start_row=exp_start, stop_at_rows=candidates_stop, max_rows=350
        )

    # Experiencia Específica (rango)
    if "exp_especifica" in label_rows:
        r_label = int(label_rows["exp_especifica"])
        exp_start = find_first_numeric_down(ws, start_row=r_label, col=3, max_rows=350)
        exp_start = exp_start if exp_start is not None else (r_label + 1)
        section_rows["exp_especifica_start_row"] = exp_start

        candidates_stop = []
        if "puntaje_total" in label_rows:
            candidates_stop.append(int(label_rows["puntaje_total"]))

        section_rows["exp_especifica_end_row"] = find_section_end_row(
            ws, start_row=exp_start, stop_at_rows=candidates_stop, max_rows=350
        )

    return {
        "generated_at": ts(),
        "template_file": template_path.name,
        "sheet_base": sheet_name,
        "header_row": 3,
        "slot_start_col": 6,
        "slot_step": 2,
        "max_postulantes_por_hoja_detectado": len(slots),
        "slots": slots,
        "label_rows_detectados": label_rows,
        "section_rows": section_rows,
    }


# -------------------------
# 2) Input hints (desde 009) - opcional, no bloquea
# -------------------------
def find_row_contains(ws, needle: str, max_rows: int = 800, max_cols: int = 25) -> Optional[int]:
    needle = norm(needle).lower()
    for r in range(1, min(ws.max_row, max_rows) + 1):
        row_t = " ".join([norm(str(ws.cell(r, c).value or "")) for c in range(1, max_cols + 1)])
        if needle in row_t.lower():
            return r
    return None

def detect_input_hints_from_excel(xlsx_path: Path) -> Dict[str, Any]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    eg = find_row_contains(ws, "experiencia general")
    ee = find_row_contains(ws, "experiencia específica") or find_row_contains(ws, "experiencia especifica")

    return {
        "generated_at": ts(),
        "source_file": str(xlsx_path),
        "sheet": ws.title,
        "anchors": {
            "experiencia_general": eg,
            "experiencia_especifica": ee,
        }
    }

def pick_one_input_excel(in_dir: Path) -> Optional[Path]:
    excels = []
    for p in in_dir.rglob("*"):
        if p.is_file() and p.suffix.lower() in (".xlsx", ".xlsm", ".xls") and not p.name.startswith("~$"):
            excels.append(p)
    excels.sort(key=lambda x: x.name.lower())
    return excels[0] if excels else None

def find_009_dir(proc_dir: Path) -> Optional[Path]:
    for name in IN_FOLDER_CANDIDATES:
        p = proc_dir / name
        if p.exists():
            return p
    return None


# -------------------------
# Main
# -------------------------
def load_global_config() -> dict:
    p = Path("configs/config.json")
    if not p.exists():
        return {}
    return json.loads(p.read_text(encoding="utf-8"))

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", type=str, default="", help="Carpeta raíz: contiene los procesos")
    ap.add_argument("--only-proc", type=str, default="", help="Procesar solo el proceso cuyo nombre contenga este texto")
    ap.add_argument("--dry-run", action="store_true", help="No escribe archivos")
    args = ap.parse_args()

    cfg = load_global_config()
    root = Path(args.root) if args.root.strip() else Path(cfg.get("input_root", ""))

    if not str(root).strip():
        raise SystemExit("Debes pasar --root o definir input_root en configs/config.json")
    if not root.exists():
        raise SystemExit(f"No existe root: {root}")

    only_filter = norm(args.only_proc).lower()

    procesos = [p for p in root.iterdir() if p.is_dir()]
    procesos.sort(key=lambda x: x.name.lower())

    ok = 0
    skip = 0
    fail = 0

    print(f"[task_00_unificado] root={root} procesos={len(procesos)}")

    for proc_dir in procesos:
        proceso = proc_dir.name
        if only_filter and only_filter not in proceso.lower():
            continue

        out_dir = proc_dir / OUT_FOLDER_NAME
        tpl = find_process_template(out_dir)

        if tpl is None:
            print(f"  - SKIP: {proceso} (no existe plantilla '{TEMPLATE_PREFIX}*.xlsx' en 011)")
            skip += 1
            continue

        try:
            template_layout = scan_template_layout(tpl)

            # input hints (opcional)
            hints = None
            in_dir = find_009_dir(proc_dir)
            if in_dir:
                one = pick_one_input_excel(in_dir)
                if one:
                    hints = detect_input_hints_from_excel(one)

            unified = {
                "generated_at": ts(),
                "process": proceso,
                "output_dir": str(out_dir),
                "template_layout": template_layout,
                "input_hints": hints,  # puede ser None
            }

            if args.dry_run:
                print(f"  - OK(dry): {proceso} | tpl={tpl.name} | slots={template_layout.get('max_postulantes_por_hoja_detectado', 0)} | hints={'YES' if hints else 'NO'}")
            else:
                ensure_dir(out_dir)
                (out_dir / OUT_CONFIG_NAME).write_text(
                    json.dumps(unified, ensure_ascii=False, indent=2),
                    encoding="utf-8"
                )
                print(f"  - OK: {proceso} -> {OUT_CONFIG_NAME} | slots={template_layout.get('max_postulantes_por_hoja_detectado', 0)} | hints={'YES' if hints else 'NO'}")

            ok += 1

        except Exception as e:
            fail += 1
            print(f"  - FAIL: {proceso} | {repr(e)}")

    print("")
    print(f"[task_00_unificado] resumen: OK={ok} SKIP={skip} FAIL={fail}")


if __name__ == "__main__":
    main()
