# tasks/task_00_scan_template.py
# -*- coding: utf-8 -*-
"""
Task 00: scan_template
Escanea la plantilla de salida por proceso (ubicada en 011. INSTALACIÓN DE COMITÉ)
y genera un archivo config_layout.json con:
- hoja base de evaluación
- cantidad de slots (postulantes por hoja) detectados
- mapeo de columnas por slot (base_col / score_col)
- detección (heurística) de filas por secciones, si existen rótulos

Uso:
  python tasks/task_00_scan_template.py --root "D:\...\ProcesoSelección"

Si no pasas --root, intenta leer configs/config.json y usar input_root.
"""

import argparse
import json
import re
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook

IN_FOLDER_NAME = "009. EDI RECIBIDA"
OUT_FOLDER_NAME = "011. INSTALACIÓN DE COMITÉ"
PROCESS_OUTPUT_TEMPLATE_NAME = "Formato_Salida_Expresion_Interes.xlsx"
TEMPLATE_PREFIX = "Revision Preliminar"
TEMPLATE_EXTS = (".xlsx", ".xlsm", ".xls")

OUT_CONFIG_NAME = "config_layout.json"


# -------------------------
# Helpers
# -------------------------


def cell_text(ws, r: int, c: int) -> str:
    return norm(str(ws.cell(row=r, column=c).value or ""))

def is_int_like(s: str) -> bool:
    s = norm(s)
    return bool(re.fullmatch(r"\d+", s))

def find_first_numeric_down(ws, start_row: int, col: int, max_rows: int = 250) -> int | None:
    """
    Busca desde start_row hacia abajo la primera fila donde la columna `col`
    contenga un número (Nro de experiencia).
    """
    for r in range(start_row, min(ws.max_row, max_rows) + 1):
        v = cell_text(ws, r, col)
        if is_int_like(v):
            return r
    return None

def find_section_end_row(ws, start_row: int, stop_at_rows: list[int] | None = None, max_rows: int = 250) -> int:
    """
    Determina el fin de una sección:
    - si existe una siguiente sección (stop_at_rows), termina en (min(stop)-1)
    - si no, termina cuando detecta un "vacío sostenido" (ej. 8 filas seguidas sin Nro)
    """
    stop_at_rows = [r for r in (stop_at_rows or []) if isinstance(r, int) and r > 0]
    stop_at = min(stop_at_rows) if stop_at_rows else None

    if stop_at and stop_at > start_row:
        return stop_at - 1

    empty_streak = 0
    last_good = start_row

    for r in range(start_row, min(ws.max_row, max_rows) + 1):
        v = cell_text(ws, r, 3)  # col C
        if v.strip() == "":
            empty_streak += 1
        else:
            empty_streak = 0
            last_good = r

        if empty_streak >= 8:
            return last_good

    return last_good



def ts():
    return datetime.now().isoformat(timespec="seconds")


def norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def col_letter(n: int) -> str:
    # 1->A, 26->Z, 27->AA ...
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def slot_columns(slot_index: int, start_col: int = 6):
    # slot 0 -> F/G (6/7), slot 1 -> H/I (8/9), ...
    base_col = start_col + slot_index * 2
    score_col = base_col + 1
    return base_col, score_col


def find_base_sheet_name(wb):
    # Prioridad: 'Evaluación CV' si existe, si no la primera hoja.
    preferred = "Evaluación CV"
    if preferred in wb.sheetnames:
        return preferred
    return wb.sheetnames[0]


def read_row_values(ws, row: int, max_col: int = 80):
    vals = []
    for c in range(1, max_col + 1):
        v = ws.cell(row=row, column=c).value
        vals.append("" if v is None else str(v))
    return vals

def row_text(ws, r: int, max_cols: int = 8) -> str:
    return " ".join(norm(str(ws.cell(row=r, column=c).value or "")) for c in range(1, max_cols + 1))

def find_first_row_down(ws, start_row: int, patterns: list[str], max_rows: int = 120, max_cols: int = 8) -> int | None:
    """
    Busca desde start_row hacia abajo la primera fila cuyo texto (A..max_cols)
    haga match con alguno de los patrones.
    """
    for r in range(start_row, min(ws.max_row, max_rows) + 1):
        t = row_text(ws, r, max_cols=max_cols)
        if not t.strip():
            continue
        for p in patterns:
            if re.search(p, t, flags=re.IGNORECASE):
                return r
    return None



def find_process_template(out_dir: Path) -> Path | None:
    """
    Busca en 011. INSTALACIÓN DE COMITÉ un archivo Excel que empiece con:
      'Revision Preliminar'
    Si hay varios, prioriza los que contienen 'SCI' y luego el más reciente.
    """
    if not out_dir.exists():
        return None

    candidates = []
    for p in out_dir.iterdir():
        if not p.is_file():
            continue
        if p.suffix.lower() not in TEMPLATE_EXTS:
            continue
        name = p.name.lower()
        if not name.startswith(TEMPLATE_PREFIX.lower()):
            continue
        if p.name.startswith("~$"):
            continue
        candidates.append(p)

    if not candidates:
        return None

    def score(p: Path):
        n = p.name.lower()
        s = 0
        if "sci" in n:
            s += 20
        # más reciente gana
        try:
            s += int(p.stat().st_mtime / 100000)  # normaliza un poco
        except Exception:
            pass
        return s

    candidates.sort(key=score, reverse=True)
    return candidates[0]


def detect_slots(ws, header_row: int = 3, start_col: int = 6, max_scan_slots: int = 80):
    """
    Heurística:
    - Slots empiezan en columna F (6), de 2 en 2 (F/H/J/...)
    - Un slot "existe" si la celda (header_row, base_col) tiene texto no vacío,
      o si las siguientes filas de criterios tienen algo. Para hacerlo robusto,
      validamos mirando 3 filas (3,6,13) en esa columna.
    """
    slots = []
    for i in range(max_scan_slots):
        base_col, score_col = slot_columns(i, start_col=start_col)

        v3 = ws.cell(row=header_row, column=base_col).value
        v6 = ws.cell(row=6, column=base_col).value
        v13 = ws.cell(row=13, column=base_col).value

        def has(x):
            return x is not None and str(x).strip() != ""

        # En plantillas, a veces el encabezado trae "NOMBRE DEL CONSULTOR"
        # o está vacío pero el slot existe igual. Por eso consideramos 3 filas.
        exists = has(v3) or has(v6) or has(v13)

        if not exists:
            # Si ya hemos detectado al menos 1 slot y encontramos vacío sostenido,
            # asumimos que ya no hay más slots.
            if len(slots) > 0:
                break
            # Si no hemos detectado ninguno todavía, seguimos buscando (por si el encabezado está raro)
            continue

        slots.append({
            "slot_index": i,
            "base_col": base_col,
            "score_col": score_col,
            "base_col_letter": col_letter(base_col),
            "score_col_letter": col_letter(score_col),
        })

    return slots


def find_label_rows(ws, max_rows: int = 80, max_cols: int = 8):
    """
    Busca rótulos en el bloque izquierdo (por defecto columnas A..H) para
    ubicar filas de secciones. Esto es heurístico: si el Excel cambia, igual
    te deja el config_layout con slots detectados.
    """
    patterns = {
        "formacion": [r"\bFORMACI[ÓO]N\b", r"\bFORMACION\b", r"\bFORMACI[ÓO]N\s+ACAD"],
        "complementarios": [r"\bESTUDIOS\b", r"\bCOMPLEMENT", r"\bCAPACIT", r"\bCURSOS\b"],
        "exp_general": [r"\bEXPERIENCIA\b.*\bGENERAL\b", r"\bEXP\.\b.*\bGENERAL\b"],
        "exp_especifica": [r"\bEXPERIENCIA\b.*\bESPEC", r"\bEXP\.\b.*\bESPEC"],
        "entrevista": [r"\bENTREVISTA\b"],
        "puntaje_total": [r"\bPUNTAJE\b.*\bTOTAL\b", r"\bTOTAL\b.*\bPUNTAJE\b"],
    }

    found = {k: None for k in patterns.keys()}

    for r in range(1, max_rows + 1):
        text = " ".join(
            norm(str(ws.cell(row=r, column=c).value or "")) for c in range(1, max_cols + 1)
        ).upper()

        if not text.strip():
            continue

        for key, pats in patterns.items():
            if found[key] is not None:
                continue
            for p in pats:
                if re.search(p, text, flags=re.IGNORECASE):
                    found[key] = r
                    break

    # Limpia None
    return {k: v for k, v in found.items() if v is not None}


def scan_one_template(template_path: Path):
    wb = load_workbook(template_path)
    sheet_name = find_base_sheet_name(wb)
    ws = wb[sheet_name]

    slots = detect_slots(ws)
    label_rows = find_label_rows(ws)

    section_rows = {}

    # --- Formación Académica ---
    if "formacion" in label_rows:
        # si el rótulo está en fila X, la data suele estar en X+1
        section_rows["fa_row"] = int(label_rows["formacion"]) + 1

    # --- Estudios Complementarios ---
    if "complementarios" in label_rows:
        start = int(label_rows["complementarios"])
        ec_base = find_first_row_down(
            ws,
            start_row=start,
            patterns=[r"\bb\s*\.\s*1\b"],   # b.1
            max_rows=250,
            max_cols=8
        )
        section_rows["ec_row_base"] = int(ec_base) if ec_base is not None else (start + 1)

    # --- Experiencia General ---
    if "exp_general" in label_rows:
        r_label = int(label_rows["exp_general"])

        # buscamos la primera fila REAL de data: Nro (col C) debajo del rótulo
        exp_start = find_first_numeric_down(ws, start_row=r_label, col=3, max_rows=350)
        # fallback: si no encuentra nro, usa rótulo+1
        exp_start = exp_start if exp_start is not None else (r_label + 1)
        section_rows["exp_general_start_row"] = exp_start

        # Para fin, usamos la siguiente sección más cercana si existe:
        candidates_stop = []
        if "exp_especifica" in label_rows:
            candidates_stop.append(int(label_rows["exp_especifica"]))
        if "entrevista" in label_rows:
            candidates_stop.append(int(label_rows["entrevista"]))
        if "puntaje_total" in label_rows:
            candidates_stop.append(int(label_rows["puntaje_total"]))

        section_rows["exp_general_end_row"] = find_section_end_row(
            ws,
            start_row=exp_start,
            stop_at_rows=candidates_stop,
            max_rows=350
        )

    # --- Experiencia Específica (si la necesitas también) ---
    if "exp_especifica" in label_rows:
        r_label = int(label_rows["exp_especifica"])
        exp_start = find_first_numeric_down(ws, start_row=r_label, col=3, max_rows=350)
        exp_start = exp_start if exp_start is not None else (r_label + 1)
        section_rows["exp_especifica_start_row"] = exp_start

        candidates_stop = []
        if "entrevista" in label_rows:
            candidates_stop.append(int(label_rows["entrevista"]))
        if "puntaje_total" in label_rows:
            candidates_stop.append(int(label_rows["puntaje_total"]))

        section_rows["exp_especifica_end_row"] = find_section_end_row(
            ws,
            start_row=exp_start,
            stop_at_rows=candidates_stop,
            max_rows=350
        )



    layout = {
        "template_file": template_path.name,
        "sheet_base": sheet_name,
        "header_row": 3,
        "slot_start_col": 6,  # F
        "slot_step": 2,
        "max_postulantes_por_hoja_detectado": len(slots),
        "slots": slots,
        "label_rows_detectados": label_rows,
        "section_rows": section_rows,
        "generated_at": ts(),
    }
    #layout["section_rows"] = section_rows

    return layout


def load_global_config() -> dict:
    p = Path("configs/config.json")
    if not p.exists():
        return {}
    return json.loads(p.read_text(encoding="utf-8"))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", type=str, default="", help="Carpeta raíz: contiene los procesos")
    ap.add_argument("--write", action="store_true", help="Escribir config_layout.json (por defecto sí)")
    ap.add_argument("--dry-run", action="store_true", help="No escribe archivos, solo imprime resumen")
    args = ap.parse_args()

    cfg_global = load_global_config()

    root = Path(args.root) if args.root.strip() else Path(cfg_global.get("input_root", ""))
    if not str(root).strip():
        raise SystemExit("Debes pasar --root o definir input_root en configs/config.json")

    if not root.exists():
        raise SystemExit(f"No existe root: {root}")

    procesos = [p for p in root.iterdir() if p.is_dir()]
    procesos.sort(key=lambda x: x.name.lower())

    print(f"[scan_template] root = {root}")
    print(f"[scan_template] procesos detectados = {len(procesos)}")

    ok = 0
    skipped = 0
    fail = 0

    for proc_dir in procesos:
        proceso = proc_dir.name
        out_dir = proc_dir / OUT_FOLDER_NAME
        #template_path = out_dir / PROCESS_OUTPUT_TEMPLATE_NAME
        template_path = find_process_template(out_dir)

        #if not template_path.exists():
        #    skipped += 1
        #    print(f"  - SKIP: {proceso} (no existe plantilla en 011: {template_path.name})")
        #    continue
        
        if template_path is None:
            skipped += 1
            print(f"  - SKIP: {proceso} (no existe plantilla que empiece con '{TEMPLATE_PREFIX}' en 011)")
            continue        

        try:
            layout = scan_one_template(template_path)
            out_config_path = out_dir / OUT_CONFIG_NAME

            if args.dry_run:
                print(f"  - OK(dry): {proceso} | hoja={layout['sheet_base']} | slots={layout['max_postulantes_por_hoja_detectado']} | labels={list(layout['label_rows_detectados'].keys())}")
            else:
                out_dir.mkdir(parents=True, exist_ok=True)
                out_config_path.write_text(json.dumps(layout, ensure_ascii=False, indent=2), encoding="utf-8")
                print(f"  - OK: {proceso} -> {out_config_path.name} | hoja={layout['sheet_base']} | slots={layout['max_postulantes_por_hoja_detectado']} | labels={list(layout['label_rows_detectados'].keys())}")
            ok += 1

        except Exception as e:
            fail += 1
            print(f"  - FAIL: {proceso} | {repr(e)}")

    print("")
    print(f"[scan_template] resumen: OK={ok} SKIP={skipped} FAIL={fail}")


if __name__ == "__main__":
    main()
