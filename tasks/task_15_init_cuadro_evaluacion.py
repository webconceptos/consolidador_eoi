# tasks/task_15_init_cuadro_evaluacion.py
# -*- coding: utf-8 -*-
"""
TASK 15 — Init + Prep Cuadro de Evaluación (por proceso)

Qué resuelve este Task (versión robusta)
========================================
1) Crea el archivo:
      011. INSTALACIÓN DE COMITÉ/Cuadro_Evaluacion_<PROCESO>.xlsx

2) Crea todas las hojas requeridas (según Task 00 o según files_selected.csv).

3) PREPARA SLOTS (LO IMPORTANTE):
   - El header (fila 3) en plantilla suele estar mergeado SOLO para algunos slots (p.ej. 8).
     Este task:
        a) crea el merge (base_col..score_col) para TODOS los slots detectados (p.ej. 20)
        b) copia el estilo del header del slot 0 a todos los slots
        c) LIMPIA también el header (datos personales), dejándolo vacío

   - El cuerpo del slot (filas típicas 4..puntaje_total) debe tener mismo estilo que el slot 0.
     Este task:
        a) copia estilo celda-a-celda (dos columnas) del slot 0 hacia cada slot
        b) replica merges internos del slot 0 (dentro del rango) hacia cada slot
        c) limpia valores de ejemplo en el cuerpo del slot (merge-safe)

4) No toca contenido fuera de slots (no destruye estructura del cuadro).

Entradas
========
- 011/config_layout.json (Task 00)
- 011/files_selected.csv (Task 10)
- Plantilla "Revision Preliminar*.xlsx" (en 011)

Salidas
=======
- 011/Cuadro_Evaluacion_<PROCESO>.xlsx
- 011/init_cuadro_summary.json

Uso
===
python tasks/task_15_init_cuadro_evaluacion.py --root "C:\\IA_Investigacion\\ProcesoSelección"
Opcionales:
--only-proc "SCI N° 067-2025"
--dry-run
--force
--no-prep-slots
"""

import argparse
import csv
import json
import math
import re
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, Any, List, Tuple
from copy import copy

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


# ---------------------------------------------------------------------
# Convenciones
# ---------------------------------------------------------------------
OUT_FOLDER_NAME = "011. INSTALACIÓN DE COMITÉ"
LAYOUT_FILE = "config_layout.json"
SELECTED_FILE = "files_selected.csv"

TEMPLATE_PREFIX = "Revision Preliminar"
TEMPLATE_EXTS = (".xlsx", ".xlsm")

OUT_SUMMARY = "init_cuadro_summary.json"


# ---------------------------------------------------------------------
# Helpers base
# ---------------------------------------------------------------------
def ts() -> str:
    return datetime.now().isoformat(timespec="seconds")


def norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def load_global_config() -> dict:
    p = Path("configs/config.json")
    if not p.exists():
        return {}
    return json.loads(p.read_text(encoding="utf-8"))


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, obj: dict) -> None:
    ensure_dir(path.parent)
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")


def safe_get(d: dict, *keys, default=None):
    cur = d
    for k in keys:
        if not isinstance(cur, dict) or k not in cur:
            return default
        cur = cur[k]
    return cur


def read_selected_count(selected_csv: Path) -> int:
    if not selected_csv.exists():
        return 0
    with selected_csv.open("r", encoding="utf-8") as f:
        r = csv.reader(f)
        header = next(r, None)
        if not header:
            return 0
        rows = [row for row in r if row and any(cell.strip() for cell in row)]
    return len(rows)


# ---------------------------------------------------------------------
# Plantilla
# ---------------------------------------------------------------------
def find_template(out_dir_011: Path) -> Optional[Path]:
    if not out_dir_011.exists():
        return None

    candidates: List[Path] = []
    for p in out_dir_011.iterdir():
        if not p.is_file():
            continue
        if p.name.startswith("~$"):
            continue
        if p.suffix.lower() not in TEMPLATE_EXTS:
            continue
        if not p.name.lower().startswith(TEMPLATE_PREFIX.lower()):
            continue
        candidates.append(p)

    if not candidates:
        return None

    candidates.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return candidates[0]


def ensure_sheet_base_exists(wb, sheet_base: str) -> str:
    return sheet_base if sheet_base in wb.sheetnames else wb.sheetnames[0]


def make_sheet_name(base: str, idx: int) -> str:
    return base if idx == 1 else f"{base} ({idx})"


def copy_base_sheet_n_times(wb, base_sheet_name: str, n_sheets: int) -> List[str]:
    base_ws = wb[base_sheet_name]
    out_names: List[str] = [base_ws.title]

    for i in range(2, n_sheets + 1):
        new_ws = wb.copy_worksheet(base_ws)
        target_name = make_sheet_name(base_ws.title, i)

        if target_name in wb.sheetnames:
            j = 2
            while f"{target_name}.{j}" in wb.sheetnames:
                j += 1
            target_name = f"{target_name}.{j}"

        new_ws.title = target_name
        out_names.append(target_name)

    return out_names


# ---------------------------------------------------------------------
# Slots desde layout
# ---------------------------------------------------------------------
def iter_slot_columns(layout: dict) -> List[Tuple[int, int]]:
    slots = safe_get(layout, "template_layout", "slots", default=[]) or []
    out: List[Tuple[int, int]] = []
    for s in slots:
        out.append((int(s["base_col"]), int(s["score_col"])))
    return out


# ---------------------------------------------------------------------
# Merge-safe cleaning (evita MergedCell read-only)
# ---------------------------------------------------------------------
def merged_anchor_for_cell(ws, row: int, col: int) -> Tuple[int, int]:
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row, rng.min_col
    return row, col


def clear_cell_value_safe(ws, row: int, col: int) -> None:
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        ar, ac = merged_anchor_for_cell(ws, row, col)
        ws.cell(row=ar, column=ac).value = None
    else:
        ws.cell(row=row, column=col).value = None


def infer_slot_body_rows(layout: dict, ws_max_row: int, header_row: int) -> Tuple[int, int]:
    """
    Rango donde replicamos estilos/merges y limpiamos valores del slot.
    Incluimos header_row para limpiar datos personales.
    Preferimos:
      end = label_rows_detectados.puntaje_total
    """
    pt_row = safe_get(layout, "template_layout", "label_rows_detectados", "puntaje_total", default=None)
    end = int(pt_row) if pt_row else min(ws_max_row, header_row + 60)
    end = min(ws_max_row, end)
    return header_row, end



from copy import copy
from openpyxl.styles import Font
# ---------------------------------------------------------------------
# Estilos: copia celda-a-celda
# ---------------------------------------------------------------------
def clone_cell_style(src_cell, dst_cell, font_size: int = 9) -> None:
    """
    Copia estilo completo. Importante usar copy() para no enlazar referencias.
    """
    dst_cell._style = copy(src_cell._style)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = copy(src_cell.protection)
    dst_cell.alignment = copy(src_cell.alignment)
    if src_cell.font:
        dst_cell.font = Font(
            name=src_cell.font.name,
            size=font_size,
            italic=src_cell.font.italic,
            vertAlign=src_cell.font.vertAlign,
            underline=src_cell.font.underline,
            strike=src_cell.font.strike,
            color=src_cell.font.color,
            bold=False,  
        )    


def copy_column_dimensions(ws, src_col: int, dst_col: int) -> None:
    """
    Copia ancho y propiedades de columna.
    """
    src_letter = ws.cell(row=1, column=src_col).column_letter
    dst_letter = ws.cell(row=1, column=dst_col).column_letter

    src_dim = ws.column_dimensions.get(src_letter)
    if src_dim is None:
        return

    dst_dim = ws.column_dimensions[dst_letter]
    # Copia atributos útiles
    dst_dim.width = src_dim.width
    dst_dim.hidden = src_dim.hidden
    dst_dim.outlineLevel = src_dim.outlineLevel
    dst_dim.collapsed = src_dim.collapsed
    dst_dim.bestFit = src_dim.bestFit


# ---------------------------------------------------------------------
# Merges: replicación del slot modelo hacia slots target
# ---------------------------------------------------------------------
def merged_ranges_within(ws, r1: int, r2: int, c1: int, c2: int):
    """
    Devuelve rangos merged que están COMPLETAMENTE dentro del rectángulo.
    """
    out = []
    for rng in ws.merged_cells.ranges:
        if (r1 <= rng.min_row <= rng.max_row <= r2) and (c1 <= rng.min_col <= rng.max_col <= c2):
            out.append(rng)
    return out


def ensure_merge(ws, min_row: int, min_col: int, max_row: int, max_col: int) -> bool:
    """
    Crea merge si no existe uno EXACTO.
    Retorna True si creó, False si ya existía.
    """
    for rng in ws.merged_cells.ranges:
        if rng.min_row == min_row and rng.max_row == max_row and rng.min_col == min_col and rng.max_col == max_col:
            return False
    ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
    return True


def replicate_slot_merges(ws, model_c1: int, model_c2: int, target_c1: int, target_c2: int,
                         row_from: int, row_to: int) -> int:
    """
    Replica merges internos que existan dentro del slot modelo al slot target,
    desplazando columnas por offset.
    """
    created = 0
    offset = target_c1 - model_c1

    for rng in merged_ranges_within(ws, row_from, row_to, model_c1, model_c2):
        new_min_row = rng.min_row
        new_max_row = rng.max_row
        new_min_col = rng.min_col + offset
        new_max_col = rng.max_col + offset

        # Garantiza que caiga dentro del target
        if new_min_col < target_c1 or new_max_col > target_c2:
            continue

        if ensure_merge(ws, new_min_row, new_min_col, new_max_row, new_max_col):
            created += 1

    return created


# ---------------------------------------------------------------------
# Preparación integral de slots (estilo + merges + limpieza)
# ---------------------------------------------------------------------
def prep_slots_full(ws, slot_cols: List[Tuple[int, int]], header_row: int,
                    row_from: int, row_to: int) -> Dict[str, Any]:
    """
    Aplica en UNA hoja:
    - Copia estilos (todo el bloque del slot) desde slot 0 hacia los demás
    - Replica merges internos del slot 0 hacia los demás
    - Asegura merge del header en cada slot (base..score)
    - Limpia valores de ejemplo en TODO el bloque del slot (incluyendo fila 3)

    Retorna métricas.
    """
    report = {
        "header_row": int(header_row),
        "row_from": int(row_from),
        "row_to": int(row_to),
        "slots_total": len(slot_cols),
        "header_merges_created": 0,
        "internal_merges_created": 0,
        "styled_cells_copied": 0,
        "column_dimensions_copied": 0,
        "values_cleared": 0,
    }

    if not slot_cols:
        return report

    # Slot modelo: slot 0
    model_base, model_score = slot_cols[0]
    model_c1, model_c2 = model_base, model_score

    # Copiar widths de columnas del modelo a todos (base y score)
    for (bc, sc) in slot_cols[1:]:
        copy_column_dimensions(ws, model_base, bc)
        copy_column_dimensions(ws, model_score, sc)
        report["column_dimensions_copied"] += 2

    # Asegurar merges del header y replicar merges internos + estilos en todo el bloque
    for idx, (bc, sc) in enumerate(slot_cols):
        # 1) Merge header base..score
        if ensure_merge(ws, header_row, bc, header_row, sc):
            report["header_merges_created"] += 1

        if idx == 0:
            # Para el modelo no replicamos sobre sí mismo, pero sí limpiaremos valores
            continue

        # 2) Replicar merges internos del slot modelo hacia slot target (dentro del rango)
        report["internal_merges_created"] += replicate_slot_merges(
            ws,
            model_c1=model_c1, model_c2=model_c2,
            target_c1=bc, target_c2=sc,
            row_from=row_from, row_to=row_to
        )

        # 3) Copiar estilos celda-a-celda para TODO el bloque [row_from..row_to] x [2 cols]
        for r in range(row_from, row_to + 1):
            # Col base
            src = ws.cell(row=r, column=model_base)
            dst = ws.cell(row=r, column=bc)
            clone_cell_style(src, dst, font_size=7)
            report["styled_cells_copied"] += 1

            # Col score
            src2 = ws.cell(row=r, column=model_score)
            dst2 = ws.cell(row=r, column=sc)
            clone_cell_style(src2, dst2, font_size=7)
            report["styled_cells_copied"] += 1

    # 4) Limpieza de valores (para TODOS los slots, incluyendo el modelo)
    #    Limpia fila 3 (datos personales) + cuerpo, preservando estilos y merges (merge-safe).
    for (bc, sc) in slot_cols:
        for r in range(row_from, row_to + 1):
            clear_cell_value_safe(ws, r, bc)
            report["values_cleared"] += 1
            clear_cell_value_safe(ws, r, sc)
            report["values_cleared"] += 1

    return report


# ---------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------
def main() -> None:
    cfg = load_global_config()

    ap = argparse.ArgumentParser()
    ap.add_argument("--root", type=str, default="", help="Carpeta raíz con procesos")
    ap.add_argument("--only-proc", type=str, default="", help="Procesar solo procesos cuyo nombre contenga este texto")
    ap.add_argument("--dry-run", action="store_true", help="Solo simula (no escribe archivos)")
    ap.add_argument("--force", action="store_true", help="Sobrescribir Cuadro_Evaluacion_*.xlsx si existe")
    ap.add_argument("--no-prep-slots", action="store_true", help="Crear hojas pero sin preparación de slots")
    args = ap.parse_args()

    root = Path(args.root) if norm(args.root) else Path(cfg.get("input_root", ""))
    if not norm(str(root)):
        raise SystemExit("Debes pasar --root o definir input_root en configs/config.json")
    root = root.resolve()
    if not root.exists():
        raise SystemExit(f"No existe root: {root}")

    only_filter = norm(args.only_proc).lower()
    prep_slots = not bool(args.no_prep_slots)

    procesos = [p for p in root.iterdir() if p.is_dir()]
    procesos.sort(key=lambda x: x.name.lower())

    ok, skip, fail = 0, 0, 0
    print(f"[task_15_init_cuadro_evaluacion] root={root} procesos={len(procesos)} (prep_slots={prep_slots})")

    for proc_dir in procesos:
        proceso = proc_dir.name
        if only_filter and only_filter not in proceso.lower():
            continue

        out_dir_011 = proc_dir / OUT_FOLDER_NAME
        layout_path = out_dir_011 / LAYOUT_FILE
        selected_path = out_dir_011 / SELECTED_FILE

        if not out_dir_011.exists():
            print(f"  - SKIP: {proceso} (no existe {OUT_FOLDER_NAME})")
            skip += 1
            continue
        if not layout_path.exists():
            print(f"  - SKIP: {proceso} (falta {LAYOUT_FILE} - ejecuta Task 00)")
            skip += 1
            continue
        if not selected_path.exists():
            print(f"  - SKIP: {proceso} (falta {SELECTED_FILE} - ejecuta Task 10)")
            skip += 1
            continue

        tpl = find_template(out_dir_011)
        if tpl is None:
            print(f"  - SKIP: {proceso} (no se encontró plantilla '{TEMPLATE_PREFIX}*.xlsx/xlsm' en 011)")
            skip += 1
            continue

        try:
            layout = read_json(layout_path)

            # Slots por hoja
            slots_per_sheet = safe_get(layout, "runtime", "slots_per_sheet", default=None)
            if slots_per_sheet is None:
                slots_per_sheet = safe_get(layout, "template_layout", "slots_per_sheet", default=None)
            slots_per_sheet = int(slots_per_sheet or 0)
            if slots_per_sheet <= 0:
                raise ValueError("slots_per_sheet inválido en config_layout.json")

            # Postulantes seleccionados (Task 10)
            selected_count = read_selected_count(selected_path)

            # Hojas requeridas (preferimos Task00; si no, calculamos)
            sheets_required = safe_get(layout, "runtime", "sheets_required", default=None)
            sheets_required = int(sheets_required or 0)
            if sheets_required <= 0:
                sheets_required = max(1, math.ceil(selected_count / slots_per_sheet)) if selected_count > 0 else 1

            # Hoja base / header row
            sheet_base = norm(safe_get(layout, "template_layout", "sheet_base", default="Evaluación CV") or "Evaluación CV")
            header_row = int(safe_get(layout, "template_layout", "header_row", default=3) or 3)

            out_xlsx = out_dir_011 / f"Cuadro_Evaluacion_{proceso}.xlsx"
            out_summary = out_dir_011 / OUT_SUMMARY

            if out_xlsx.exists() and not args.force:
                print(f"  - SKIP: {proceso} (ya existe {out_xlsx.name}; usa --force)")
                skip += 1
                continue

            if args.dry_run:
                print(
                    f"  - OK(dry): {proceso} | tpl={tpl.name} | selected={selected_count} "
                    f"| slots={slots_per_sheet} | sheets_required={sheets_required} | sheet_base='{sheet_base}'"
                )
                ok += 1
                continue

            # Cargar plantilla
            wb = load_workbook(tpl)
            base_name_real = ensure_sheet_base_exists(wb, sheet_base)
            created_sheets = copy_base_sheet_n_times(wb, base_name_real, sheets_required)

            # Preparación de slots integral
            slot_cols = iter_slot_columns(layout)
            prep_report = {"enabled": bool(prep_slots), "by_sheet": {}}

            if prep_slots and slot_cols:
                for sname in created_sheets:
                    ws = wb[sname]
                    row_from, row_to = infer_slot_body_rows(layout, ws.max_row, header_row=header_row)

                    rep = prep_slots_full(
                        ws,
                        slot_cols=slot_cols,
                        header_row=header_row,
                        row_from=row_from,
                        row_to=row_to
                    )
                    prep_report["by_sheet"][sname] = rep

            ensure_dir(out_dir_011)
            wb.save(out_xlsx)

            summary = {
                "generated_at": ts(),
                "process": proceso,
                "paths": {
                    "process_dir": str(proc_dir),
                    "out_dir_011": str(out_dir_011),
                    "template_file": tpl.name,
                    "layout_file": str(layout_path),
                    "selected_file": str(selected_path),
                    "output_xlsx": str(out_xlsx),
                },
                "inputs": {
                    "selected_count": int(selected_count),
                    "slots_per_sheet": int(slots_per_sheet),
                },
                "result": {
                    "sheets_required": int(sheets_required),
                    "sheet_base_requested": sheet_base,
                    "sheet_base_used": base_name_real,
                    "sheets_created": created_sheets,
                    "prep": prep_report,
                },
                "notes": {
                    "prep_copies_full_slot_style_from_slot0": True,
                    "prep_replicates_internal_merges_from_slot0": True,
                    "prep_clears_header_and_body_values": True,
                    "task_40_should_only_fill_values": True
                }
            }
            write_json(out_summary, summary)

            print(
                f"  - OK: {proceso} -> {out_xlsx.name} | selected={selected_count} | sheets={sheets_required} | prep_slots={'YES' if prep_slots else 'NO'}"
            )
            ok += 1

        except Exception as e:
            fail += 1
            print(f"  - FAIL: {proceso} | {repr(e)}")

    print("")
    print(f"[task_15_init_cuadro_evaluacion] resumen: OK={ok} SKIP={skip} FAIL={fail}")


if __name__ == "__main__":
    main()
