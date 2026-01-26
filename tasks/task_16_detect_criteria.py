# task_05_build_criteria.py
# -*- coding: utf-8 -*-
"""
TASK_05 v2 (robusto)
- Lee plantilla + config_layout.json + init_cuadro_summary.json (por proceso)
- Detecta criterios en columna C:
  * Formación (FA) -> C{fa_row}
  * Estudios Complementarios (EC) -> desde ec_row_base, detecta bloques B.1, B.2, ...
  * Experiencia General (EG) -> detecta título por texto ("EXPERIENCIA" + "GENERAL")
  * Experiencia Específica (EE) -> detecta título por texto ("EXPERIENCIA" + "ESPEC")
- STOP: "PUNTAJE TOTAL CV DOCUMENTADO"
- Graba: 011.../procesados/criteria_evaluacion.json
- Debug máximo: imprime rutas, filas detectadas y textos recortados.

Uso:
  python -m tasks.task_05_build_criteria --root "C:\\IA_Investigacion\\ProcesoSelección" --only-proc "SCI N° 085-2025"
"""

import argparse
import json
import re
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

OUT_FOLDER_NAME = "011. INSTALACIÓN DE COMITÉ"
PROCESADOS_SUBFOLDER = "procesados"

CONFIG_LAYOUT_NAME = "config_layout.json"
INIT_SUMMARY_NAME = "init_cuadro_summary.json"
CRITERIA_OUT_NAME = "criteria_evaluacion.json"

DEFAULT_SHEET_BASE = "Evaluación CV"
CRITERIA_COL = "C"  # siempre

TEMPLATE_PREFIXES = [
    "Cuadro_Evaluacion",      # tu plantilla creada
    "Revision Preliminar",    # fallback
]
TEMPLATE_EXTS = (".xlsx", ".xlsm", ".xls")

EVALMODE_COL = "D"

# ---------------------------------------------------------------------
def ts() -> str:
    return datetime.now().isoformat(timespec="seconds")


def ensure_dir(p: Path):
    p.mkdir(parents=True, exist_ok=True)


def norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def read_json(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def short(text: str, n: int = 140) -> str:
    t = norm(text)
    return t if len(t) <= n else t[:n].rstrip() + "..."

# ---------------------------------------------------------------------
# Excel helpers (merged-cells safe)
def cell_value(ws, row: int, col_letter: str):
    """
    Retorna valor incluso si la celda está merged (si cae en un rango merged, devuelve el valor del ancla).
    """
    cell = ws[f"{col_letter}{row}"]
    if not isinstance(cell, MergedCell):
        return cell.value

    # si es merged, buscamos rango que lo contiene y devolvemos valor del ancla
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= cell.column <= rng.max_col:
            anchor = ws.cell(row=rng.min_row, column=rng.min_col)
            return anchor.value
    return None


def parse_eval_mode(d_text: str):
    """
    Retorna (modo_evaluacion, valor)
    - modo_evaluacion: "Cumple_NoCumple" | "Puntaje"
    - valor: "Cumple_NoCumple" o el contenido de la celda D (normalizado)
    """
    s = norm(d_text or "")
    up = s.upper()

    if not s:
        # si está vacío, asumimos puntaje pero sin valor explícito
        return "Puntaje", ""

    # variantes típicas
    if ("CUMPLE" in up and "NO" in up) or ("CUMPLE/NO" in up) or ("NO CUMPLE" in up) or ("CUMPLE - NO" in up):
        return "Cumple_NoCumple", "Cumple_NoCumple"

    # si aparece explícitamente "PUNTAJE"
    if "PUNTAJE" in up:
        return "Puntaje", s

    # si hay dígitos, casi seguro es puntaje (ej: 10, 0-10, Hasta 15, etc.)
    if re.search(r"\d", s):
        return "Puntaje", s

    # fallback
    return "Puntaje", s


def read_criterion_evalmeta(ws, row: int, criteria_col: str = CRITERIA_COL, eval_col: str = EVALMODE_COL):
    c_txt = cell_value(ws, row, criteria_col)
    d_txt = cell_value(ws, row, eval_col)
    modo, valor = parse_eval_mode(str(d_txt) if d_txt is not None else "")
    return {
        "row": row,
        "criteria": {"col": criteria_col, "text": norm(str(c_txt)) if c_txt is not None else ""},
        "eval": {"col": eval_col, "text": norm(str(d_txt)) if d_txt is not None else ""},
        "modo_evaluacion": modo,
        "valor": valor,
    }

def find_row_contains(ws, col_letter: str, pattern: str, start_row: int = 1, end_row: int | None = None):
    """
    Encuentra la primera fila donde col contiene el patrón (case-insensitive).
    """
    end = end_row or ws.max_row
    rx = re.compile(pattern, re.IGNORECASE)
    for r in range(start_row, end + 1):
        v = cell_value(ws, r, col_letter)
        s = norm(str(v)) if v is not None else ""
        if s and rx.search(s):
            return r, s
    return None, ""


def find_section_title(ws, must_have: list[str], col_letter: str = CRITERIA_COL,
                       start_row: int = 1, end_row: int | None = None):
    """
    Encuentra fila donde el texto contiene todas las palabras/fragmentos de must_have (case-insensitive).
    """
    end = end_row or ws.max_row
    must = [m.upper() for m in must_have]
    #print(f"[task_05] find_section_title() buscando en {col_letter}{start_row}:{col_letter}{end} los tokens: {must}")
    

    for r in range(start_row, end + 1):
        v = cell_value(ws, r, col_letter)
        #print(f"  - buscando título en fila {r}: '{short(str(v))}'")
        s = norm(str(v)) if v is not None else ""
        if not s:
            continue
        up = s.upper()
        #print(f"    => '{short(s)}' => '{up}'")
        if all(m in up for m in must):
            #print(f"[task_05] título encontrado en {col_letter}{r}: '{short(s)}'")
            return r, s
    #exit()
    return None, ""


def collect_criterion_lines(ws, col_letter: str, start_row: int, end_row: int,
                           hard_stop_keywords: list[str] | None = None,
                           max_rows: int = 6) -> list[dict]:
    """
    Recolecta líneas de criterio desde start_row hasta end_row (inclusive),
    deteniéndose si encuentra keyword de hard_stop_keywords.
    Devuelve lista de {row, text}.
    """
    hard_stop_keywords = [k.upper() for k in (hard_stop_keywords or [])]

    out = []
    for r in range(start_row, min(end_row, start_row + max_rows - 1) + 1):
        v = cell_value(ws, r, col_letter)
        s = norm(str(v)) if v is not None else ""
        if not s:
            continue

        up = s.upper()
        if any(k in up for k in hard_stop_keywords):
            break
        
        ec_lineas_criterios = read_criterion_evalmeta(ws, r, CRITERIA_COL, EVALMODE_COL)
        
        out.append({
            "criterio_item": {"row": r, "col": CRITERIA_COL, "text": s},
            "modo_evaluacion": ec_lineas_criterios["modo_evaluacion"],
            "valor": ec_lineas_criterios["valor"],
            "eval_cell": {"col": EVALMODE_COL, "text": ec_lineas_criterios["eval"]["text"]},
        })        
        #out.append({"row": r, "col": col_letter, "text": s})
    return out


# ---------------------------------------------------------------------
def parse_layout_min(layout: dict):
    """
    Lee config_layout.json de forma tolerante (sin asumir estructura rígida).
    """
    sheet_base = layout.get("sheet_base") or layout.get("sheet_name") or DEFAULT_SHEET_BASE

    slot_start_col = int(layout.get("slot_start_col", 6))
    slot_step_cols = int(layout.get("slot_step_cols", 2))
    header_row = int(layout.get("header_row", 3))

    fields = layout.get("fields", {}) if isinstance(layout.get("fields", {}), dict) else {}
    rows = layout.get("label_rows_detectados", {}) if isinstance(layout.get("label_rows_detectados", {}), dict) else {}

    def pick_row(*names, default=None):
        for n in names:
            if n in rows and isinstance(rows[n], int):
                return rows[n]
            if n in layout and isinstance(layout[n], int):
                return layout[n]
            if n in fields and isinstance(fields[n], int):
                return fields[n]
        return default

    # Formación (FA)
    fa_row = pick_row("FA", "FA_ROW", "formacion_obligatoria_row", default=6)

    # Estudios complementarios: base (fila inicial)
    # (algunos layouts guardan EC_ROWS, otros ec_row_base, otros label_rows_detectados["EC"])
    ec_row_base = pick_row("EC_BASE", "ec_row_base", "estudios_complementarios_row", default=8)

    # Si existe EC_ROWS como lista, igual la guardamos (pero en v2 preferimos detectar bloques)
    ec_rows = []
    v_ec = layout.get("EC_ROWS") or rows.get("EC_ROWS") or fields.get("EC_ROWS") or rows.get("EC")
    if isinstance(v_ec, list):
        ec_rows = [int(x) for x in v_ec if isinstance(x, int)]

    # Fallbacks numéricos (si no hay detección por texto)
    eg_start_row = pick_row("EG_START", "eg_start_row", "exp_general_start_row", default=14)
    ee_start_row = pick_row("EE_START", "ee_start_row", "exp_especifica_start_row", default=18)

    return {
        "sheet_base": sheet_base,
        "slot_start_col": slot_start_col,
        "slot_step_cols": slot_step_cols,
        "header_row": header_row,
        "fa_row": fa_row,
        "ec_row_base": ec_row_base,
        "ec_rows": ec_rows,
        "eg_start_row": eg_start_row,
        "ee_start_row": ee_start_row,
    }


# ---------------------------------------------------------------------
def find_template_in_011(out_dir: Path) -> Path | None:
    """
    Busca la plantilla en 011 priorizando Cuadro_Evaluacion* antes que Revision Preliminar*
    """
    candidates = []
    for p in out_dir.iterdir():
        if not p.is_file():
            continue
        if p.name.startswith("~$"):
            continue
        if p.suffix.lower() not in TEMPLATE_EXTS:
            continue
        candidates.append(p)

    if not candidates:
        return None

    for pref in TEMPLATE_PREFIXES:
        pref_cands = [p for p in candidates if p.name.lower().startswith(pref.lower())]
        if pref_cands:
            pref_cands.sort(key=lambda x: x.stat().st_mtime, reverse=True)
            return pref_cands[0]

    candidates.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return candidates[0]

def detect_ec_blocks(ws, ec_row_base: int, stop_row: int,
                     col_letter: str = CRITERIA_COL,
                     hard_stop_row: int | None = None):
    """
    Detecta criterios de Estudios Complementarios en columna C.

    Modo 1 (etiquetado): B.1 / B1 / B-1 / B 1 ...
    Modo 2 (no etiquetado): texto directo en filas consecutivas (C8, C9, C10...)

    IMPORTANTE: corta EC al encontrar títulos/labels de secciones siguientes
    (p.ej. EXPERIENCIA GENERAL / EXPERIENCIA ESPECIFICA / PUNTAJE TOTAL).
    """
    end = min(stop_row, hard_stop_row) if hard_stop_row else stop_row
    print(f"[task_05] detect_ec_blocks() desde {col_letter}{ec_row_base} hasta {col_letter}{end}")

    # stops típicos: cuando aparecen, EC terminó
    STOP_TOKENS = [
        "EXPERIENCIA GENERAL",
        "EXPERIENCIA ESPECIFICA",
        "EXPERIENCIA ESPECÍFICA",
        "PUNTAJE TOTAL CV DOCUMENTADO",
        "PUNTAJE TOTAL",
    ]

    def is_stop_line(txt: str) -> bool:
        u = (txt or "").strip().upper()
        return any(tok in u for tok in STOP_TOKENS)

    # Regex tolerante para B1 / B.1 / B-1 / B 1
    rx = re.compile(r"^\s*(B)\s*[\.\-\s]?\s*(\d+)\s*:?\s*(.*)\s*$", re.IGNORECASE)

    blocks = []
    found_tagged = False

    r = ec_row_base
    while r <= end:
        v = cell_value(ws, r, col_letter)
        s_raw = (str(v) if v is not None else "").strip()
        s = norm(s_raw)

        print(f"  - fila {r} => '{short(s_raw)}'")

        if not s:
            r += 1
            continue

        # si aparece título/stop, EC terminó
        if is_stop_line(s):
            print(f"[task_16] STOP EC en {col_letter}{r}: '{short(s)}'")
            break

        m = rx.match(s)
        if m:
            found_tagged = True
            bid = f"B.{int(m.group(2))}"
            rest = norm(m.group(3))
            print(f"    => bloque detectado: {bid} resto: '{rest}'")

            crit_row = r
            crit_text = rest

            # si la celda solo tiene "B.1:" sin texto, tomar siguiente no vacía
            if not crit_text:
                rr = r + 1
                while rr <= end:
                    vv = cell_value(ws, rr, col_letter)
                    ss = norm(str(vv)) if vv is not None else ""
                    if not ss:
                        rr += 1
                        continue
                    if is_stop_line(ss):
                        break
                    crit_row = rr
                    crit_text = ss
                    break

            blocks.append({
                "id": bid,
                "criterion": {"row": crit_row, "col": col_letter, "text": crit_text}
            })

        r += 1

    # Modo 2: no etiquetado -> cada fila no vacía (hasta stop) es un bloque EC.*
    if not found_tagged:
        k = 1
        rr = ec_row_base
        while rr <= end:
            vv = cell_value(ws, rr, col_letter)
            ss = norm(str(vv)) if vv is not None else ""
            if not ss:
                rr += 1
                continue

            if is_stop_line(ss):
                print(f"[task_16] STOP EC(no-tag) en {col_letter}{rr}: '{short(ss)}'")
                break

            blocks.append({
                "id": f"EC.{k}",
                "criterion": {"row": rr, "col": col_letter, "text": ss}
            })
            k += 1
            rr += 1
    print(f"[task_16] detect_ec_blocks() => {len(blocks)} bloques encontrados")

    return blocks

def build_criteria_for_process(proc_dir: Path, template_path: Path, layout_path: Path, init_summary_path: Path,
                               verbose: bool = True):
    out_dir = proc_dir / OUT_FOLDER_NAME
    processed_dir = out_dir / PROCESADOS_SUBFOLDER
    ensure_dir(processed_dir)

    layout = read_json(layout_path) if layout_path.exists() else {}
    init_summary = read_json(init_summary_path) if init_summary_path.exists() else {}

    layout_min = parse_layout_min(layout)

    wb = load_workbook(template_path)
    sheet_name = layout_min["sheet_base"]
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    # STOP row
    stop_row, stop_txt = find_row_contains(
        ws,
        CRITERIA_COL,
        r"PUNTAJE\s+TOTAL\s+CV\s+DOCUMENTADO",
        start_row=1
    )
    if not stop_row:
        stop_row = ws.max_row

    # Titles for EG/EE by text (robusto)
    # Buscar FA (Formación Académica) primero para acotar búsqueda
    fa_title_row, fa_title_txt = find_section_title(
        ws,
        must_have=["FORMACIÓN ACADÉMICA"],
        col_letter=CRITERIA_COL,
        start_row=max(1, layout_min["fa_row"] - 2),
        end_row=stop_row
    )

    # Buscar EC (Estudios Complementarios) primero para acotar búsqueda
    ec_title_row, ec_title_txt = find_section_title(
        ws,
        must_have=["ESTUDIOS COMPLEMENTARIOS"],
        col_letter=CRITERIA_COL,
        start_row=max(1, layout_min["fa_row"] - 2),
        end_row=stop_row
    )

    # Buscar EG (Experiencia General)
    eg_title_row, eg_title_txt = find_section_title(
        ws,
        must_have=["EXPERIENCIA GENERAL"],
        col_letter=CRITERIA_COL,
        start_row=max(1, layout_min["fa_row"] - 2),
        end_row=stop_row
    )

    # Buscar EE (Experiencia Específica) desde después de EG si existe
    ee_search_start = (eg_title_row + 1) if eg_title_row else max(1, layout_min["ee_start_row"] - 2)
    ee_title_row, ee_title_txt = find_section_title(
        ws,
        must_have=["EXPERIENCIA ESPECIFICA"],
        col_letter=CRITERIA_COL,
        start_row=ee_search_start,
        end_row=stop_row
    )

    # Fallback si por texto no encontró (usa layout_min)
    if not eg_title_row:
        eg_title_row = layout_min["eg_start_row"]
        eg_title_txt = norm(str(cell_value(ws, eg_title_row, CRITERIA_COL) or ""))

    if not ee_title_row:
        ee_title_row = layout_min["ee_start_row"]
        ee_title_txt = norm(str(cell_value(ws, ee_title_row, CRITERIA_COL) or ""))

    # Formación (FA): normalmente en la fila fa_row es el criterio (no el título)
    fa_row = layout_min["fa_row"]
    fa_txt = norm(str(cell_value(ws, fa_row, CRITERIA_COL) or ""))
    fa_criterios = read_criterion_evalmeta(ws, fa_row, CRITERIA_COL, EVALMODE_COL)

    # Estudios Complementarios (EC)
    ec_row_base = layout_min["ec_row_base"]
    ec_txt_base = norm(str(cell_value(ws, ec_row_base, CRITERIA_COL) or ""))

    # Detectar bloques B.1.. antes de EG_TITLE (porque EC va antes)
    ec_blocks = detect_ec_blocks(
        ws,
        ec_row_base=ec_row_base,
        stop_row=stop_row,
        col_letter=CRITERIA_COL,
        hard_stop_row=(eg_title_row - 1) if eg_title_row else None
    )

    # Si no hay bloques, tratamos EC como criterio único en ec_row_base
    ec_criteria = []

######
#                "criterio_item": {"row": fa_row, "col": CRITERIA_COL, "text": fa_criterios["criteria"]["text"]},
#                "modo_evaluacion": fa_criterios["modo_evaluacion"],
#                "valor": fa_criterios["valor"],
#                "eval_cell": {"col": EVALMODE_COL, "text": fa_criterios["eval"]["text"]},
#####    
    if ec_blocks:
        for b in ec_blocks:
            ec_blocks_criterios = read_criterion_evalmeta(ws, b["criterion"]["row"], CRITERIA_COL, EVALMODE_COL)

            ec_criteria.append({
                "id": b["id"],
                "criterio_item": {"row": b["criterion"]["row"], "col": CRITERIA_COL, "text": b["criterion"]["text"]},
                "modo_evaluacion": ec_blocks_criterios["modo_evaluacion"],
                "valor": ec_blocks_criterios["valor"],
                "eval_cell": {"col": EVALMODE_COL, "text": ec_blocks_criterios["eval"]["text"]},
            })

    # EG: criterio base y extras debajo hasta antes de EE
    eg_base_start = eg_title_row + 1
    eg_end = (ee_title_row - 1) if ee_title_row and ee_title_row > eg_title_row else stop_row
    eg_lines = collect_criterion_lines(
        ws,
        col_letter=CRITERIA_COL,
        start_row=eg_base_start,
        end_row=eg_end,
        hard_stop_keywords=["EXPERIENCIA ESPECIFICA", "PUNTAJE TOTAL CV DOCUMENTADO"],
        max_rows=6
    )

    # EE: criterio base y extras debajo hasta STOP
    ee_base_start = ee_title_row + 1
    ee_lines = collect_criterion_lines(
        ws,
        col_letter=CRITERIA_COL,
        start_row=ee_base_start,
        end_row=stop_row,
        hard_stop_keywords=["PUNTAJE TOTAL CV DOCUMENTADO"],
        max_rows=8
    )

    criteria = {
        "_meta": {
            "task": "task_16_v3",
            "generated_at": ts(),
            "proceso": proc_dir.name,
            "template": str(template_path),
            "config_layout": str(layout_path),
            "init_summary": str(init_summary_path),
            "sheet_used": ws.title,
            "stop_row": stop_row,
            "stop_cell": f"{CRITERIA_COL}{stop_row}",
            "stop_text": stop_txt,
        },
        "titles": {
            "FA": {"row": fa_title_row, "col": CRITERIA_COL, "text": fa_title_txt},
            "EC": {"row": ec_title_row, "col": CRITERIA_COL, "text": ec_title_txt},
            "EG": {"row": eg_title_row, "col": CRITERIA_COL, "text": eg_title_txt},
            "EE": {"row": ee_title_row, "col": CRITERIA_COL, "text": ee_title_txt},
        },
        "criterios": {
            "FA": {
                "criterio_item": {"row": fa_row, "col": CRITERIA_COL, "text": fa_criterios["criteria"]["text"]},
                "modo_evaluacion": fa_criterios["modo_evaluacion"],
                "valor": fa_criterios["valor"],
                "eval_cell": {"col": EVALMODE_COL, "text": fa_criterios["eval"]["text"]},
            },
            "EC": {
                "blocks": ec_criteria,
                "blocks_detected": len(ec_blocks),
            },
            "EG": {
                "lines": eg_lines,  # lista {row,col,text}
            },
            "EE": {
                "lines": ee_lines,  # lista {row,col,text}
            },
        },
    }

    out_path = processed_dir / CRITERIA_OUT_NAME
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(criteria, f, ensure_ascii=False, indent=2)

    if verbose:
        print(f"[task_15] parse_layout_min() =>")
        print(f"  sheet_base={layout_min['sheet_base']}")
        print(f"  slot_start_col={layout_min['slot_start_col']} slot_step_cols={layout_min['slot_step_cols']} header_row={layout_min['header_row']}")
        print(f"  fa_row={layout_min['fa_row']}")
        print(f"  ec_row_base={layout_min['ec_row_base']} ec_rows={layout_min['ec_rows']}")
        print(f"  eg_start_row={layout_min['eg_start_row']} (det by text -> {eg_title_row})")
        print(f"  ee_start_row={layout_min['ee_start_row']} (det by text -> {ee_title_row})")
        print(f"[task_16] STOP encontrado en {CRITERIA_COL}{stop_row}: '{short(stop_txt, 120)}'")
        print(f"[task_16] {proc_dir.name} => guardado: {out_path}")
        print(f"[task_16] ---- RESUMEN DEBUG ----")
        print(f"  plantilla: {template_path}")
        print(f"  layout: {layout_path}")
        print(f"  init_summary: {init_summary_path}")
        print(f"  sheet: {ws.title}")
        print(f"  stop_row: {stop_row}")
        print(f"  criterios:")
        print(f"    - FA {CRITERIA_COL}{fa_row}: {short(fa_txt)}")
        if ec_criteria:
            if ec_blocks:
                print(f"    - EC base {CRITERIA_COL}{ec_row_base}: {short(ec_txt_base)}")
                for b in ec_criteria:
                    print(f"      * {b['id']} {b['criterio_item']['col']}{b['criterio_item']['row']}: {short(b['criterio_item']['text'])}")
            else:
                print(f"    - EC {CRITERIA_COL}{ec_row_base}: {short(ec_txt_base)}")
        else:
            print(f"    - EC: (no detectado / vacío)")
        print(f"    - EG title {CRITERIA_COL}{eg_title_row}: {short(eg_title_txt)}")
        for ln in eg_lines:
            print(f"      * EG {ln['criterio_item']['col']}{ln['criterio_item']['row']}: {short(ln['criterio_item']['text'])}")
        print(f"    - EE title {CRITERIA_COL}{ee_title_row}: {short(ee_title_txt)}")
        for ln in ee_lines:
            print(f"      * EE {ln['criterio_item']['col']}{ln['criterio_item']['row']}: {short(ln['criterio_item']['text'])}")
        print(f"  EC bloques detectados: {len(ec_blocks)}")

    return out_path

# ---------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", required=True, help="Ruta raíz donde están los procesos")
    ap.add_argument("--only-proc", default="", help="Nombre exacto del proceso (opcional)")
    ap.add_argument("--template", default="", help="Ruta explícita de plantilla (opcional)")
    ap.add_argument("--layout", default="", help="Ruta explícita de config_layout.json (opcional)")
    ap.add_argument("--init-summary", default="", help="Ruta explícita de init_cuadro_summary.json (opcional)")
    ap.add_argument("--quiet", action="store_true", help="Menos logs")
    args = ap.parse_args()

    root = Path(args.root)
    if not root.exists():
        raise SystemExit(f"No existe root: {root}")

    procesos = [p for p in root.iterdir() if p.is_dir()]
    procesos.sort(key=lambda p: p.name.lower())

    any_ok = False

    for proc_dir in procesos:
        if args.only_proc and proc_dir.name != args.only_proc:
            continue

        out_dir = proc_dir / OUT_FOLDER_NAME
        if not out_dir.exists():
            print(f"[task_16] SKIP {proc_dir.name}: no existe 011")
            continue

        # template
        if args.template:
            template_path = Path(args.template)
            if not template_path.exists():
                print(f"[task_16] SKIP {proc_dir.name}: --template no existe: {template_path}")
                continue
        else:
            template_path = find_template_in_011(out_dir)
            if not template_path:
                print(f"[task_16] SKIP {proc_dir.name}: no hay plantilla en 011")
                continue

        # layout
        if args.layout:
            layout_path = Path(args.layout)
        else:
            layout_path = out_dir / CONFIG_LAYOUT_NAME

        if not layout_path.exists():
            print(f"[task_16] SKIP {proc_dir.name}: no existe {CONFIG_LAYOUT_NAME} en 011")
            continue

        # init summary
        if args.init_summary:
            init_summary_path = Path(args.init_summary)
        else:
            init_summary_path = out_dir / INIT_SUMMARY_NAME

        if not init_summary_path.exists():
            print(f"[task_16] SKIP {proc_dir.name}: no existe {INIT_SUMMARY_NAME} en 011")
            continue

        build_criteria_for_process(
            proc_dir=proc_dir,
            template_path=template_path,
            layout_path=layout_path,
            init_summary_path=init_summary_path,
            verbose=(not args.quiet),
        )
        any_ok = True

    if not any_ok:
        raise SystemExit("[task_16] No se procesó ningún proceso (revisa --root/--only-proc)")


if __name__ == "__main__":
    main()
