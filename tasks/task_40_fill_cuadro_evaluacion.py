# tasks/task_40_fill_cuadro_evaluacion.py
# -*- coding: utf-8 -*-

import argparse
import json
import re
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

# -------------------------
# Constantes (tu estructura)
# -------------------------
OUT_FOLDER_NAME = "011. INSTALACIÓN DE COMITÉ"
PROCESADOS_SUBFOLDER = "procesados"
TEMPLATE_PREFIX = "Revision Preliminar"
TEMPLATE_EXTS = (".xlsx", ".xlsm", ".xls")

EVAL_SHEET_TEMPLATE = "Evaluación CV"  # si no existe, usa la primera hoja
OUT_FILENAME_PREFIX = "Cuadro_Evaluacion"

LAYOUT_NAME = "config_layout.json"

# -------------------------
# Helpers generales
# -------------------------
def ts() -> str:
    return datetime.now().isoformat(timespec="seconds")

def norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())

def ensure_dir(p: Path):
    p.mkdir(parents=True, exist_ok=True)

def log_append(path: Path | None, msg: str, also_print: bool = True):
    """
    Log simple a archivo (y opcionalmente a consola).
    Si path es None, solo imprime.
    """
    line = f"[{ts()}] {msg}"
    if also_print:
        print(line)
    if path is None:
        return
    ensure_dir(path.parent)
    with path.open("a", encoding="utf-8") as f:
        f.write(line + "\n")

def safe_filename(s: str, max_len: int = 140) -> str:
    s = norm(s)
    out = []
    for ch in s:
        out.append(ch if ch.isalnum() or ch in "._- " else "_")
    s2 = "".join(out).replace(" ", "_")
    return s2[:max_len]

def log(msg: str):
    print(f"[task_40] {msg}")

# -------------------------
# Layout: compatibilidad viejo/nuevo
# -------------------------
def load_layout(out_dir: Path) -> dict:
    p = out_dir / LAYOUT_NAME
    if not p.exists():
        return {}
    return json.loads(p.read_text(encoding="utf-8"))

def get_template_layout(layout: dict) -> dict:
    """
    Compatibilidad:
    - Si guardas layout nuevo como {"template_layout": {...}, ...} => usa eso
    - Si es layout viejo plano => usa layout directamente
    """
    if isinstance(layout.get("template_layout"), dict):
        return layout["template_layout"]
    return layout

def get_section_rows(tpl: dict) -> dict:
    return tpl.get("section_rows", {}) or {}

def get_sheet_base(tpl: dict) -> str:
    return tpl.get("sheet_base") or tpl.get("sheet") or EVAL_SHEET_TEMPLATE

# -------------------------
# Template finder (011)
# -------------------------
def find_process_template(out_dir: Path) -> Path | None:
    if not out_dir.exists():
        return None

    cands = []
    for p in out_dir.iterdir():
        if not p.is_file():
            continue
        if p.suffix.lower() not in TEMPLATE_EXTS:
            continue
        if p.name.startswith("~$"):
            continue
        if not p.name.lower().startswith(TEMPLATE_PREFIX.lower()):
            continue
        cands.append(p)

    if not cands:
        return None

    # más reciente gana
    cands.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return cands[0]

# -------------------------
# Slots
# -------------------------
def slot_columns(slot_index: int, slot_start_col: int = 6, slot_step_cols: int = 2):
    base_col = slot_start_col + slot_index * slot_step_cols
    score_col = base_col + 1
    return base_col, score_col

def next_free_slot_old(ws, max_slots: int, slot_start_col: int = 6, slot_step_cols: int = 2, header_row: int = 3):
    for i in range(max_slots):
        base_col, _ = slot_columns(i, slot_start_col, slot_step_cols)
        v = ws.cell(row=header_row, column=base_col).value
        if v is None or str(v).strip() == "" or "NOMBRE DEL CONSULTOR" in str(v).upper():
            return i
    return None

def next_free_slot(ws, max_slots: int, slot_start_col: int = 6, slot_step_cols: int = 2, header_row: int = 3):
    """
    Detecta el siguiente slot libre.
    Considera "libre" si:
      - celda header está vacía, o
      - contiene textos placeholder del template (NOMBRE DEL CONSULTOR, etc.)
    """
    PLACEHOLDERS = (
        "NOMBRE DEL CONSULTOR",
        "APELLIDOS Y NOMBRES",
        "POSTULANTE",
        "NOMBRE",
        "DNI",
    )

    for i in range(max_slots):
        base_col = slot_start_col + i * slot_step_cols
        v = ws.cell(row=header_row, column=base_col).value
        s = (str(v).strip() if v is not None else "")

        # libre si está vacío
        if not s:
            return i

        # libre si es placeholder
        up = s.upper()
        if any(ph in up for ph in PLACEHOLDERS):
            return i

    return None


def get_or_create_eval_sheet(wb, template_name: str, sheet_index: int):
    # hoja 1: usa la base
    if sheet_index == 1:
        if template_name in wb.sheetnames:
            return wb[template_name]
        return wb[wb.sheetnames[0]]

    # hojas siguientes: copia
    title = f"{template_name} ({sheet_index})"
    if title in wb.sheetnames:
        return wb[title]

    base = wb[template_name] if template_name in wb.sheetnames else wb[wb.sheetnames[0]]
    ws = wb.copy_worksheet(base)
    ws.title = title
    return ws

# -------------------------
# Styles / formato slot
# -------------------------
def build_slot_styles(cfg: dict):
    slot_cfg = cfg.get("slot_style", {})

    fill_base = PatternFill("solid", fgColor=slot_cfg.get("fill_base", "FFF2CC"))
    fill_score = PatternFill("solid", fgColor=slot_cfg.get("fill_score", "D9E1F2"))

    font_base = Font(
        name=slot_cfg.get("font_name", "Calibri"),
        size=int(slot_cfg.get("font_size", 10)),
        bold=bool(slot_cfg.get("font_bold", False)),
    )
    font_header = Font(
        name=slot_cfg.get("font_name", "Calibri"),
        size=int(slot_cfg.get("header_font_size", 11)),
        bold=True,
    )

    align = Alignment(horizontal=slot_cfg.get("align_h", "left"),
                      vertical=slot_cfg.get("align_v", "top"),
                      wrap_text=True)

    thin = Side(style="thin", color=slot_cfg.get("border_color", "000000"))
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "fill_base": fill_base,
        "fill_score": fill_score,
        "font_base": font_base,
        "font_header": font_header,
        "align": align,
        "border": border
    }

def apply_slot_format(ws, base_col: int, score_col: int, cfg: dict, styles: dict):
    slot_cfg = cfg.get("slot_style", {})
    start_row = int(slot_cfg.get("start_row", 3))
    end_row = int(slot_cfg.get("end_row", 22))

    base_width = float(slot_cfg.get("base_col_width", 34))
    score_width = float(slot_cfg.get("score_col_width", 14))

    ws.column_dimensions[get_column_letter(base_col)].width = base_width
    ws.column_dimensions[get_column_letter(score_col)].width = score_width

    if bool(slot_cfg.get("set_row_heights", True)):
        ws.row_dimensions[3].height = float(slot_cfg.get("row3_height", 18))
        ws.row_dimensions[6].height = float(slot_cfg.get("row6_height", 34))

    for r in range(start_row, end_row + 1):
        cb = ws.cell(row=r, column=base_col)
        cs = ws.cell(row=r, column=score_col)

        cb.fill = styles["fill_base"]
        cb.font = styles["font_base"]
        cb.alignment = styles["align"]
        cb.border = styles["border"]

        cs.fill = styles["fill_score"]
        cs.font = styles["font_base"]
        cs.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cs.border = styles["border"]

    ws.cell(row=3, column=base_col).font = styles["font_header"]
    ws.cell(row=3, column=score_col).font = styles["font_header"]

# -------------------------
# Escritura merge-safe
# -------------------------
def _cell_is_in_merged_range(ws, cell_coord: str):
    for rng in ws.merged_cells.ranges:
        if cell_coord in rng:
            return rng
    return None

def write_value_safe(ws, row: int, col: int, value):
    """
    Escribe value en (row,col) SOLO si:
    - la celda no es merged, o
    - la celda es merged y (row,col) es el anchor del merge (top-left).
    """
    cell = ws.cell(row=row, column=col)
    coord = cell.coordinate
    rng = _cell_is_in_merged_range(ws, coord)

    if rng:
        anchor = (rng.min_row, rng.min_col)
        if (row, col) != anchor:
            # no escribimos para no pisar otra sección
            return False
        ws.cell(row=rng.min_row, column=rng.min_col).value = value
        return True

    cell.value = value
    return True

def apply_wrap(ws, row: int, col: int, vertical="top"):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    if cell.alignment:
        cell.alignment = cell.alignment.copy(wrap_text=True, vertical=vertical)
    else:
        cell.alignment = Alignment(wrap_text=True, vertical=vertical)

# -------------------------
# Parse consolidado.jsonl
# -------------------------
def find_consolidado_jsonl(proc_dir: Path) -> Path | None:
    """
    Busca consolidado.jsonl en el proceso (donde lo deje task_20).
    """
    cands = list(proc_dir.rglob("consolidado.jsonl"))
    if not cands:
        return None
    # el más reciente
    cands.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return cands[0]

def read_jsonl(path: Path) -> list[dict]:
    rows = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            rows.append(json.loads(line))
    return rows

# -------------------------
# Formateo EC
# -------------------------
def block_index(block_id: str) -> int:
    m = re.search(r"\b[bB]\s*\.\s*(\d+)\b", str(block_id))
    return int(m.group(1)) if m else 1

def is_header_row_ec(it: dict) -> bool:
    nro = norm(it.get("nro", "")).lower()
    centro = norm(it.get("centro", "")).lower()
    cap = norm(it.get("capacitacion", "")).lower()
    fi = norm(it.get("fecha_inicio", "")).lower()
    if nro in ("no.", "n°", "nº", "nro", "nro."):
        return True
    if centro in ("centro de estudios", "nombre de la entidad ó empresa"):
        return True
    if cap in ("capacitacion", "capacitación", "nombre del proyecto"):
        return True
    if "fecha de inicio" in fi:
        return True
    return False

def format_course_line(x: dict) -> str:
    centro = norm(x.get("centro", ""))
    cap = norm(x.get("capacitacion", ""))
    fi = norm(x.get("fecha_inicio", ""))
    ff = norm(x.get("fecha_fin", ""))
    horas = x.get("horas", "")

    out = []
    if centro: out.append(centro)
    if cap: out.append(cap)
    fechas = " - ".join([p for p in [fi, ff] if p]).strip()
    if fechas: out.append(fechas)

    h = str(horas).strip() if horas is not None else ""
    if h and h not in ("0", "0.0"):
        if not h.lower().endswith("h"):
            h = f"{h}h"
        out.append(h)

    return " | ".join(out).strip()

# -------------------------
# Writer principal (Stage 1)
# -------------------------
def write_postulante_stage1(ws, data: dict, cfg: dict, styles: dict):
    max_slots = int(cfg.get("max_postulantes_por_hoja", 20))
    slot_start_col = int(cfg.get("slot_start_col", 6))
    slot_step_cols = int(cfg.get("slot_step_cols", 2))
    header_row = int(cfg.get("slot_header_row", 3))

    fa_row = int(cfg.get("fa_row", 6))
    ec_rows = cfg.get("ec_rows", [8, 9, 10, 11])
    if isinstance(ec_rows, str):
        ec_rows = [int(x.strip()) for x in ec_rows.split(",") if x.strip()]

    slot = next_free_slot(ws, max_slots, slot_start_col, slot_step_cols, header_row)
    if slot is None:
        return False, "NO_HAY_SLOT"

    base_col, score_col = slot_columns(slot, slot_start_col, slot_step_cols)
    apply_slot_format(ws, base_col, score_col, cfg, styles)

    # Cabecera
    display_name = norm(data.get("nombre_full", "")) or norm(data.get("nombre", "")) or ""
    if not display_name:
        display_name = norm(data.get("dni", "")) or norm(data.get("source_label", "")) or "SIN_NOMBRE"

    dni = norm(data.get("dni", ""))
    cel = norm(data.get("celular", ""))
    email = norm(data.get("email", ""))

    parts = [display_name]
    if dni: parts.append(f"DNI: {dni}")
    if cel: parts.append(f"Cel: {cel}")
    if email: parts.append(f"Email: {email}")
    header_text = " | ".join(parts).strip()

    write_value_safe(ws, header_row, base_col, header_text)
    apply_wrap(ws, header_row, base_col, vertical="center")

    # Formación Académica
    fa = norm(data.get("formacion_resumen") or "")
    if not fa:
        fa = "SIN DATOS DE FORMACIÓN (revisar EDI)"
    write_value_safe(ws, fa_row, base_col, fa)
    apply_wrap(ws, fa_row, base_col, vertical="top")

    # Estudios Complementarios
    ec = data.get("estudios_complementarios") or {}
    blocks = ec.get("blocks") or []

    # limpia filas EC definidas
    for rr in ec_rows:
        write_value_safe(ws, rr, base_col, None)

    # escribe cada bloque b.1..b.N en su fila
    for b in blocks:
        bid = b.get("id", "")
        idx = block_index(bid)
        if 1 <= idx <= len(ec_rows):
            rr = ec_rows[idx - 1]
        else:
            rr = ec_rows[-1]  # si vienen más bloques, apílalos en la última (mejor que perderlos)

        items = b.get("items") or []

        seen = set()
        lines = []
        for it in items:
            if is_header_row_ec(it):
                continue
            if not (norm(it.get("centro", "")) or norm(it.get("capacitacion", ""))):
                continue

            key = (
                norm(it.get("centro", "")).lower(),
                norm(it.get("capacitacion", "")).lower(),
                norm(it.get("fecha_inicio", "")),
                norm(it.get("fecha_fin", "")),
                str(it.get("horas", "")).strip()
            )
            if key in seen:
                continue
            seen.add(key)
            lines.append(format_course_line(it))

        resumen = "\n".join([x for x in lines if x]).strip()
        if not resumen:
            resumen = "(sin cursos declarados)"

        # si rr ya tenía algo (porque varios bloques cayeron a la última fila), concatena
        existing = ws.cell(row=rr, column=base_col).value
        if existing and str(existing).strip() and existing != "(sin cursos declarados)":
            resumen = str(existing).rstrip() + "\n" + resumen

        write_value_safe(ws, rr, base_col, resumen)
        apply_wrap(ws, rr, base_col, vertical="top")

    #return True, f"SLOT_{slot}"
    return True, {"slot": slot, "base_col": base_col, "score_col": score_col}


def write_experiencia_general_range(ws, base_col: int, data: dict, cfg: dict):
    """
    Llena Experiencia General en el mismo slot (base_col) del postulante.
    Usa rango detectado por layout:
      cfg["exp_general_start_row"], cfg["exp_general_end_row"]
    Fallback: si no existe, usa cfg["exp_general_row"] o no hace nada.
    """

    start_row = cfg.get("exp_general_start_row")
    end_row = cfg.get("exp_general_end_row")

    # fallback simple si no tienes rango (plantillas antiguas)
    if not start_row or not end_row:
        eg_row = cfg.get("exp_general_row")
        if not eg_row:
            return False
        start_row = int(eg_row)
        end_row = int(eg_row)

    start_row = int(start_row)
    end_row = int(end_row)

    # Texto fuente (según cómo lo venga dejando task_20)
    eg = data.get("exp_general") or {}
    print("Experiencia general:")
    print(eg)

    texto = norm(
        data.get("exp_general_resumen")
        or eg.get("resumen")
        or ""
    )

    if not texto:
        texto = "(sin experiencia general declarada)"

    # Limpia el rango antes de escribir
    for r in range(start_row, end_row + 1):
        write_value_safe(ws, r, base_col, None)

    # Escribe repartiendo líneas por fila (si el rango tiene varias filas)
    lines = [ln.strip() for ln in texto.splitlines() if ln.strip()]
    if not lines:
        lines = [texto]

    max_rows = end_row - start_row + 1

    # Si hay 1 fila, escribe todo ahí.
    if max_rows <= 1:
        write_value_safe(ws, start_row, base_col, "\n".join(lines))
        apply_wrap(ws, start_row, base_col, vertical="top")
        return True

    # Si hay varias filas: 1 línea por fila; si sobran líneas, se concatenan en la última
    for i in range(max_rows):
        r = start_row + i
        if i < len(lines):
            val = lines[i]
        else:
            val = ""
        write_value_safe(ws, r, base_col, val)
        apply_wrap(ws, r, base_col, vertical="top")

    # Si sobran líneas, concatena en la última fila
    if len(lines) > max_rows:
        tail = "\n".join(lines[max_rows:])
        last_val = ws.cell(row=end_row, column=base_col).value or ""
        merged = (str(last_val).rstrip() + "\n" + tail).strip()
        write_value_safe(ws, end_row, base_col, merged)
        apply_wrap(ws, end_row, base_col, vertical="top")

    return True

def write_experiencia_general_range_new(ws, base_col: int, data: dict, cfg: dict, debug_log= None):
    """
    Escribe Experiencia General así:
      EMPRESA | CARGO | YYYY-MM-DD - YYYY-MM-DD
    y en la última fila del bloque (la "celda de abajo") escribe:
      TOTAL EG (sin superposición): XXXX días (≈ Yy Mm Dd)

    Soporta 2 entradas:
    - data["exp_general_resumen"] (texto largo como el que pegaste)
    - data["exp_general"]["resumen"]
    """

    import re
    from datetime import datetime, date

    start_row = cfg.get("exp_general_start_row")
    end_row = cfg.get("exp_general_end_row")

    if not start_row or not end_row:
        log_append(debug_log, "[EG] SKIP: no hay exp_general_start_row/end_row en cfg", also_print=False)
        return False

    start_row = int(start_row)
    end_row = int(end_row)

    raw = (
        data.get("exp_general_resumen")
        or (data.get("exp_general") or {}).get("resumen")
        or ""
    ).strip()

    if not raw:
        raw = "(sin experiencia general declarada)"

    # ---------------------------------------------------------
    # 1) Parsear experiencias del texto bruto
    #    Formato típico detectado:
    #    EMPRESA | CARGO | 2025-08-01 00:00:00 - 2025-10-31 00:00:00 | 91 Desc: ...
    # ---------------------------------------------------------
    def _to_date(s: str) -> date | None:
        s = (s or "").strip()
        if not s:
            return None
        # acepta 'YYYY-MM-DD' o 'YYYY-MM-DD 00:00:00'
        s = s.split(" ")[0].strip()
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except Exception:
            return None

    # Captura bloques (empresa | cargo | fi - ff | dias) y corta antes de "Desc:"
    # OJO: empresa/cargo pueden tener guiones y tildes, por eso usamos [^|]+
    pat = re.compile(
        r"(?P<empresa>[^|]+?)\s*\|\s*(?P<cargo>[^|]+?)\s*\|\s*"
        r"(?P<fi>\d{4}-\d{2}-\d{2})(?:\s+\d{2}:\d{2}:\d{2})?\s*-\s*"
        r"(?P<ff>\d{4}-\d{2}-\d{2})(?:\s+\d{2}:\d{2}:\d{2})?\s*\|\s*"
        r"(?P<dias>\d+)",
        flags=re.IGNORECASE
    )

    items = []
    for m in pat.finditer(raw):
        empresa = norm(m.group("empresa"))
        cargo = norm(m.group("cargo"))
        fi = _to_date(m.group("fi"))
        ff = _to_date(m.group("ff"))
        dias = int(m.group("dias"))

        # Validación mínima
        if not empresa or not cargo or not fi or not ff:
            continue

        items.append({
            "empresa": empresa,
            "cargo": cargo,
            "fi": fi,
            "ff": ff,
            "dias": dias,
        })

    # Fallback: si no matcheó nada, escribe el raw recortado (pero igual sin desc)
    if not items:
        # corta cualquier "Desc:" si existe
        raw2 = re.split(r"\bDesc\s*:", raw, maxsplit=1, flags=re.IGNORECASE)[0].strip()
        lines_out = [raw2] if raw2 else ["(sin experiencia general declarada)"]
        total_days_net = 0
    else:
        # ---------------------------------------------------------
        # 2) Armar líneas “limpias”: empresa | cargo | fi - ff
        # ---------------------------------------------------------
        lines_out = [
            f"{it['empresa']} | {it['cargo']} | {it['fi'].isoformat()} - {it['ff'].isoformat()}"
            for it in items
        ]

        # ---------------------------------------------------------
        # 3) Calcular días netos sin superposición (unión de intervalos)
        #    Nota: tratamos intervalos como inclusivos en fechas.
        # ---------------------------------------------------------
        intervals = []
        for it in items:
            a = it["fi"]
            b = it["ff"]
            if b < a:
                a, b = b, a
            intervals.append((a, b))

        intervals.sort(key=lambda x: x[0])

        merged = []
        for a, b in intervals:
            if not merged:
                merged.append([a, b])
                continue
            la, lb = merged[-1]
            # si se solapan o son contiguos (lb + 1 día >= a), merge
            if (lb.toordinal() + 1) >= a.toordinal():
                merged[-1][1] = max(lb, b)
            else:
                merged.append([a, b])

        total_days_net = 0
        for a, b in merged:
            total_days_net += (b.toordinal() - a.toordinal() + 1)

    # ---------------------------------------------------------
    # 4) Escribir en Excel:
    #    - Usamos filas start_row..end_row
    #    - Reservamos la ÚLTIMA fila (end_row) como "celda de abajo" para el total
    # ---------------------------------------------------------
    total_row = end_row
    list_rows = max(0, (end_row - start_row))  # filas disponibles para listar (sin la última)

    # Limpieza
    for r in range(start_row, end_row + 1):
        write_value_safe(ws, r, base_col, None)

    # Escribe listado (hasta list_rows)
    for i in range(min(list_rows, len(lines_out))):
        r = start_row + i
        write_value_safe(ws, r, base_col, lines_out[i])
        apply_wrap(ws, r, base_col, vertical="top")

    # Si sobran líneas, las compactamos en la última línea del listado
    if len(lines_out) > list_rows and list_rows > 0:
        tail = "\n".join(lines_out[list_rows-1:])  # desde la última fila lista
        r = start_row + (list_rows - 1)
        write_value_safe(ws, r, base_col, tail)
        apply_wrap(ws, r, base_col, vertical="top")

    # Convertir días a aproximación Y/M/D (simple: 365/30)
    def approx_ymd(days: int):
        y = days // 365
        rem = days % 365
        m = rem // 30
        d = rem % 30
        return y, m, d

    y, m, d = approx_ymd(int(total_days_net))

    total_text = f"TOTAL EG (sin superposición): {total_days_net} días (≈ {y}a {m}m {d}d)"
    write_value_safe(ws, total_row, base_col, total_text)
    apply_wrap(ws, total_row, base_col, vertical="center")

    log_append(
        debug_log,
        f"[EG] OK r{start_row}..r{end_row} col={base_col} items={len(items)} total_net_days={total_days_net}",
        also_print=False
    )
    return True


# -------------------------
# Cfg base
# -------------------------
def load_global_config() -> dict:
    p = Path("configs/config.json")
    if not p.exists():
        return {}
    return json.loads(p.read_text(encoding="utf-8"))

# -------------------------
# Main
# -------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", required=True, help="Ruta raíz de ProcesoSelección")
    ap.add_argument("--only-proc", default="", help="Nombre exacto de proceso (opcional)")
    ap.add_argument("--limit", type=int, default=0, help="Limitar postulantes (0=sin limite)")
    args = ap.parse_args()

    cfg_global = load_global_config()
    cfg = dict(cfg_global)  # copia
    styles = build_slot_styles(cfg)

    root = Path(args.root)
    if not root.exists():
        raise SystemExit(f"No existe root: {root}")

    procesos = [p for p in root.iterdir() if p.is_dir()]
    print(procesos)
    procesos.sort(key=lambda p: p.name.lower())

    ok_proc = 0
    skip_proc = 0

    log(f"root={root} procesos={len(procesos)}")

    for proc_dir in procesos:
        proceso = proc_dir.name
        if args.only_proc and proceso != args.only_proc:
            continue

        out_dir = proc_dir / OUT_FOLDER_NAME
        if not out_dir.exists():
            log(f"SKIP: {proceso} (no existe 011)")
            skip_proc += 1
            continue

        # 1) layout
        layout = load_layout(out_dir)
        tpl = get_template_layout(layout)
        section_rows = get_section_rows(tpl)

        # inyecta config dinámico sin romper tu writer
        cfg["sheet_base"] = get_sheet_base(tpl)
        cfg["slot_header_row"] = int(tpl.get("header_row", 3))
        cfg["slot_start_col"] = int(tpl.get("slot_start_col", 6))
        cfg["slot_step_cols"] = int(tpl.get("slot_step", tpl.get("slot_step_cols", 2)))
        cfg["max_postulantes_por_hoja"] = int(tpl.get("max_postulantes_por_hoja_detectado", cfg.get("max_postulantes_por_hoja", 20)))

        if "fa_row" in section_rows:
            cfg["fa_row"] = int(section_rows["fa_row"])

        # EC: si tienes ec_rows explícitas, úsalo; si solo tienes ec_row_base, arma 4 filas
        if "ec_rows" in section_rows and isinstance(section_rows["ec_rows"], list) and section_rows["ec_rows"]:
            cfg["ec_rows"] = [int(x) for x in section_rows["ec_rows"]]
        else:
            ec_base = int(section_rows.get("ec_row_base", cfg.get("row_ec_base", 8)))
            cfg["ec_rows"] = [ec_base, ec_base + 1, ec_base + 2, ec_base + 3]

        if "exp_general_start_row" in section_rows:
            cfg["exp_general_start_row"] = int(section_rows["exp_general_start_row"])
        if "exp_general_end_row" in section_rows:
            cfg["exp_general_end_row"] = int(section_rows["exp_general_end_row"])


        # 2) template en 011
        tpl_xlsx = find_process_template(out_dir)
        if not tpl_xlsx:
            log(f"SKIP: {proceso} (no encuentro plantilla '{TEMPLATE_PREFIX}*.xlsx' en 011)")
            skip_proc += 1
            continue

        # 3) consolidado.jsonl del task_20
        cons = find_consolidado_jsonl(proc_dir)
        if not cons:
            log(f"SKIP: {proceso} (no encuentro consolidado.jsonl de task_20)")
            skip_proc += 1
            continue

        data_rows = read_jsonl(cons)
        if args.limit and args.limit > 0:
            data_rows = data_rows[:args.limit]

        if not data_rows:
            log(f"SKIP: {proceso} (consolidado.jsonl vacío)")
            skip_proc += 1
            continue

        log(f"PROCESO: {proceso}")
        log(f"  template={tpl_xlsx.name}")
        log(f"  consolidado={cons}")
        log(f"  postulantes={len(data_rows)}")
        log(f"  cfg(sheet_base={cfg['sheet_base']}, header_row={cfg['slot_header_row']}, fa_row={cfg.get('fa_row')}, ec_rows={cfg.get('ec_rows')})")

        wb = load_workbook(tpl_xlsx)
        sheet_idx = 1
        ws = get_or_create_eval_sheet(wb, cfg["sheet_base"], sheet_idx)

        for d in data_rows:
            # etiqueta fuente (si existe)
            if "source_label" not in d:
                d["source_label"] = norm(d.get("folder", "") or d.get("source_file", "") or "")

##            ok, where = write_postulante_stage1(ws, d, cfg, styles)
##            if not ok and where == "NO_HAY_SLOT":
##                sheet_idx += 1
##                ws = get_or_create_eval_sheet(wb, cfg["sheet_base"], sheet_idx)
##                ok, where = write_postulante_stage1(ws, d, cfg, styles)
##
##            if not ok:
##                log(f"  WARN: no se pudo escribir postulante (reason={where})")
                ok, info = write_postulante_stage1(ws, d, cfg, styles)

                if not ok and info == "NO_HAY_SLOT":
                    sheet_idx += 1
                    ws = get_or_create_eval_sheet(wb, cfg["sheet_base"], sheet_idx)
                    ok, info = write_postulante_stage1(ws, d, cfg, styles)

                if not ok:
                    log(f"  WARN: no se pudo escribir postulante (reason={info})")
                    continue

                # ✅ MISMO SLOT: llena experiencia general
                base_col = info["base_col"]
                write_experiencia_general_range(ws, base_col, d, cfg)


        # salida
        out_xlsx = out_dir / f"{OUT_FILENAME_PREFIX}_{safe_filename(proceso)}.xlsx"
        wb.save(out_xlsx)
        log(f"  ✅ guardado: {out_xlsx}")

        ok_proc += 1

    log(f"RESUMEN: OK={ok_proc} SKIP={skip_proc}")

if __name__ == "__main__":
    main()
