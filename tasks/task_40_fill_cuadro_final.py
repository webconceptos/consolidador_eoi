# task_40_fill_cuadro_evaluacion.py
# -*- coding: utf-8 -*-

import argparse
import json
import re
from pathlib import Path
from datetime import datetime, date, timedelta
from typing import Dict, Any, List, Tuple, Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.cell.cell import MergedCell

OUT_FOLDER_NAME = "011. INSTALACIÓN DE COMITÉ"
PROCESADOS_SUBFOLDER = "procesados"

SUMMARY_NAME = "init_cuadro_summary.json"
LAYOUT_NAME = "config_layout.json"
PARSED_JSONL_NAME = "parsed_postulantes.jsonl"

DEFAULT_SHEET_BASE = "Evaluación CV"
_DATE_FMT = "%d/%m/%Y"
_CAL_ANCHOR = date(2000, 1, 1)  # ancla fija para convertir días -> (y,m,d) real

try:
    from dateutil.relativedelta import relativedelta
except Exception:
    relativedelta = None  # si no está dateutil instalado

def ts() -> str:
    return datetime.now().isoformat(timespec="seconds")

def norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())

def ensure_dir(p: Path):
    p.mkdir(parents=True, exist_ok=True)

def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))

def read_jsonl(path: Path) -> list[dict]:
    rows = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows

def safe_preview(x, n=180):
    s = "" if x is None else str(x)
    s = s.replace("\r\n", "\n")
    return (s[:n] + "…") if len(s) > n else s

def write_value_safe(ws, row: int, col: int, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        # escribir solo si es ancla del rango merged
        for rng in ws.merged_cells.ranges:
            if (rng.min_row <= row <= rng.max_row) and (rng.min_col <= col <= rng.max_col):
                if (row, col) == (rng.min_row, rng.min_col):
                    c = ws.cell(row=rng.min_row, column=rng.min_col)
                    c.value = value
                    c.alignment = Alignment(wrap_text=True, vertical="top")
                return
        return

    cell.value = value
    cell.alignment = Alignment(wrap_text=True, vertical="top")

def detect_max_slots(ws, slot_start_col: int, slot_step_cols: int) -> int:
    max_col = ws.max_column
    if max_col < slot_start_col:
        return 0
    return ((max_col - slot_start_col) // slot_step_cols) + 1

def find_next_slot(ws, max_slots: int, header_row: int, slot_start_col: int, slot_step_cols: int):
    """
    Busca primer slot libre mirando cabecera (header_row) en base_col.
    """
    for i in range(max_slots):
        base_col = slot_start_col + i * slot_step_cols
        c = ws.cell(row=header_row, column=base_col)
        if isinstance(c, MergedCell):
            continue
        v = c.value
        if v is None or str(v).strip() == "":
            return i
        s = str(v).strip().upper()
        if "POSTULANTE" in s or "NOMBRE DEL CONSULTOR" in s:
            return i
    return None

def get_eval_sheet(wb, base_name: str, idx: int):
    """
    idx=1 -> hoja base
    idx>1 -> crea (base_name (idx))
    """
    if idx == 1:
        if base_name in wb.sheetnames:
            return wb[base_name]
        return wb[wb.sheetnames[0]]

    title = f"{base_name} ({idx})"
    if title in wb.sheetnames:
        return wb[title]

    base = wb[base_name] if base_name in wb.sheetnames else wb[wb.sheetnames[0]]
    ws = wb.copy_worksheet(base)
    ws.title = title
    return ws

def coalesce(d: dict, keys: list[str]):
    for k in keys:
        if k in d and d.get(k) not in (None, "", [], {}):
            return k, d.get(k)
    return None, None

def resolve_process_files(proc_dir: Path):
    out_dir = proc_dir / OUT_FOLDER_NAME
    if not out_dir.exists():
        return None

    summary_path = out_dir / SUMMARY_NAME
    if not summary_path.exists():
        return None

    summary = read_json(summary_path)

    # layout
    layout_path = Path(summary.get("paths", {}).get("layout_file", "")) if summary.get("paths") else None
    if not layout_path or not layout_path.exists():
        # fallback local
        layout_path = out_dir / LAYOUT_NAME
        if not layout_path.exists():
            layout_path = None

    layout = read_json(layout_path) if layout_path else {}

    # output excel ya preparado por task_15
    out_xlsx = Path(summary.get("paths", {}).get("output_xlsx", "")) if summary.get("paths") else None
    if not out_xlsx or not out_xlsx.exists():
        # fallback: buscar Cuadro_Evaluacion*.xlsx en 011
        cands = sorted([p for p in out_dir.glob("Cuadro_Evaluacion*.xlsx") if p.is_file() and not p.name.startswith("~$")],
                       key=lambda p: p.stat().st_mtime, reverse=True)
        out_xlsx = cands[0] if cands else None

    # jsonl
    jsonl = out_dir / PROCESADOS_SUBFOLDER / PARSED_JSONL_NAME
    if not jsonl.exists():
        jsonl = proc_dir / PROCESADOS_SUBFOLDER / PARSED_JSONL_NAME
    if not jsonl.exists():
        # última opción: rglob
        cands = sorted(proc_dir.rglob(PARSED_JSONL_NAME), key=lambda p: p.stat().st_mtime, reverse=True)
        jsonl = cands[0] if cands else None

    return out_dir, summary_path, summary, layout_path, layout, out_xlsx, jsonl

def parse_layout_min(layout: dict):
    """
    Lee layout real generado por task_15 / init, priorizando:
      - template_layout (si existe)
      - section_rows (si existe)
    y devuelve lo mínimo para task_40.
    """

    # Si el JSON viene como {"template_layout": {...}}, entramos ahí
    tl = layout.get("template_layout") if isinstance(layout.get("template_layout"), dict) else layout

    sheet_base = tl.get("sheet_base") or tl.get("sheet_name") or DEFAULT_SHEET_BASE

    slot_start_col = int(tl.get("slot_start_col", 6))

    # OJO: en tu JSON es slot_step, no slot_step_cols
    slot_step_cols = int(tl.get("slot_step_cols", tl.get("slot_step", 2)))

    header_row = int(tl.get("header_row", 3))

    sr = tl.get("section_rows", {}) if isinstance(tl.get("section_rows", {}), dict) else {}

    # --- filas principales ---
    fa_row = int(sr.get("fa_row", tl.get("fa_row", 6)))

    # EC: algunos procesos pueden tener 1, 2, 4... aquí lo resolvemos con:
    # 1) ec_rows explícito si existe
    # 2) ec_row_base + ec_row_count si existe
    # 3) default 4 (tu caso clásico)
    ec_rows = None

    if isinstance(sr.get("ec_row_base"), list) and sr["ec_row_base"]:
        ec_rows = [int(x) for x in sr["ec_rows"]]
    else:
        ec_base = sr.get("ec_row_base", tl.get("ec_row_base", 8))
        abcd=sr.get("ec_row_base")
        defg=sr.get("exp_general_start_row")        
        ec_count = defg - abcd-1
        try:
            ec_base = int(ec_base)
            ec_count = int(ec_count)
            ec_rows = list(range(ec_base, ec_base + ec_count))
        except Exception:
            ec_rows = [8, 9, 10]

    # Experiencia general/específica: en tu JSON están como objetos con summary_row/total_row
    eg = sr.get("exp_general", {}) if isinstance(sr.get("exp_general", {}), dict) else {}
    ee = sr.get("exp_especifica", {}) if isinstance(sr.get("exp_especifica", {}), dict) else {}

    eg_detail_row = int(eg.get("summary_row", tl.get("eg_detail_row", 15)))
    #eg_total_row  = int(eg.get("total_row",   tl.get("eg_total_row", 14)))-1
    eg_total_row  = int(eg.get("summary_row", tl.get("eg_detail_row", 15)))+1

    ee_detail_row = int(ee.get("summary_row", tl.get("ee_detail_row", 19)))
    #ee_total_row  = int(ee.get("total_row",   tl.get("ee_total_row", 18)))-1
    ee_total_row  = int(ee.get("summary_row", tl.get("ee_detail_row", 19)))+1
    


    return {
        "sheet_base": sheet_base,
        "slot_start_col": slot_start_col,
        "slot_step_cols": slot_step_cols,
        "header_row": header_row,
        "fa_row": fa_row,
        "ec_rows": ec_rows,
        "eg_total_row": eg_total_row,
        "eg_detail_row": eg_detail_row,
        "ee_total_row": ee_total_row,
        "ee_detail_row": ee_detail_row,
    }

##a task20
def _parse_date(s: str) -> Optional[date]:
    s = (s or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, _DATE_FMT).date()
    except Exception:
        return None
##a task20
def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())

##a task20
def _merge_intervals(intervals: List[Tuple[date, date]]) -> List[Tuple[date, date]]:
    """
    intervals: lista de (start, end) con end INCLUSIVO.
    Mergea superposiciones y adyacentes.
    """
    if not intervals:
        return []
    intervals = sorted(intervals, key=lambda x: (x[0], x[1]))
    merged = [intervals[0]]
    for s, e in intervals[1:]:
        ps, pe = merged[-1]
        # si se superpone o es adyacente (pe + 1 día >= s), unir
        if s <= pe + timedelta(days=1):
            merged[-1] = (ps, max(pe, e))
        else:
            merged.append((s, e))
    return merged

##a task20
def _days_inclusive(s: date, e: date) -> int:
    return (e - s).days + 1

##a task20
def _days_to_ymd_calendar_real(total_days: int, anchor: date = _CAL_ANCHOR) -> Tuple[int, int, int]:
    """
    Convierte días -> (años, meses, días) en calendario real usando un anchor fijo.
    Requiere python-dateutil.
    """
    if relativedelta is None:
        raise RuntimeError("Falta python-dateutil. Instala con: pip install python-dateutil")

    end = anchor + timedelta(days=total_days)
    rd = relativedelta(end, anchor)
    return rd.years, rd.months, rd.days

def _days_to_ymd(total_days: int) -> Tuple[int, int, int]:
    """
    Conversión estable tipo "RRHH":
    1 año = 365 días, 1 mes = 30 días.
    """
    years = total_days // 365
    rem = total_days % 365
    months = rem // 30
    days = rem % 30
    return years, months, days

# --- helper bonito para texto final ---
def format_ymd(y: int, m: int, d: int) -> str:
    return f"{y} año(s), {m} mes(es), {d} día(s)"

def compute_experience_summary_and_total(exp_block: Dict[str, Any]) -> Tuple[str, Tuple[int, int, int], int, List[Tuple[date, date]],str]:
    """
    exp_block: dict con estructura como:
      {"items":[{... fecha_inicio, fecha_fin, entidad, cargo, descripcion ...}], ...}

    Retorna:
      - resumen_text (str)
      - (años, meses, días)
      - total_days_acumulado (int)  # ya sin duplicados
      - merged_intervals (list[(start,end)])  # útil para depuración
    """
    items = exp_block.get("items") or []
    if not isinstance(items, list):
        items = []

    resumen_parts: List[str] = []
    detalle_parts: List[str] = []
    raw_intervals: List[Tuple[date, date]] = []

    for it in items:
        if not isinstance(it, dict):
            continue

        entidad = _norm(it.get("entidad", ""))
        cargo = _norm(it.get("cargo", ""))
        f1s = _norm(it.get("fecha_inicio", ""))
        f2s = _norm(it.get("fecha_fin", ""))

        d1 = _parse_date(f1s)
        d2 = _parse_date(f2s)

        # --- armar resumen (aunque falten fechas) ---
        header = f"{entidad} - {cargo}".strip(" -")
        if f1s or f2s:
            header += f" | {f1s or '?'} a {f2s or '?'}"

        desc = (it.get("descripcion") or "").strip()
        if desc:
            resumen_parts.append(f"{header}\n  Desc: {desc}")
        else:
            resumen_parts.append(f"{header}")

        detalle_parts.append(f"- {header}")

        # --- intervalos para acumulado ---
        # Solo contamos si ambas fechas son válidas y ordenables
        if d1 and d2:
            if d2 < d1:
                d1, d2 = d2, d1
            raw_intervals.append((d1, d2))

    resumen_text = "\n\n".join([p for p in resumen_parts if p.strip()]).strip()
    detalle_text = "\n".join([p for p in detalle_parts if p.strip()]).strip()

    merged = _merge_intervals(raw_intervals)
    total_days = sum(_days_inclusive(s, e) for s, e in merged)
    y, m, d = _days_to_ymd(total_days)

    return resumen_text, (y, m, d), total_days, merged, detalle_text

def compute_experience_summary_and_total_calendar_real(
    exp_block: Dict[str, Any],
    anchor: date = _CAL_ANCHOR
) -> Tuple[str, Tuple[int, int, int], int, List[Tuple[date, date]],str]:
    """
    Retorna:
      - resumen_text
      - (años, meses, días) calendario real (con anchor fijo)
      - total_days_unicos (sin superposición)
      - merged_intervals (para depuración)
    """
    items = exp_block.get("items") or []
    if not isinstance(items, list):
        items = []

    resumen_parts: List[str] = []
    detalle_parts: List[str] = []
    raw_intervals: List[Tuple[date, date]] = []

    for it in items:
        if not isinstance(it, dict):
            continue

        entidad = _norm(it.get("entidad", ""))
        cargo = _norm(it.get("cargo", ""))
        f1s = _norm(it.get("fecha_inicio", ""))
        f2s = _norm(it.get("fecha_fin", ""))

        d1 = _parse_date(f1s)
        d2 = _parse_date(f2s)

        # Resumen
        header = f"{entidad} - {cargo}".strip(" -")
        if f1s or f2s:
            header += f" | {f1s or '?'} a {f2s or '?'}"

        desc = (it.get("descripcion") or "").strip()
        if desc:
            resumen_parts.append(f"{header}\n  Desc: {desc}")
        else:
            resumen_parts.append(header)

        detalle_parts.append(f"- {header}")

        # Intervalos (solo si hay fechas válidas)
        if d1 and d2:
            if d2 < d1:
                d1, d2 = d2, d1
            raw_intervals.append((d1, d2))

    resumen_text = "\n\n".join([p for p in resumen_parts if p.strip()]).strip()
    detalle_text = "\n".join([p for p in detalle_parts if p.strip()]).strip()    

    merged = _merge_intervals(raw_intervals)
    total_days = sum(_days_inclusive(s, e) for s, e in merged)
    y, m, d = _days_to_ymd_calendar_real(total_days, anchor=anchor)

    return resumen_text, (y, m, d), total_days, merged, detalle_text

def fill_slot(ws, slot_idx: int, payload: dict, lay: dict, debug_item: dict):
    base_col = lay["slot_start_col"] + slot_idx * lay["slot_step_cols"]

    # HEADER
    nombre = norm(payload.get("nombre_full", "")) or norm(payload.get("nombres", ""))
    dni = norm(payload.get("dni", ""))
    header = nombre if not dni else f"{nombre}\nDNI: {dni}"
    write_value_safe(ws, lay["header_row"], base_col, header)

    # Formación
    write_value_safe(ws, lay["fa_row"], base_col, payload.get("formacion_obligatoria_resumen", "") or "")

    # Estudios complementarios: NO asumimos B.1/B.2.
    ec_rows = lay["ec_rows"] if isinstance(lay["ec_rows"], list) else []


    ec_text = payload.get("estudios_complementarios_resumen", "") or ""
    blocks = split_b_blocks(ec_text)

    labels = [f"B.{i}" for i in range(1, 1 + len(ec_rows))]

    for r, lab in zip(ec_rows, labels):
        write_value_safe(ws, r, base_col, blocks.get(lab, "") or "(sin cursos declarados)")

    debug_item["ec_rows"] = ec_rows

    # EXPERIENCIA GENERAL: buscamos varias llaves posibles (para depurar)
    eg_total_key, eg_total = coalesce(payload, [
        "exp_general_total_text",
        "exp_general_total",
        "exp_general_total_texto",
        "exp_general_total_str",
    ])
    eg_detail_key, eg_detail = coalesce(payload, [
        "exp_general_detalle_text",
        "exp_general_resumen",
        "exp_general_text",
        "exp_general_detalle_text",
    ])
    # Si solo viene estructura exp_general{items,resumen,total...}
    if eg_total is None or eg_detail is None:
        eg_struct = payload.get("exp_general")
        if isinstance(eg_struct, dict):
            if eg_total is None and "total_text" in eg_struct:
                eg_total_key, eg_total = "exp_general.total_text", eg_struct.get("total_text")
            if eg_detail is None and "resumen" in eg_struct:
                eg_detail_key, eg_detail = "exp_general.resumen", eg_struct.get("resumen")

    debug_item["eg_total_key"] = eg_total_key
    debug_item["eg_detail_key"] = eg_detail_key
    debug_item["eg_total_preview"] = safe_preview(eg_total)
    debug_item["eg_detail_preview"] = safe_preview(eg_detail)

    write_value_safe(ws, lay["eg_total_row"], base_col, eg_total or "")
    write_value_safe(ws, lay["eg_detail_row"], base_col, eg_detail or "")

    # EXPERIENCIA ESPECIFICA
    ee_total_key, ee_total = coalesce(payload, [
        "exp_especifica_total_text",
        "exp_especifica_total",
        "exp_especifica_total_texto",
        "exp_especifica_total_str",
    ])
    ee_detail_key, ee_detail = coalesce(payload, [
        "exp_especifica_detalle_text",
        "exp_especifica_resumen",
        "exp_especifica_text",
        "exp_especifica_detalle_text",
    ])
    if ee_total is None or ee_detail is None:
        ee_struct = payload.get("exp_especifica")
        if isinstance(ee_struct, dict):
            if ee_total is None and "total_text" in ee_struct:
                ee_total_key, ee_total = "exp_especifica.total_text", ee_struct.get("total_text")
            if ee_detail is None and "resumen" in ee_struct:
                ee_detail_key, ee_detail = "exp_especifica.resumen", ee_struct.get("resumen")

    debug_item["ee_total_key"] = ee_total_key
    debug_item["ee_detail_key"] = ee_detail_key
    debug_item["ee_total_preview"] = safe_preview(ee_total)
    debug_item["ee_detail_preview"] = safe_preview(ee_detail)

    write_value_safe(ws, lay["ee_total_row"], base_col, ee_total or "")
    write_value_safe(ws, lay["ee_detail_row"], base_col, ee_detail or "")

import re

def split_b_blocks(text: str) -> dict:
    """
    Extrae B.1, B.2, B.3, B.4 desde estudios_complementarios_resumen.
    Soporta:
      - "B.1:" en línea sola (como tu ejemplo)
      - doble salto de línea entre bloques
      - mantiene saltos internos de cada bloque
    """
    if not text:
        return {}

    t = text.replace("\r\n", "\n").replace("\r", "\n").strip()

    # Normaliza: asegura que cada "B.x:" arranque en nueva línea
    t = re.sub(r"(?i)\n?\s*(B\.\d)\s*:\s*", r"\n\1:\n", t).strip()

    blocks = {}
    current = None
    acc = []

    for line in t.split("\n"):
        s = line.strip()

        m = re.match(r"(?i)^(B\.\d)\s*:\s*$", s)
        if m:
            if current:
                blocks[current] = "\n".join(acc).strip()
            current = m.group(1).upper()
            acc = []
            continue

        if current is not None:
            acc.append(line)  # conserva saltos reales dentro del bloque

    if current:
        blocks[current] = "\n".join(acc).strip()

    return blocks


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", required=True, help="Ruta raíz donde están los procesos (carpeta que contiene SCI N° ...)")
    ap.add_argument("--only-proc", default="", help="Nombre exacto del proceso (opcional)")
    ap.add_argument("--limit", type=int, default=0, help="Limitar postulantes (0=sin limite)")
    ap.add_argument("--debug", action="store_true", help="Imprime depuración máxima")
    args = ap.parse_args()

    root = Path(args.root)
    if not root.exists():
        raise SystemExit(f"No existe root: {root}")

    procesos = [p for p in root.iterdir() if p.is_dir()]
    procesos.sort(key=lambda p: p.name.lower())

    for proc_dir in procesos:
        if args.only_proc and proc_dir.name != args.only_proc:
            continue

        resolved = resolve_process_files(proc_dir)

        if not resolved:
            print(f"[task_40] SKIP {proc_dir.name}: no encuentro 011/{SUMMARY_NAME}")
            continue

        out_dir, summary_path, summary, layout_path, layout, out_xlsx, jsonl = resolved

        if not out_xlsx or not out_xlsx.exists():
            print(f"[task_40] SKIP {proc_dir.name}: no encuentro output_xlsx preparado (task_15)")
            continue
        if not jsonl or not jsonl.exists():
            print(f"[task_40] SKIP {proc_dir.name}: no encuentro {PARSED_JSONL_NAME}")
            continue

        lay = parse_layout_min(layout)

        rows = read_jsonl(jsonl)
        
        if args.limit and args.limit > 0:
            rows = rows[:args.limit]

        if not rows:
            print(f"[task_40] SKIP {proc_dir.name}: jsonl vacío")
            continue

        # Debug master
        debug = {
            "generated_at": ts(),
            "process_dir": str(proc_dir),
            "out_dir_011": str(out_dir),
            "summary_path": str(summary_path),
            "layout_path": str(layout_path) if layout_path else "",
            "output_xlsx": str(out_xlsx),
            "jsonl": str(jsonl),
            "layout_min": lay,
            "postulantes": len(rows),
            "items": []
        }

        wb = load_workbook(out_xlsx)

        sheet_idx = 1
        ws = get_eval_sheet(wb, lay["sheet_base"], sheet_idx)

        max_slots = detect_max_slots(ws, lay["slot_start_col"], lay["slot_step_cols"])
        if max_slots <= 0:
            raise SystemExit(f"Plantilla sin slots detectables: {out_xlsx}")

        print(f"\n[task_40] PROCESO {proc_dir.name}")
        print(f"          out_xlsx: {out_xlsx.name}")
        print(f"          layout: {layout_path.name if layout_path else '(sin layout)'}")
        print(f"          jsonl: {jsonl}")
        print(f"          postulantes: {len(rows)}")
        print(f"          sheet_base: {lay['sheet_base']}")
        print(f"          slots_por_hoja: {max_slots}")
        print(f"          slot_start_col={lay['slot_start_col']} step={lay['slot_step_cols']} header_row={lay['header_row']}")

        for idx, rec in enumerate(rows, start=1):
            payload = rec.get("_fill_payload", rec)
       
            # slot
            slot = find_next_slot(ws, max_slots, lay["header_row"], lay["slot_start_col"], lay["slot_step_cols"])
            
            if slot is None:
                sheet_idx += 1
                ws = get_eval_sheet(wb, lay["sheet_base"], sheet_idx)
                max_slots = detect_max_slots(ws, lay["slot_start_col"], lay["slot_step_cols"])
                slot = find_next_slot(ws, max_slots, lay["header_row"], lay["slot_start_col"], lay["slot_step_cols"])

            if slot is None:
                raise SystemExit("No hay slots disponibles ni en hoja nueva (revisar plantilla)")

            dbg_item = {
                "i": idx,
                "slot": slot,
                "sheet": ws.title,
                "dni": payload.get("dni", ""),
                "nombre_full": payload.get("nombre_full", payload.get("nombres", "")),
                "keys_record": sorted(list(rec.keys())),
                "keys_payload": sorted(list(payload.keys())),
            }

            # Llenado
            fill_slot(ws, slot, payload, lay, dbg_item)
            debug["items"].append(dbg_item)

            if args.debug:
                print(f"  - [{idx}] {norm(dbg_item['nombre_full'])} DNI={dbg_item['dni']} -> {ws.title} slot={slot}")
                print(f"      EG_total({dbg_item.get('eg_total_key')}): {dbg_item.get('eg_total_preview')}")
                print(f"      EG_det  ({dbg_item.get('eg_detail_key')}): {dbg_item.get('eg_detail_preview')}")
                print(f"      EE_total({dbg_item.get('ee_total_key')}): {dbg_item.get('ee_total_preview')}")
                print(f"      EE_det  ({dbg_item.get('ee_detail_key')}): {dbg_item.get('ee_detail_preview')}")
                print(f"      EC_items_key={dbg_item.get('ec_items_key')} EC_text_key={dbg_item.get('ec_text_key')}")

        # guardar
        ensure_dir(out_dir / PROCESADOS_SUBFOLDER)

        out_path = out_dir / PROCESADOS_SUBFOLDER / f"Cuadro_Evaluacion_LLENO_{proc_dir.name}.xlsx"
        wb.save(out_path)

        debug_path = out_dir / PROCESADOS_SUBFOLDER / f"task_40_debug_{proc_dir.name}.json"
        debug_path.write_text(json.dumps(debug, ensure_ascii=False, indent=2), encoding="utf-8")

        print(f"[task_40]  guardado: {out_path}")
        print(f"[task_40]  debug:    {debug_path}")

if __name__ == "__main__":
    main()
