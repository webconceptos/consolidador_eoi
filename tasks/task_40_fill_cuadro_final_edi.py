# task_40_fill_cuadro_evaluacion.py
# -*- coding: utf-8 -*-

import argparse
import json
import re
from pathlib import Path
from datetime import datetime, date, timedelta
from typing import Dict, Any, List, Tuple, Optional

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.cell.cell import MergedCell

OUT_FOLDER_NAME = "011. INSTALACIÓN DE COMITÉ"
PROCESADOS_SUBFOLDER = "procesados"

# Carpeta donde llegan EDIs
EDI_FOLDER_HINTS = [
    "09 EDI RECIBIDAS",
    "09. EDI RECIBIDAS",
    "09_EDI_RECIBIDAS",
    "09 EDI RECIBIDA",
    "09. EDI RECIBIDA",
]

SUMMARY_NAME = "init_cuadro_summary.json"
LAYOUT_NAME = "config_layout.json"
PARSED_JSONL_NAME = "parsed_postulantes.jsonl"

DEFAULT_SHEET_BASE = "Evaluación CV"
_DATE_FMT = "%d/%m/%Y"
_CAL_ANCHOR = date(2000, 1, 1)

try:
    from dateutil.relativedelta import relativedelta
except Exception:
    relativedelta = None


# -------------------------
# Utils base
# -------------------------
def ts() -> str:
    return datetime.now().isoformat(timespec="seconds")

def norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())

def ensure_dir(p: Path):
    p.mkdir(parents=True, exist_ok=True)

def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))

def read_jsonl(path: Path) -> list[dict]:
    rows = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows

def safe_sheet_name(name: str, max_len: int = 31) -> str:
    # Excel: max 31, no : \ / ? * [ ]
    name = re.sub(r"[:\\/?*\[\]]", "_", name.strip())
    name = re.sub(r"\s+", " ", name).strip()
    return name[:max_len]

def safe_preview(x, n=140):
    s = "" if x is None else str(x)
    s = s.replace("\r\n", "\n")
    return (s[:n] + "…") if len(s) > n else s


# -------------------------
# Excel write merge-safe
# -------------------------
def write_value_safe(ws, row: int, col: int, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        # escribir solo si es ancla del rango merged
        for rng in ws.merged_cells.ranges:
            if (rng.min_row <= row <= rng.max_row) and (rng.min_col <= col <= rng.max_col):
                if (row, col) == (rng.min_row, rng.min_col):
                    c = ws.cell(row=rng.min_row, column=rng.min_col)
                    c.value = value
                    c.alignment = Alignment(wrap_text=True, vertical="top")
                return
        return

    cell.value = value
    cell.alignment = Alignment(wrap_text=True, vertical="top")


# -------------------------
# Slots detect
# -------------------------
def detect_max_slots(ws, slot_start_col: int, slot_step_cols: int) -> int:
    max_col = ws.max_column
    if max_col < slot_start_col:
        return 0
    return ((max_col - slot_start_col) // slot_step_cols) + 1

def find_next_slot(ws, max_slots: int, header_row: int, slot_start_col: int, slot_step_cols: int):
    for i in range(max_slots):
        base_col = slot_start_col + i * slot_step_cols
        c = ws.cell(row=header_row, column=base_col)
        if isinstance(c, MergedCell):
            continue
        v = c.value
        if v is None or str(v).strip() == "":
            return i
        s = str(v).strip().upper()
        if "POSTULANTE" in s or "NOMBRE DEL CONSULTOR" in s:
            return i
    return None

def get_eval_sheet(wb, base_name: str, idx: int):
    if idx == 1:
        if base_name in wb.sheetnames:
            return wb[base_name]
        return wb[wb.sheetnames[0]]

    title = f"{base_name} ({idx})"
    if title in wb.sheetnames:
        return wb[title]

    base = wb[base_name] if base_name in wb.sheetnames else wb[wb.sheetnames[0]]
    ws = wb.copy_worksheet(base)
    ws.title = title
    return ws


# -------------------------
# Layout resolver
# -------------------------
def resolve_process_files(proc_dir: Path):
    out_dir = proc_dir / OUT_FOLDER_NAME
    if not out_dir.exists():
        return None

    summary_path = out_dir / SUMMARY_NAME
    if not summary_path.exists():
        return None

    summary = read_json(summary_path)

    layout_path = Path(summary.get("paths", {}).get("layout_file", "")) if summary.get("paths") else None
    if not layout_path or not layout_path.exists():
        layout_path = out_dir / LAYOUT_NAME
        if not layout_path.exists():
            layout_path = None

    layout = read_json(layout_path) if layout_path else {}

    out_xlsx = Path(summary.get("paths", {}).get("output_xlsx", "")) if summary.get("paths") else None
    if not out_xlsx or not out_xlsx.exists():
        cands = sorted(
            [p for p in out_dir.glob("Cuadro_Evaluacion*.xlsx") if p.is_file() and not p.name.startswith("~$")],
            key=lambda p: p.stat().st_mtime, reverse=True
        )
        out_xlsx = cands[0] if cands else None

    jsonl = out_dir / PROCESADOS_SUBFOLDER / PARSED_JSONL_NAME
    if not jsonl.exists():
        jsonl = proc_dir / PROCESADOS_SUBFOLDER / PARSED_JSONL_NAME
    if not jsonl.exists():
        cands = sorted(proc_dir.rglob(PARSED_JSONL_NAME), key=lambda p: p.stat().st_mtime, reverse=True)
        jsonl = cands[0] if cands else None

    return out_dir, summary_path, summary, layout_path, layout, out_xlsx, jsonl

def parse_layout_min(layout: dict):
    tl = layout.get("template_layout") if isinstance(layout.get("template_layout"), dict) else layout

    sheet_base = tl.get("sheet_base") or tl.get("sheet_name") or DEFAULT_SHEET_BASE
    slot_start_col = int(tl.get("slot_start_col", 6))
    slot_step_cols = int(tl.get("slot_step_cols", tl.get("slot_step", 2)))
    header_row = int(tl.get("header_row", 3))

    sr = tl.get("section_rows", {}) if isinstance(tl.get("section_rows", {}), dict) else {}

    fa_row = int(sr.get("fa_row", tl.get("fa_row", 6)))

    # EC rows
    ec_rows = None
    if isinstance(sr.get("ec_rows"), list) and sr["ec_rows"]:
        ec_rows = [int(x) for x in sr["ec_rows"]]
    else:
        # fallback a tu clásico
        ec_base = int(sr.get("ec_row_base", tl.get("ec_row_base", 8)))
        # si no hay count real, usamos 4 por defecto
        ec_count = int(sr.get("ec_row_count", tl.get("ec_row_count", 4)) or 4)
        ec_rows = list(range(ec_base, ec_base + ec_count))

    eg = sr.get("exp_general", {}) if isinstance(sr.get("exp_general", {}), dict) else {}
    ee = sr.get("exp_especifica", {}) if isinstance(sr.get("exp_especifica", {}), dict) else {}

    # Importante: en tu plantilla, a menudo EG y EE están mergeadas (total+detalle juntos)
    eg_anchor_row = int(eg.get("summary_row", tl.get("eg_anchor_row", 14)))  # ancla
    ee_anchor_row = int(ee.get("summary_row", tl.get("ee_anchor_row", 18)))  # ancla

    # Si tu layout trae filas explícitas, úsalo:
    eg_total_row = int(eg.get("total_row", eg_anchor_row))
    eg_detail_row = int(eg.get("summary_row", eg_anchor_row + 1))
    ee_total_row = int(ee.get("total_row", ee_anchor_row))
    ee_detail_row = int(ee.get("summary_row", ee_anchor_row + 1))

    return {
        "sheet_base": sheet_base,
        "slot_start_col": slot_start_col,
        "slot_step_cols": slot_step_cols,
        "header_row": header_row,
        "fa_row": fa_row,
        "ec_rows": ec_rows,
        "eg_total_row": eg_total_row,
        "eg_detail_row": eg_detail_row,
        "ee_total_row": ee_total_row,
        "ee_detail_row": ee_detail_row,
    }


# -------------------------
# EDI: localizar archivos recibidos
# -------------------------
def find_edi_dir(proc_dir: Path) -> Optional[Path]:
    # buscar por nombres típicos
    for name in EDI_FOLDER_HINTS:
        p = proc_dir / name
        if p.exists() and p.is_dir():
            return p
    # fallback: buscar carpeta que contenga "EDI" y "RECIB" en nombre
    for p in proc_dir.iterdir():
        if p.is_dir():
            up = p.name.upper()
            if "EDI" in up and ("RECIB" in up or "RECIBID" in up):
                return p
    return None

def guess_edi_file_for_postulante(edi_dir: Path, postulante: dict) -> Tuple[Optional[Path], str]:
    """
    Retorna (path, kind) donde kind: 'excel'|'pdf'|'none'
    Estrategia:
      1) Si JSON trae 'source_file' o 'source_path', usar eso si existe en EDI dir.
      2) Buscar por DNI en nombre.
      3) Buscar por apellido/nombre tokens (si existe) de forma suave.
    """
    dni = norm(postulante.get("dni", "")).replace(" ", "")
    nombre = norm(postulante.get("nombre_full", postulante.get("nombres", "")))

    # 1) source hints
    for k in ["source_path", "source_file", "original_file", "file_name", "filename"]:
        v = postulante.get(k)
        if v:
            cand = (edi_dir / Path(v).name)
            if cand.exists():
                ext = cand.suffix.lower()
                if ext in (".xlsx", ".xlsm", ".xls"):
                    return cand, "excel"
                if ext == ".pdf":
                    return cand, "pdf"

    # 2) buscar por DNI
    if dni:
        excel_cands = []
        pdf_cands = []
        for p in edi_dir.rglob("*"):
            if not p.is_file() or p.name.startswith("~$"):
                continue
            if dni in re.sub(r"\D", "", p.stem):  # stem numeric match
                ext = p.suffix.lower()
                if ext in (".xlsx", ".xlsm", ".xls"):
                    excel_cands.append(p)
                elif ext == ".pdf":
                    pdf_cands.append(p)
        if excel_cands:
            excel_cands.sort(key=lambda x: x.stat().st_mtime, reverse=True)
            return excel_cands[0], "excel"
        if pdf_cands:
            pdf_cands.sort(key=lambda x: x.stat().st_mtime, reverse=True)
            return pdf_cands[0], "pdf"

    # 3) buscar por tokens de nombre (suave)
    tokens = [t for t in re.split(r"\s+", nombre.upper()) if len(t) >= 4]
    if tokens:
        excel_cands = []
        pdf_cands = []
        for p in edi_dir.rglob("*"):
            if not p.is_file() or p.name.startswith("~$"):
                continue
            stem_up = p.stem.upper()
            hits = sum(1 for t in tokens[:3] if t in stem_up)  # máximo 3 tokens
            if hits >= 2:
                ext = p.suffix.lower()
                if ext in (".xlsx", ".xlsm", ".xls"):
                    excel_cands.append((hits, p))
                elif ext == ".pdf":
                    pdf_cands.append((hits, p))
        if excel_cands:
            excel_cands.sort(key=lambda x: (x[0], x[1].stat().st_mtime), reverse=True)
            return excel_cands[0][1], "excel"
        if pdf_cands:
            pdf_cands.sort(key=lambda x: (x[0], x[1].stat().st_mtime), reverse=True)
            return pdf_cands[0][1], "pdf"

    return None, "none"


# -------------------------
# Copiar hojas entre workbooks (openpyxl)
# -------------------------
def copy_sheet_to_wb(src_ws, dst_wb: Workbook, new_title: str):
    new_title = safe_sheet_name(new_title)
    base_title = new_title
    k = 2
    while new_title in dst_wb.sheetnames:
        new_title = safe_sheet_name(f"{base_title}_{k}")
        k += 1

    dst_ws = dst_wb.create_sheet(title=new_title)

    # merges
    for rng in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(rng))

    # freeze panes
    dst_ws.freeze_panes = src_ws.freeze_panes

    # copiar celdas (valores + estilos)
    for row in src_ws.iter_rows():
        for cell in row:
            dst_cell = dst_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)
            if cell.has_style:
                dst_cell._style = cell._style
                dst_cell.font = cell.font
                dst_cell.border = cell.border
                dst_cell.fill = cell.fill
                dst_cell.number_format = cell.number_format
                dst_cell.protection = cell.protection
                dst_cell.alignment = cell.alignment
            if cell.comment:
                dst_cell.comment = cell.comment
            if cell.hyperlink:
                dst_cell.hyperlink = cell.hyperlink

    # dimensiones
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
        dst_ws.column_dimensions[col_letter].hidden = dim.hidden

    for row_idx, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_idx].height = dim.height
        dst_ws.row_dimensions[row_idx].hidden = dim.hidden

    dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    dst_ws.sheet_view.zoomScale = src_ws.sheet_view.zoomScale

    return dst_ws


# -------------------------
# EC split
# -------------------------
def split_b_blocks(text: str) -> dict:
    if not text:
        return {}
    t = text.replace("\r\n", "\n").replace("\r", "\n").strip()
    t = re.sub(r"(?i)\n?\s*(B\.\d)\s*:\s*", r"\n\1:\n", t).strip()

    blocks = {}
    current = None
    acc = []
    for line in t.split("\n"):
        s = line.strip()
        m = re.match(r"(?i)^(B\.\d)\s*:\s*$", s)
        if m:
            if current:
                blocks[current] = "\n".join(acc).strip()
            current = m.group(1).upper()
            acc = []
            continue
        if current is not None:
            acc.append(line)
    if current:
        blocks[current] = "\n".join(acc).strip()
    return blocks


# -------------------------
# Fill slot (con enumeración + experiencia merge-safe)
# -------------------------
def fill_slot(ws, slot_idx: int, payload: dict, lay: dict, postulante_n: int, debug_item: dict):
    base_col = lay["slot_start_col"] + slot_idx * lay["slot_step_cols"]

    # HEADER (enumerado)
    nombre = norm(payload.get("nombre_full", "")) or norm(payload.get("nombres", ""))
    dni = norm(payload.get("dni", ""))
    pref = f"[{postulante_n:03d}] "
    header = pref + (nombre if not dni else f"{nombre}\nDNI: {dni}")
    write_value_safe(ws, lay["header_row"], base_col, header)

    # Formación
    write_value_safe(ws, lay["fa_row"], base_col, payload.get("formacion_obligatoria_resumen", "") or "")

    # Estudios complementarios
    ec_rows = lay["ec_rows"] if isinstance(lay["ec_rows"], list) else []
    ec_text = payload.get("estudios_complementarios_resumen", "") or ""
    blocks = split_b_blocks(ec_text)
    labels = [f"B.{i}" for i in range(1, 1 + len(ec_rows))]
    for r, lab in zip(ec_rows, labels):
        write_value_safe(ws, r, base_col, blocks.get(lab, "") or "(sin cursos declarados)")

    # Experiencia: en tu data nueva, normalmente viene como:
    # exp_general_total_text + exp_general_resumen_text
    # exp_especifica_total_text + exp_especifica_resumen_text
    eg_total = payload.get("exp_general_total_text") or ""
    eg_det   = payload.get("exp_general_resumen_text") or ""
    ee_total = payload.get("exp_especifica_total_text") or ""
    ee_det   = payload.get("exp_especifica_resumen_text") or ""

    # Si en la plantilla EG está mergeada (total+detalle en la celda ancla), escribimos todo en la fila total
    eg_text = "\n".join([x for x in [eg_total, eg_det] if norm(x)]).strip()
    ee_text = "\n".join([x for x in [ee_total, ee_det] if norm(x)]).strip()

    write_value_safe(ws, lay["eg_total_row"], base_col, eg_text)
    # NO escribir en eg_detail_row si es merged
    write_value_safe(ws, lay["ee_total_row"], base_col, ee_text)
    # NO escribir en ee_detail_row si es merged

    debug_item["eg_total_preview"] = safe_preview(eg_total)
    debug_item["eg_det_preview"]   = safe_preview(eg_det)
    debug_item["ee_total_preview"] = safe_preview(ee_total)
    debug_item["ee_det_preview"]   = safe_preview(ee_det)


# -------------------------
# MAIN
# -------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", required=True, help="Ruta raíz: 'Procesos de Selección'")
    ap.add_argument("--only-proc", default="", help="Proceso exacto (opcional)")
    ap.add_argument("--limit", type=int, default=0, help="Limitar postulantes (0=sin limite)")
    ap.add_argument("--debug", action="store_true", help="Depuración")
    ap.add_argument("--copy-edi", action="store_true", help="Copia EDI al consolidado (recomendado: ON)")
    args = ap.parse_args()

    root = Path(args.root)
    if not root.exists():
        raise SystemExit(f"No existe root: {root}")

    procesos = [p for p in root.iterdir() if p.is_dir()]
    procesos.sort(key=lambda p: p.name.lower())

    for proc_dir in procesos:
        if args.only_proc and proc_dir.name != args.only_proc:
            continue

        resolved = resolve_process_files(proc_dir)
        if not resolved:
            print(f"[task_40] SKIP {proc_dir.name}: no encuentro 011/{SUMMARY_NAME}")
            continue

        out_dir, summary_path, summary, layout_path, layout, out_xlsx, jsonl = resolved

        if not out_xlsx or not out_xlsx.exists():
            print(f"[task_40] SKIP {proc_dir.name}: no encuentro output_xlsx preparado (task_15)")
            continue
        if not jsonl or not jsonl.exists():
            print(f"[task_40] SKIP {proc_dir.name}: no encuentro {PARSED_JSONL_NAME}")
            continue

        lay = parse_layout_min(layout)

        rows = read_jsonl(jsonl)
        if args.limit and args.limit > 0:
            rows = rows[:args.limit]
        if not rows:
            print(f"[task_40] SKIP {proc_dir.name}: jsonl vacío")
            continue

        # EDI directory
        edi_dir = find_edi_dir(proc_dir)

        wb = load_workbook(out_xlsx)
        sheet_idx = 1
        ws = get_eval_sheet(wb, lay["sheet_base"], sheet_idx)

        max_slots = detect_max_slots(ws, lay["slot_start_col"], lay["slot_step_cols"])
        if max_slots <= 0:
            raise SystemExit(f"Plantilla sin slots detectables: {out_xlsx}")

        print(f"\n[task_40] PROCESO {proc_dir.name}")
        print(f"          base_xlsx: {out_xlsx.name}")
        print(f"          jsonl: {jsonl}")
        print(f"          postulantes: {len(rows)}")
        print(f"          sheet_base: {lay['sheet_base']}")
        print(f"          slots_por_hoja: {max_slots}")
        print(f"          edi_dir: {edi_dir if edi_dir else '(no encontrada)'}")

        debug = {
            "generated_at": ts(),
            "process_dir": str(proc_dir),
            "output_base_xlsx": str(out_xlsx),
            "jsonl": str(jsonl),
            "edi_dir": str(edi_dir) if edi_dir else "",
            "postulantes": len(rows),
            "items": []
        }

        # 1) Llenar Evaluación por slots
        for n, rec in enumerate(rows, start=1):
            payload = rec.get("_fill_payload", rec)

            slot = find_next_slot(ws, max_slots, lay["header_row"], lay["slot_start_col"], lay["slot_step_cols"])
            if slot is None:
                sheet_idx += 1
                ws = get_eval_sheet(wb, lay["sheet_base"], sheet_idx)
                max_slots = detect_max_slots(ws, lay["slot_start_col"], lay["slot_step_cols"])
                slot = find_next_slot(ws, max_slots, lay["header_row"], lay["slot_start_col"], lay["slot_step_cols"])

            if slot is None:
                raise SystemExit("No hay slots disponibles ni en hoja nueva (revisar plantilla)")

            dbg_item = {
                "n": n,
                "slot": slot,
                "sheet_eval": ws.title,
                "dni": payload.get("dni", ""),
                "nombre_full": payload.get("nombre_full", payload.get("nombres", "")),
                "edi_match": None,
                "edi_kind": None,
            }

            fill_slot(ws, slot, payload, lay, postulante_n=n, debug_item=dbg_item)
            debug["items"].append(dbg_item)

            if args.debug:
                print(f"  - [{n:03d}] {norm(dbg_item['nombre_full'])} DNI={dbg_item['dni']} -> {ws.title} slot={slot}")

        # 2) Copiar EDI como hojas nuevas enumeradas 001..N (si copy-edi)
        if args.copy_edi:
            for n, rec in enumerate(rows, start=1):
                payload = rec.get("_fill_payload", rec)
                dni = norm(payload.get("dni", ""))

                edi_path = None
                edi_kind = "none"
                if edi_dir:
                    edi_path, edi_kind = guess_edi_file_for_postulante(edi_dir, payload)

                # registrar en debug
                debug["items"][n-1]["edi_match"] = str(edi_path) if edi_path else ""
                debug["items"][n-1]["edi_kind"] = edi_kind

                sheet_name = f"{n:03d}"  # <- requerido por ti

                if edi_kind == "excel" and edi_path and edi_path.exists():
                    try:
                        edi_wb = load_workbook(edi_path, data_only=False)
                        # estrategia: copiar la primera hoja (o la más relevante)
                        # si existe alguna hoja con 'expres' o 'interes', preferirla
                        best = None
                        for s in edi_wb.sheetnames:
                            print(s)
                            exit()
                            up = s.upper()
                            if "EXP" in up and ("INTER" in up or "INT" in up):
                                best = s
                                break
                        src_ws = edi_wb[best] if best else edi_wb[edi_wb.sheetnames[0]]

                        copy_sheet_to_wb(src_ws, wb, sheet_name)
                        edi_wb.close()
                    except Exception as e:
                        # si falla copia, crear hoja placeholder
                        ws_edi = wb.create_sheet(title=safe_sheet_name(sheet_name))
                        ws_edi["A1"] = f"EDI [{n:03d}] - ERROR copiando Excel: {edi_path.name}"
                        ws_edi["A2"] = str(e)

                elif edi_kind == "pdf":
                    # hoja vacía placeholder
                    ws_edi = wb.create_sheet(title=safe_sheet_name(sheet_name))
                    ws_edi["A1"] = f"EDI [{n:03d}] - El postulante envió PDF (pendiente de extracción)"
                    if edi_path:
                        ws_edi["A2"] = f"Archivo: {edi_path.name}"
                    if dni:
                        ws_edi["A3"] = f"DNI: {dni}"

                else:
                    # no encontrado: hoja vacía placeholder
                    ws_edi = wb.create_sheet(title=safe_sheet_name(sheet_name))
                    ws_edi["A1"] = f"EDI [{n:03d}] - No se encontró Excel de EDI en '09 EDI RECIBIDAS'"
                    if dni:
                        ws_edi["A2"] = f"DNI: {dni}"

        # Guardar consolidado
        ensure_dir(out_dir / PROCESADOS_SUBFOLDER)
        out_path = out_dir / PROCESADOS_SUBFOLDER / f"Cuadro_Evaluacion_CONSOLIDADO_{proc_dir.name}.xlsx"
        wb.save(out_path)

        debug_path = out_dir / PROCESADOS_SUBFOLDER / f"task_40_debug_{proc_dir.name}.json"
        debug_path.write_text(json.dumps(debug, ensure_ascii=False, indent=2), encoding="utf-8")

        print(f"[task_40] ✅ guardado: {out_path}")
        print(f"[task_40] 🧾 debug:   {debug_path}")


if __name__ == "__main__":
    main()
