# tasks/task_40_fill_output.py
# -*- coding: utf-8 -*-

import argparse
import csv
import json
import shutil
import sys
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------
# Fix imports when running from /tasks (Windows common issue)
# ------------------------------------------------------------
REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

# ------------------------------------------------------------
# Folders
# ------------------------------------------------------------
IN_FOLDER_CANDIDATES = ["009. EDI RECIBIDAS", "009. EDI RECIBIDA"]
OUT_FOLDER_NAME = "011. INSTALACIÓN DE COMITÉ"
PROCESADOS_SUBFOLDER = "procesados"

PROCESS_OUTPUT_TEMPLATE_PREFIX = "Revision Preliminar"
EVAL_SHEET_TEMPLATE = "Evaluación CV"

CONSOLIDADO_JSONL = "consolidado.jsonl"

# ------------------------------------------------------------
# Helpers
# ------------------------------------------------------------
def ts() -> str:
    return datetime.now().isoformat(timespec="seconds")

def norm(s: str) -> str:
    return " ".join((s or "").strip().split())

def ensure_dir(p: Path):
    p.mkdir(parents=True, exist_ok=True)

def safe_filename(s: str, max_len: int = 140) -> str:
    s = norm(s)
    out = []
    for ch in s:
        out.append(ch if ch.isalnum() or ch in "._- " else "_")
    s2 = "".join(out).replace(" ", "_")
    return s2[:max_len]

def write_csv(path: Path, header, rows):
    ensure_dir(path.parent)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)

def log_append(path: Path, msg: str, also_print: bool = True):
    ensure_dir(path.parent)
    line = f"[{ts()}] {msg}"
    with path.open("a", encoding="utf-8") as f:
        f.write(line + "\n")
    if also_print:
        print(line)

def safe_save_workbook(wb, out_xlsx: Path, debug_log: Path = None):
    try:
        wb.save(out_xlsx)
        if debug_log:
            log_append(debug_log, f"[SAVE] OK -> {out_xlsx}", also_print=True)
        return out_xlsx
    except PermissionError:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_alt = out_xlsx.with_name(f"{out_xlsx.stem}_{stamp}{out_xlsx.suffix}")
        wb.save(out_alt)
        if debug_log:
            log_append(debug_log, f"[SAVE] BLOQUEADO -> guardado alternativo: {out_alt}", also_print=True)
        return out_alt

def resolve_in_dir(proc_dir: Path) -> Path:
    for name in IN_FOLDER_CANDIDATES:
        p = proc_dir / name
        if p.exists():
            return p
    return proc_dir / IN_FOLDER_CANDIDATES[0]


LAYOUT_NAME = "config_layout.json"

def load_layout_for_process(out_dir: Path) -> dict:
    p = out_dir / LAYOUT_NAME
    if not p.exists():
        return {}
    return json.loads(p.read_text(encoding="utf-8"))

def apply_layout_over_cfg(cfg: dict, layout: dict) -> dict:
    """
    Convierte layout->cfg operativo:
    - header_row
    - slot_start_col
    - slot_step_cols
    - max_postulantes_por_hoja (detectado)
    - fa_row
    - ec_rows (desde ec_row_base)
    - exp_general_row (si quieres un único row resumen; si no, lo manejamos como rango)
    """
    cfg2 = dict(cfg)

    section = (layout or {}).get("section_rows", {}) or {}
    header_row = (layout or {}).get("header_row", None)
    slot_start_col = (layout or {}).get("slot_start_col", None)
    slot_step = (layout or {}).get("slot_step", None)
    max_slots = (layout or {}).get("max_postulantes_por_hoja_detectado", None)

    if header_row: cfg2["slot_header_row"] = int(header_row)
    if slot_start_col: cfg2["slot_start_col"] = int(slot_start_col)
    if slot_step: cfg2["slot_step_cols"] = int(slot_step)
    if max_slots: cfg2["max_postulantes_por_hoja"] = int(max_slots)

    # Formación Académica
    if "fa_row" in section:
        cfg2["fa_row"] = int(section["fa_row"])

    # Estudios complementarios: base b.1
    # Tu scan guarda: ec_row_base (ojo: en tu layout se llama ec_row_base o ec_row_base??)
    # En tu código: section_rows["ec_row_base"] (pero lo guardas como "ec_row_base" o "ec_row_base"?)
    # Veo: section_rows["ec_row_base"] = ...
    # entonces:
    if "ec_row_base" in section:
        base = int(section["ec_row_base"])
        # Asumimos b.1..b.4 (ajustable)
        cfg2["ec_rows"] = [base, base+1, base+2, base+3]

    # Experiencia General (si quieres un rango)
    if "exp_general_start_row" in section:
        cfg2["exp_general_start_row"] = int(section["exp_general_start_row"])
    if "exp_general_end_row" in section:
        cfg2["exp_general_end_row"] = int(section["exp_general_end_row"])

    # Experiencia Específica (si te interesa)
    if "exp_especifica_start_row" in section:
        cfg2["exp_especifica_start_row"] = int(section["exp_especifica_start_row"])
    if "exp_especifica_end_row" in section:
        cfg2["exp_especifica_end_row"] = int(section["exp_especifica_end_row"])

    return cfg2

def load_layout(out_dir: Path) -> dict:
    p = out_dir / "config_layout.json"
    if not p.exists():
        return {}
    return json.loads(p.read_text(encoding="utf-8"))

def get_template_layout(layout: dict) -> dict:
    """
    Compatibilidad:
    - Layout nuevo: layout["template_layout"] contiene el layout real de la plantilla (011)
    - Layout viejo: layout ya venía plano (sheet_base, section_rows, etc.)
    """
    if "template_layout" in layout and isinstance(layout["template_layout"], dict):
        return layout["template_layout"]
    return layout  # fallback al formato antiguo

def get_section_rows(tpl: dict) -> dict:
    return tpl.get("section_rows", {}) or {}



# ------------------------------------------------------------
# Slot logic (template)
# ------------------------------------------------------------
def slot_columns(slot_index: int, slot_start_col: int = 6, slot_step_cols: int = 2):
    base_col = slot_start_col + slot_index * slot_step_cols
    score_col = base_col + 1
    return base_col, score_col

def next_free_slot(ws, max_slots: int, slot_start_col: int = 6, slot_step_cols: int = 2, header_row: int = 3):
    for i in range(max_slots):
        base_col, _ = slot_columns(i, slot_start_col, slot_step_cols)
        v = ws.cell(row=header_row, column=base_col).value
        if v is None or str(v).strip() == "" or "NOMBRE DEL CONSULTOR" in str(v).upper():
            return i
    return None

def get_or_create_eval_sheet(wb, template_name: str, sheet_index: int):
    if sheet_index == 1:
        return wb[template_name]
    title = f"{template_name} ({sheet_index})"
    if title in wb.sheetnames:
        return wb[title]
    ws = wb.copy_worksheet(wb[template_name])
    ws.title = title
    return ws

# ------------------------------------------------------------
# Styles (same idea as your old task_30)
# ------------------------------------------------------------
def build_slot_styles(cfg: dict):
    slot_cfg = cfg.get("slot_style", {})

    fill_base = PatternFill("solid", fgColor=slot_cfg.get("fill_base", "FFF2CC"))
    fill_score = PatternFill("solid", fgColor=slot_cfg.get("fill_score", "D9E1F2"))

    font_base = Font(
        name=slot_cfg.get("font_name", "Calibri"),
        size=int(slot_cfg.get("font_size", 10)),
        bold=bool(slot_cfg.get("font_bold", False)),
    )
    font_header = Font(
        name=slot_cfg.get("font_name", "Calibri"),
        size=int(slot_cfg.get("header_font_size", 11)),
        bold=True,
    )

    align = Alignment(horizontal=slot_cfg.get("align_h", "left"),
                      vertical=slot_cfg.get("align_v", "top"),
                      wrap_text=True)

    thin = Side(style="thin", color=slot_cfg.get("border_color", "000000"))
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "fill_base": fill_base,
        "fill_score": fill_score,
        "font_base": font_base,
        "font_header": font_header,
        "align": align,
        "border": border
    }

def apply_slot_format(ws, base_col: int, score_col: int, cfg: dict, styles: dict):
    slot_cfg = cfg.get("slot_style", {})
    start_row = int(slot_cfg.get("start_row", 3))
    end_row = int(slot_cfg.get("end_row", 22))

    base_width = float(slot_cfg.get("base_col_width", 34))
    score_width = float(slot_cfg.get("score_col_width", 14))

    ws.column_dimensions[get_column_letter(base_col)].width = base_width
    ws.column_dimensions[get_column_letter(score_col)].width = score_width

    if bool(slot_cfg.get("set_row_heights", True)):
        ws.row_dimensions[3].height = float(slot_cfg.get("row3_height", 18))
        ws.row_dimensions[6].height = float(slot_cfg.get("row6_height", 34))

    for r in range(start_row, end_row + 1):
        cb = ws.cell(row=r, column=base_col)
        cs = ws.cell(row=r, column=score_col)

        cb.fill = styles["fill_base"]
        cb.font = styles["font_base"]
        cb.alignment = styles["align"]
        cb.border = styles["border"]

        cs.fill = styles["fill_score"]
        cs.font = styles["font_base"]
        cs.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cs.border = styles["border"]

    ws.cell(row=3, column=base_col).font = styles["font_header"]
    ws.cell(row=3, column=score_col).font = styles["font_header"]

# ------------------------------------------------------------
# Template finder
# ------------------------------------------------------------
def find_process_template(out_dir: Path) -> Path | None:
    if not out_dir.exists():
        return None

    cands = sorted(
        [p for p in out_dir.iterdir()
         if p.is_file()
         and p.suffix.lower() == ".xlsx"
         and p.name.lower().startswith(PROCESS_OUTPUT_TEMPLATE_PREFIX.lower())]
    )
    return cands[0] if cands else None

# ------------------------------------------------------------
# Merge-safe writing
# ------------------------------------------------------------
def _cell_merge_anchor(ws, row: int, col: int):
    coord = ws.cell(row=row, column=col).coordinate
    for rng in ws.merged_cells.ranges:
        if coord in rng:
            return rng.min_row, rng.min_col
    return row, col

def write_value_safe(ws, row: int, col: int, value):
    ar, ac = _cell_merge_anchor(ws, row, col)
    # Escribe SOLO en el anchor real
    ws.cell(row=ar, column=ac).value = value
    return ws.cell(row=ar, column=ac)

def clear_value_safe(ws, row: int, col: int):
    return write_value_safe(ws, row, col, None)

def apply_wrap(cell, vertical="top"):
    if cell.alignment:
        cell.alignment = cell.alignment.copy(wrap_text=True, vertical=vertical)
    else:
        cell.alignment = Alignment(wrap_text=True, vertical=vertical)

# ------------------------------------------------------------
# Fill logic (DP + FA + EC + optional EG)
# ------------------------------------------------------------
def _block_index(block_id: str) -> int:
    import re
    m = re.search(r"\b[bB]\s*\.\s*(\d+)\b", str(block_id))
    return int(m.group(1)) if m else 1

def _is_header_like_course(it: dict) -> bool:
    nro = norm(it.get("nro", "")).lower()
    centro = norm(it.get("centro", "")).lower()
    cap = norm(it.get("capacitacion", "")).lower()
    fi = norm(it.get("fecha_inicio", "")).lower()
    if nro in ("no.", "n°", "nº", "nro", "nro."):
        return True
    if centro.startswith("centro de estudios") or "nombre de la entidad" in centro:
        return True
    if cap in ("capacitacion", "capacitación", "nombre del proyecto"):
        return True
    if fi.startswith("fecha de"):
        return True
    return False

def _format_course_line(it: dict) -> str:
    centro = norm(it.get("centro", ""))
    cap = norm(it.get("capacitacion", ""))
    fi = norm(it.get("fecha_inicio", ""))
    ff = norm(it.get("fecha_fin", ""))
    horas = it.get("horas", "")

    out = []
    if centro: out.append(centro)
    if cap: out.append(cap)

    fechas = " - ".join([x for x in [fi, ff] if x]).strip()
    if fechas: out.append(fechas)

    h = str(horas).strip() if horas is not None else ""
    if h and h not in ("0", "0.0"):
        if not h.lower().endswith("h"):
            h = f"{h}h"
        out.append(h)

    return " | ".join(out).strip()

def fill_postulante(ws, data: dict, cfg: dict, styles: dict, debug_log: Path):
    """
    Llena un slot:
    - header_row: cabecera con nombre/DNI/cel/email
    - fa_row: formación resumen
    - ec_rows: filas para b.1..b.N (si no hay, usa auto-detect por texto "DETALLAR LOS CURSOS DECLARADOS")
    - eg_row: si existe en cfg, escribe experiencia general resumen
    """
    max_slots = int(cfg.get("max_postulantes_por_hoja", 20))
    slot_start_col = int(cfg.get("slot_start_col", 6))
    slot_step_cols = int(cfg.get("slot_step_cols", 2))
    header_row = int(cfg.get("slot_header_row", 3))

    fa_row = int(cfg.get("fa_row", 6))
    eg_row = int(cfg.get("exp_general_row", 0))  # 0=disabled

    ec_rows = cfg.get("ec_rows", [8, 9, 10, 11])
    if isinstance(ec_rows, str):
        ec_rows = [int(x.strip()) for x in ec_rows.split(",") if x.strip()]

    slot = next_free_slot(ws, max_slots, slot_start_col, slot_step_cols, header_row)
    if slot is None:
        return False, "NO_HAY_SLOT"

    base_col, score_col = slot_columns(slot, slot_start_col, slot_step_cols)
    apply_slot_format(ws, base_col, score_col, cfg, styles)

    # 1) Cabecera
    display_name = norm(data.get("nombre_full", "")) or norm(data.get("source_label", "")) or "SIN_NOMBRE"
    dni = norm(data.get("dni", ""))
    cel = norm(data.get("celular", ""))
    email = norm(data.get("email", ""))

    parts = [display_name]
    if dni: parts.append(f"DNI: {dni}")
    if cel: parts.append(f"Cel: {cel}")
    if email: parts.append(f"Email: {email}")
    header_text = " | ".join(parts)

    c = write_value_safe(ws, header_row, base_col, header_text)
    apply_wrap(c, vertical="center")

    # 2) Formación Académica
    fa = norm(data.get("formacion_resumen") or "")
    if not fa:
        fa = "SIN DATOS DE FORMACIÓN"
    c = write_value_safe(ws, fa_row, base_col, fa)
    apply_wrap(c, vertical="top")

    # 3) Estudios Complementarios
    ec = data.get("estudios_complementarios") or {}
    blocks = ec.get("blocks") or []

    # Limpia filas EC destino
    for rr in ec_rows:
        clear_value_safe(ws, rr, base_col)

    # Escribe b.1..b.N en ec_rows
    for b in blocks:
        bid = b.get("id", "")
        idx = _block_index(bid)
        if idx < 1:
            continue
        if idx > len(ec_rows):
            # si hay más bloques que filas, los “apila” en el último
            rr = ec_rows[-1]
        else:
            rr = ec_rows[idx - 1]

        items = b.get("items") or []
        seen = set()
        lines = []
        for it in items:
            if _is_header_like_course(it):
                continue
            if not (norm(it.get("centro", "")) or norm(it.get("capacitacion", ""))):
                continue

            key = (
                norm(it.get("centro", "")).lower(),
                norm(it.get("capacitacion", "")).lower(),
                norm(it.get("fecha_inicio", "")),
                norm(it.get("fecha_fin", "")),
                str(it.get("horas", "")).strip()
            )
            if key in seen:
                continue
            seen.add(key)
            lines.append(_format_course_line(it))

        resumen = "\n".join(lines).strip() if lines else "(sin cursos declarados)"
        c = write_value_safe(ws, rr, base_col, resumen)
        apply_wrap(c, vertical="top")

    # 4) Experiencia General (opcional)
    if eg_row and int(eg_row) > 0:
        eg = data.get("exp_general") or {}
        eg_text = norm(data.get("exp_general_resumen") or eg.get("resumen") or "")
        if not eg_text:
            eg_text = "(sin experiencia general declarada)"
        c = write_value_safe(ws, int(eg_row), base_col, eg_text)
        apply_wrap(c, vertical="top")

    log_append(debug_log, f"[FILL] sheet='{ws.title}' slot={slot} base_col={base_col} name='{display_name}' blocks={len(blocks)}", also_print=False)
    return True, f"SLOT_{slot}"

# ------------------------------------------------------------
# Read consolidado.jsonl
# ------------------------------------------------------------
def read_jsonl(path: Path):
    items = []
    if not path.exists():
        return items
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            items.append(json.loads(line))
    return items

# ------------------------------------------------------------
# Main
# ------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", required=True, help="Ruta raíz de ProcesoSelección")
    ap.add_argument("--only-proc", default="", help="Nombre exacto del proceso (opcional)")
    ap.add_argument("--limit", type=int, default=0, help="Limitar postulantes (0=sin limite)")
    args = ap.parse_args()

    cfg = json.loads((REPO_ROOT / "configs" / "config.json").read_text(encoding="utf-8"))
    root = Path(args.root)

    processed_mode = cfg.get("processed_mode", "copy")
    if processed_mode not in ("copy", "move"):
        processed_mode = "copy"

    styles = build_slot_styles(cfg)

    print(f"[task_40] root={root} processed_mode={processed_mode}")

    procesos = [p for p in root.iterdir() if p.is_dir()]
    procesos.sort(key=lambda p: p.name.lower())
    print(f"[task_40] procesos detectados={len(procesos)}")

    ok_proc = 0
    skip_proc = 0

    for proc_dir in procesos:
        proceso = proc_dir.name
        if args.only_proc and proceso != args.only_proc:
            continue

        in_dir = resolve_in_dir(proc_dir)
        out_dir = proc_dir / OUT_FOLDER_NAME

        print(f"\n[task_40] --- PROCESO: {proceso} ---")
        print(f"  in_dir = {in_dir}")
        print(f"  out_dir= {out_dir}")

        if not in_dir.exists():
            print("  [SKIP] No existe carpeta 009 (EDI)")
            skip_proc += 1
            continue
        if not out_dir.exists():
            print("  [SKIP] No existe carpeta 011. INSTALACIÓN DE COMITÉ")
            skip_proc += 1
            continue

        debug_log = out_dir / "debug_task_40.log"
        ensure_dir(out_dir / PROCESADOS_SUBFOLDER)
        log_append(debug_log, f"== TASK 40 | PROCESO: {proceso} ==")

        tpl = find_process_template(out_dir)
        if not tpl:
            print(f"  [SKIP] No encuentro plantilla '{PROCESS_OUTPUT_TEMPLATE_PREFIX}*.xlsx' en 011")
            log_append(debug_log, f"[SKIP] No encuentro plantilla en {out_dir}")
            skip_proc += 1
            continue

        consolidado_path = out_dir / CONSOLIDADO_JSONL
        postulantes = read_jsonl(consolidado_path)
        if not postulantes:
            print(f"  [SKIP] No existe o está vacío: {consolidado_path.name}")
            log_append(debug_log, f"[SKIP] No existe o vacío: {consolidado_path}")
            skip_proc += 1
            continue

        if args.limit and args.limit > 0:
            postulantes = postulantes[: args.limit]

        print(f"  template = {tpl.name}")
        print(f"  postulantes (jsonl) = {len(postulantes)}")

        wb = load_workbook(tpl)
        sheet_idx = 1
        ws = get_or_create_eval_sheet(wb, EVAL_SHEET_TEMPLATE, sheet_idx)

        rows_log = []
        rows_flat = []

        for data in postulantes:
            try:
                # etiqueta visible (carpeta postulante si task_20 la guardó)
                if "source_label" not in data:
                    data["source_label"] = Path(data.get("source_file", "")).parent.name if data.get("source_file") else ""

                ok, where = fill_postulante(ws, data, cfg, styles, debug_log)
                if not ok and where == "NO_HAY_SLOT":
                    sheet_idx += 1
                    ws = get_or_create_eval_sheet(wb, EVAL_SHEET_TEMPLATE, sheet_idx)
                    log_append(debug_log, f"[INFO] Nueva hoja: {ws.title}", also_print=False)
                    ok, where = fill_postulante(ws, data, cfg, styles, debug_log)

                if not ok:
                    rows_log.append([ts(), data.get("source_file",""), "JSONL", "ERROR", where])
                    continue

                rows_log.append([ts(), data.get("source_file",""), "JSONL", "OK", f"{ws.title}:{where}"])
                rows_flat.append([
                    proceso,
                    data.get("source_file",""),
                    ws.title,
                    where,
                    data.get("dni",""),
                    data.get("nombre_full",""),
                    data.get("celular",""),
                    data.get("email",""),
                    (data.get("estudios_complementarios") or {}).get("total_horas", 0),
                    data.get("exp_general_dias", 0),
                ])

            except Exception as e:
                rows_log.append([ts(), data.get("source_file",""), "JSONL", "ERROR", repr(e)])

        out_xlsx = out_dir / f"Cuadro_Evaluacion_{safe_filename(proceso)}.xlsx"
        out_saved = safe_save_workbook(wb, out_xlsx, debug_log)
        print(f"  ✅ guardado: {out_saved}")

        write_csv(out_dir / "task40_log.csv", ["fecha","archivo","tipo","estado","detalle"], rows_log)
        write_csv(out_dir / "task40_consolidado.csv",
                  ["proceso","archivo_origen","hoja","slot","dni","nombre","celular","email","ec_total_horas","exp_general_dias"],
                  rows_flat)

        log_append(debug_log, f"[DONE] Guardado: {out_saved}")
        ok_proc += 1

    print(f"\n[task_40] OK_PROCESOS={ok_proc} SKIP_PROCESOS={skip_proc}")

if __name__ == "__main__":
    main()
