# tasks/task_30_fill_output.py
# -*- coding: utf-8 -*-

import argparse
import csv
import json
import shutil
from pathlib import Path
from openpyxl.cell.cell import MergedCell

from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

from parsers.eoi_excel import parse_eoi_excel
from parsers.eoi_pdf import parse_eoi_pdf


IN_FOLDER_NAME = "009. EDI RECIBIDAS"
OUT_FOLDER_NAME = "011. INSTALACIÓN DE COMITÉ"
PROCESADOS_SUBFOLDER = "procesados"

PROCESS_OUTPUT_TEMPLATE_PREFIX = "Revision Preliminar"
EVAL_SHEET_TEMPLATE = "Evaluación CV"


# -------------------------
# Helpers
# -------------------------
def ts() -> str:
    return datetime.now().isoformat(timespec="seconds")

def norm(s: str) -> str:
    return " ".join((s or "").strip().split())

def ensure_dir(p: Path):
    p.mkdir(parents=True, exist_ok=True)

def safe_filename(s: str, max_len: int = 140) -> str:
    s = norm(s)
    out = []
    for ch in s:
        out.append(ch if ch.isalnum() or ch in "._- " else "_")
    s2 = "".join(out).replace(" ", "_")
    return s2[:max_len]

def write_csv(path: Path, header, rows):
    ensure_dir(path.parent)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)

def log_append(path: Path, msg: str, also_print: bool = True):
    ensure_dir(path.parent)
    line = f"[{ts()}] {msg}"
    with path.open("a", encoding="utf-8") as f:
        f.write(line + "\n")
    if also_print:
        print(line)

def safe_save_workbook(wb, out_xlsx: Path, debug_log: Path = None):
    try:
        wb.save(out_xlsx)
        if debug_log:
            log_append(debug_log, f"[SAVE] OK -> {out_xlsx}", also_print=True)
        return out_xlsx
    except PermissionError:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_alt = out_xlsx.with_name(f"{out_xlsx.stem}_{stamp}{out_xlsx.suffix}")
        wb.save(out_alt)
        if debug_log:
            log_append(debug_log, f"[SAVE] BLOQUEADO -> guardado alternativo: {out_alt}", also_print=True)
        return out_alt



# -------------------------
# Slot logic
# -------------------------
def slot_columns(slot_index: int, slot_start_col: int = 6, slot_step_cols: int = 2):
    base_col = slot_start_col + slot_index * slot_step_cols
    score_col = base_col + 1
    return base_col, score_col

def next_free_slot(ws, max_slots: int, slot_start_col: int = 6, slot_step_cols: int = 2, header_row: int = 3):
    for i in range(max_slots):
        base_col, _ = slot_columns(i, slot_start_col, slot_step_cols)
        v = ws.cell(row=header_row, column=base_col).value
        if v is None or str(v).strip() == "" or "NOMBRE DEL CONSULTOR" in str(v).upper():
            return i
    return None

def get_or_create_eval_sheet(wb, template_name: str, sheet_index: int):
    if sheet_index == 1:
        return wb[template_name]
    title = f"{template_name} ({sheet_index})"
    if title in wb.sheetnames:
        return wb[title]
    ws = wb.copy_worksheet(wb[template_name])
    ws.title = title
    return ws


# -------------------------
# Styles
# -------------------------
def build_slot_styles(cfg: dict):
    slot_cfg = cfg.get("slot_style", {})

    fill_base = PatternFill("solid", fgColor=slot_cfg.get("fill_base", "FFF2CC"))
    fill_score = PatternFill("solid", fgColor=slot_cfg.get("fill_score", "D9E1F2"))

    font_base = Font(
        name=slot_cfg.get("font_name", "Calibri"),
        size=int(slot_cfg.get("font_size", 10)),
        bold=bool(slot_cfg.get("font_bold", False)),
    )
    font_header = Font(
        name=slot_cfg.get("font_name", "Calibri"),
        size=int(slot_cfg.get("header_font_size", 11)),
        bold=True,
    )

    align = Alignment(horizontal=slot_cfg.get("align_h", "left"),
                      vertical=slot_cfg.get("align_v", "top"),
                      wrap_text=True)

    thin = Side(style="thin", color=slot_cfg.get("border_color", "000000"))
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {"fill_base": fill_base, "fill_score": fill_score, "font_base": font_base,
            "font_header": font_header, "align": align, "border": border}

def apply_slot_format(ws, base_col: int, score_col: int, cfg: dict, styles: dict):
    slot_cfg = cfg.get("slot_style", {})
    start_row = int(slot_cfg.get("start_row", 3))
    end_row = int(slot_cfg.get("end_row", 22))

    base_width = float(slot_cfg.get("base_col_width", 34))
    score_width = float(slot_cfg.get("score_col_width", 14))

    ws.column_dimensions[get_column_letter(base_col)].width = base_width
    ws.column_dimensions[get_column_letter(score_col)].width = score_width

    if bool(slot_cfg.get("set_row_heights", True)):
        ws.row_dimensions[3].height = float(slot_cfg.get("row3_height", 18))
        ws.row_dimensions[6].height = float(slot_cfg.get("row6_height", 34))

    for r in range(start_row, end_row + 1):
        cb = ws.cell(row=r, column=base_col)
        cs = ws.cell(row=r, column=score_col)

        cb.fill = styles["fill_base"]
        cb.font = styles["font_base"]
        cb.alignment = styles["align"]
        cb.border = styles["border"]

        cs.fill = styles["fill_score"]
        cs.font = styles["font_base"]
        cs.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cs.border = styles["border"]

    ws.cell(row=3, column=base_col).font = styles["font_header"]
    ws.cell(row=3, column=score_col).font = styles["font_header"]

# -------------------------
# Template finder
# -------------------------
def find_process_template(out_dir: Path) -> Path | None:
    if not out_dir.exists():
        return None

    cands = sorted(
        [p for p in out_dir.iterdir()
         if p.is_file()
         and p.suffix.lower() == ".xlsx"
         and p.name.lower().startswith(PROCESS_OUTPUT_TEMPLATE_PREFIX.lower())]
    )
    return cands[0] if cands else None

# -------------------------
# Choose one file per postulante folder (Excel preferred)
# -------------------------
def choose_one_file_per_postulante_folder(in_dir: Path):
    if not in_dir.exists():
        return []

    by_folder = {}
    for p in in_dir.rglob("*"):
        if not p.is_file():
            continue
        if p.name.startswith("~$"):
            continue
        ext = p.suffix.lower()
        if ext not in (".xlsx", ".xlsm", ".xls", ".pdf"):
            continue
        by_folder.setdefault(p.parent, []).append(p)

    def score_excel(f: Path):
        name = f.name.lower()
        ext = f.suffix.lower()
        ext_score = {".xlsx": 30, ".xlsm": 20, ".xls": 10}.get(ext, 0)
        bonus = 0
        if any(k in name for k in ("formatocv", "formato", "cv", "edi")):
            bonus += 8
        if any(k in name for k in ("plantilla", "template", "blank")):
            bonus -= 8
        return ext_score + bonus

    def is_bad_pdf_name(name: str) -> bool:
        name = name.lower()
        return any(k in name for k in ("correo", "presentacion", "presentación", "mail", "mensaje"))

    def score_pdf(f: Path):
        name = f.name.lower()
        bonus = 0
        if any(k in name for k in ("formatocv", "cv", "expresion", "expresión", "edi")):
            bonus += 10
        if is_bad_pdf_name(name):
            bonus -= 50
        return bonus

    chosen = []
    for folder, files in by_folder.items():
        excels = [f for f in files if f.suffix.lower() in (".xlsx", ".xlsm", ".xls")]
        pdfs = [f for f in files if f.suffix.lower() == ".pdf"]

        if excels:
            excels.sort(key=score_excel, reverse=True)
            chosen.append(excels[0])
        elif pdfs:
            pdfs.sort(key=score_pdf, reverse=True)
            best = pdfs[0]
            if is_bad_pdf_name(best.name):
                continue
            chosen.append(best)

    chosen.sort(key=lambda x: str(x).lower())
    return chosen

def find_ec_detail_rows(ws, col=6, rmin=1, rmax=200):
    """
    Busca filas donde en la columna F diga:
      'DETALLAR LOS CURSOS DECLARADOS'
    Devuelve lista de rows.
    """
    rows = []
    for r in range(rmin, rmax + 1):
        v = ws.cell(r, col).value
        if v and "DETALLAR LOS CURSOS DECLARADOS" in str(v).upper():
            rows.append(r)
    return rows

def fill_estudios_complementarios(ws, base_col, data, debug_log=None):
    """
    base_col = columna del slot (F, H, J...)
    Se escribe en base_col + (offset) pero aquí ya trabajamos directo con base_col.
    """
    blocks = data.get("estudios_complementarios_blocks", [])
    if not blocks:
        return

    # 1) buscar filas "DETALLAR LOS CURSOS DECLARADOS" dentro del slot
    # OJO: base_col apunta a F/H/J..., entonces buscamos en esa misma columna.
    rows = []
    for r in range(1, 250):
        v = ws.cell(r, base_col).value
        if v and "DETALLAR LOS CURSOS DECLARADOS" in str(v).upper():
            rows.append(r)

    if debug_log:
        log_file_append(debug_log, f"[EC] filas DETALLAR en col={base_col}: {rows}")

    if not rows:
        if debug_log:
            log_file_append(debug_log, "[EC] No se encontraron filas para detallar cursos")
        return

    # 2) escribir bloques en orden
    for i, r in enumerate(rows):
        if i >= len(blocks):
            break

        resumen = blocks[i].get("resumen", "").strip()
        if not resumen:
            resumen = "(sin cursos declarados)"

        ws.cell(r, base_col).value = resumen

        if debug_log:
            log_file_append(debug_log, f"[EC] escrito bloque {blocks[i].get('id')} en fila {r}")

# -------------------------
# Write Stage 1 (DP)
# -------------------------
def is_cell_filled(v) -> bool:
    return v is not None and str(v).strip() != ""

def format_course_line(it: dict) -> str:
    centro = norm(it.get("centro", ""))
    cap = norm(it.get("capacitacion", ""))
    fi = norm(it.get("fi", "")) or norm(it.get("fecha_inicio", ""))
    ff = norm(it.get("ff", "")) or norm(it.get("fecha_fin", ""))
    horas = norm(it.get("horas", ""))

    parts = []
    if centro: parts.append(centro)
    if cap: parts.append(cap)
    rango = ""
    if fi or ff:
        rango = f"{fi or '?'}–{ff or '?'}"
        parts.append(rango)
    if horas:
        parts.append(f"{horas}h")

    return "• " + " | ".join(parts).strip()

def write_value_safe_old(ws, row: int, col: int, value):
    """
    Escribe en (row,col) aunque sea celda combinada.
    Si cae en una MergedCell (read-only), busca el top-left del merge y escribe ahí.
    """
    cell = ws.cell(row=row, column=col)

    # Caso normal
    if not isinstance(cell, MergedCell):
        cell.value = value
        return (row, col)

    # Caso merged: hallar rango que contiene esta celda
    for r in ws.merged_cells.ranges:
        if r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col:
            tl = ws.cell(row=r.min_row, column=r.min_col)
            tl.value = value
            return (r.min_row, r.min_col)

    # Si por alguna razón no se encontró el rango, no escribimos
    raise RuntimeError(f"No se encontró rango merged para celda ({row},{col})")

def get_top_left_cell(ws, row: int, col: int):
    """Si (row,col) cae dentro de un merged range, retorna la celda top-left."""
    for r in ws.merged_cells.ranges:
        if r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col:
            return ws.cell(row=r.min_row, column=r.min_col)
    return ws.cell(row=row, column=col)

def clear_value_safe(ws, row: int, col: int):
    cell = get_top_left_cell(ws, row, col)
    cell.value = None
    return cell

def _anchor_of_merged(ws, row: int, col: int):
    """
    Si (row,col) cae dentro de un rango combinado, devuelve (min_row, min_col)
    del rango (la celda "real" editable). Si no, devuelve (row,col).
    """
    coord = ws.cell(row=row, column=col).coordinate
    for r in ws.merged_cells.ranges:
        if coord in r:
            return r.min_row, r.min_col
    return row, col

def write_value_safe(ws, row: int, col: int, value):
    """
    Escribe solo si:
    - no es merged, o
    - es merged PERO el anchor (min_row/min_col) coincide con (row,col).
    Si el anchor está en otra fila, NO escribe (evita malograr otra sección).
    """
    coord = ws.cell(row=row, column=col).coordinate

    for r in ws.merged_cells.ranges:
        if coord in r:
            # anchor real
            ar, ac = r.min_row, r.min_col
            if (ar, ac) != (row, col):
                # ⚠️ No escribas: estarías escribiendo en otra fila
                return False
            ws.cell(row=ar, column=ac).value = value
            return True

    ws.cell(row=row, column=col).value = value
    return True

def write_postulante_stage1_OLD(ws, data: dict, cfg: dict, styles: dict, debug_log: Path):
    """
    Stage 1 (carga de datos):
    - Cabecera (fila header_row): Nombre | DNI | Cel | Email
    - Formación académica: fila 6
    - Estudios complementarios: filas ec_rows (por defecto 8,9,10 para b.1,b.2,b.3)
    """

    max_slots = int(cfg.get("max_postulantes_por_hoja", 20))
    slot_start_col = int(cfg.get("slot_start_col", 6))
    slot_step_cols = int(cfg.get("slot_step_cols", 2))
    header_row = int(cfg.get("slot_header_row", 3))

    # En tu template, FA se coloca en fila 6 (columna del slot)
    fa_row = int(cfg.get("fa_row", 6))

    # En tu template, b.1/b.2/b.3 suelen estar en filas 8/9/10
    # Puedes sobreescribir en config: "ec_rows": [8,9,10]
    ec_rows = cfg.get("ec_rows", [8, 9, 10])
    if isinstance(ec_rows, str):
        # por si te llega "8,9,10"
        ec_rows = [int(x.strip()) for x in ec_rows.split(",") if x.strip()]

    slot = next_free_slot(
        ws,
        max_slots=max_slots,
        slot_start_col=slot_start_col,
        slot_step_cols=slot_step_cols,
        header_row=header_row,
    )
    if slot is None:
        return False, "NO_HAY_SLOT"

    base_col, score_col = slot_columns(slot, slot_start_col, slot_step_cols)

    # Formatea todo el bloque del slot (fondo, bordes, ancho, wrap, etc.)
    apply_slot_format(ws, base_col, score_col, cfg, styles)

    # -------------------------
    # CABECERA (fila 3)
    # -------------------------
    display_name = norm(data.get("nombre_full", ""))
    if not display_name:
        dni_fb = norm(data.get("dni", ""))
        src_fb = norm(data.get("source_label", "")) or norm(data.get("source_file", ""))
        display_name = dni_fb or src_fb or "SIN_NOMBRE"

    dni = norm(data.get("dni", ""))
    cel = norm(data.get("celular", ""))
    email = norm(data.get("email", ""))

    parts = [display_name]
    if dni:
        parts.append(f"DNI: {dni}")
    if cel:
        parts.append(f"Cel: {cel}")
    if email:
        parts.append(f"Email: {email}")

    header_text = " | ".join(parts).strip()
    c_header = ws.cell(row=header_row, column=base_col)
    c_header.value = header_text
    c_header.alignment = c_header.alignment.copy(wrap_text=True, vertical="center")

    # -------------------------
    # FORMACIÓN ACADÉMICA (fila 6)
    # -------------------------
    fa = data.get("formacion_resumen") or ""
    fa = norm(fa)
    if not fa:
        fa = "SIN DATOS DE FORMACIÓN (revisar tabla 47–56)"

    c_fa = ws.cell(row=fa_row, column=base_col)
    c_fa.value = fa
    c_fa.alignment = c_fa.alignment.copy(wrap_text=True, vertical="top")



    # ==========================
    # ESTUDIOS COMPLEMENTARIOS (EC) por bloque b.1..b.N
    # ==========================
    ec = data.get("estudios_complementarios", {}) or {}
    
    blocks = ec.get("blocks", []) or []

    ROW_BASE_EC = int(cfg.get("row_ec_base", 8))  # por defecto b.1 empieza en fila 8

    def format_course_line(x: dict) -> str:
        centro = norm(x.get("centro", ""))
        cap = norm(x.get("capacitacion", ""))
        fi = norm(x.get("fecha_inicio", ""))
        ff = norm(x.get("fecha_fin", ""))
        horas = x.get("horas", "")

        parts = []
        if centro: parts.append(centro)
        if cap: parts.append(cap)
        if fi or ff:
            parts.append(f"{fi} - {ff}".strip(" -"))
        if horas not in ("", None, 0, "0"):
            parts.append(f"{horas}h")
        return " | ".join(parts)

    def block_index(block_id: str) -> int:
        # 'b.1' -> 1, 'b.2' -> 2 ...
        m = re.search(r"\b[bB]\s*\.\s*(\d+)\b", str(block_id))
        return int(m.group(1)) if m else 1

    # Limpia filas EC del slot (para no dejar basura de ejecuciones previas)
    # (limpiamos 10 filas por seguridad)
    #for rr in range(ROW_BASE_EC, ROW_BASE_EC + 10):
        #ws.cell(row=rr, column=base_col).value = None
    for rr in range(ROW_BASE_EC, ROW_BASE_EC + 10):
        write_value_safe(ws, rr, base_col, None)

        # OJO: no toco score_col todavía (por ahora)

    # Escribe cada bloque en su fila
    for b in blocks:
        bid = b.get("id", "")          # 'b.1', 'b.2', ...
        idx = block_index(bid)         # 1..N
        rr = ROW_BASE_EC + (idx - 1)   # fila destino dentro del slot

        items = b.get("items", []) or []
        lines = [format_course_line(x) for x in items if (norm(x.get("centro","")) or norm(x.get("capacitacion","")))]
        resumen = "\n".join(lines).strip() if lines else "(sin cursos declarados)"

        #ws.cell(row=rr, column=base_col).value = resumen
        write_value_safe(ws, rr, base_col, resumen)
        # Asegura que se vea multilinea
        cell = ws.cell(row=rr, column=base_col)
        if cell.alignment:
            cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")
        else:
            from openpyxl.styles import Alignment
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # (Opcional) Si tu plantilla usa solo b.1..b.4, y quieres forzar a 4 filas:
    # for idx in range(1, 5):
    #     rr = ROW_BASE_EC + (idx - 1)
    #     if ws.cell(row=rr, column=base_col).value is None:
    #         ws.cell(row=rr, column=base_col).value = "(sin cursos declarados)"




    # -------------------------
    # LOG
    # -------------------------
    log_append(
        debug_log,
        f"[WRITE_STAGE1] sheet='{ws.title}' slot={slot} base_col={base_col} "
        f"name='{display_name}' dni='{dni}' fa_len={len(fa)} blocks={len(blocks)}",
        also_print=False
    )

    return True, f"SLOT_{slot}"

def write_postulante_stage1(ws, data: dict, cfg: dict, styles: dict, debug_log: Path):
    """
    STAGE 1 (Carga de datos al template):
    1) Datos Personales   -> Cabecera (header_row) en la columna del slot
    2) Formación Académica-> Fila fa_row (por defecto 6)
    3) Estudios Complementarios (EC) -> Filas ec_rows (por defecto 8,9,10,11 para b.1..b.4)

    ⚠️ Nota crítica (openpyxl + celdas combinadas):
    - Si intentas escribir dentro de una celda combinada que NO sea el "anchor" (top-left),
      openpyxl devuelve un MergedCell y su .value es read-only.
    - Para evitar "malograr" otras filas (ej: que EC termine escribiendo sobre FA),
      write_value_safe() SOLO escribe si:
        a) No es merged, o
        b) Es merged y la celda destino ES el anchor del merge.
      Caso contrario, NO escribe y deja log de warning (debug).
    """

    # ============================================================
    # 0) IMPORTS LOCALES (por si este archivo está modularizado)
    # ============================================================
    import re
    import json
    from openpyxl.styles import Alignment

    # ============================================================
    # 1) VARIABLES / CONFIG
    # ============================================================
    max_slots = int(cfg.get("max_postulantes_por_hoja", 20))
    slot_start_col = int(cfg.get("slot_start_col", 6))
    slot_step_cols = int(cfg.get("slot_step_cols", 2))
    header_row = int(cfg.get("slot_header_row", 3))

    # Formación Académica en el template (por defecto fila 6)
    fa_row = int(cfg.get("fa_row", 6))

    # Estudios Complementarios: por defecto b.1..b.4 en filas 8..11
    ec_rows = cfg.get("ec_rows", [8, 9, 10, 11])
    if isinstance(ec_rows, str):
        ec_rows = [int(x.strip()) for x in ec_rows.split(",") if x.strip()]

    # ============================================================
    # 2) HELPERS (merge-safe + prints estructurados)
    # ============================================================
    def _safe_json(obj) -> str:
        """Pretty-print robusto (evita romper por tipos raros)."""
        try:
            return json.dumps(obj, ensure_ascii=False, indent=2, default=str)
        except Exception:
            return str(obj)

    def _cell_is_in_merged_range(cell_coord: str):
        """Retorna el merged range (openpyxl range) si coord pertenece, sino None."""
        for rng in ws.merged_cells.ranges:
            if cell_coord in rng:
                return rng
        return None

    def write_value_safe(row: int, col: int, value, label: str = ""):
        """
        Escribe value en (row,col) SOLO si:
        - la celda no es merged, o
        - la celda es merged y (row,col) es el anchor del merge (top-left).
        Si no se puede, no escribe y lo reporta.
        Retorna siempre el "cell" de ws.cell(row,col) para que puedas setear alignment,
        pero ojo: si es MergedCell no debes tocar .value directamente.
        """
        cell = ws.cell(row=row, column=col)
        coord = cell.coordinate
        rng = _cell_is_in_merged_range(coord)

        if rng:
            anchor = (rng.min_row, rng.min_col)
            if (row, col) != anchor:
                # NO escribimos para no pisar otra fila/columna por anchor "lejano"
                print(f"[WARN] write_value_safe SKIP (MergedCell read-only) "
                      f"{label} coord={coord} in_merge={rng} anchor={anchor} value_preview='{str(value)[:50]}'")
                return cell  # devolvemos cell (puede ser MergedCell)
            # escribimos en el anchor real:
            anchor_cell = ws.cell(row=rng.min_row, column=rng.min_col)
            anchor_cell.value = value
            return anchor_cell  # devolvemos el anchor (es escribible)

        # No merge -> escritura directa
        cell.value = value
        return cell

    def clear_value_safe(row: int, col: int, label: str = ""):
        """Limpieza merge-safe (respeta la misma lógica de write_value_safe)."""
        write_value_safe(row, col, None, label=label)

    def apply_wrap(cell, vertical="top"):
        """Aplica wrap_text y vertical sin romper si no hay alignment previo."""
        if cell is None:
            return
        if getattr(cell, "alignment", None):
            cell.alignment = cell.alignment.copy(wrap_text=True, vertical=vertical)
        else:
            cell.alignment = Alignment(wrap_text=True, vertical=vertical)

    def block_index(block_id: str) -> int:
        """Convierte 'b.1' -> 1, 'b.2' -> 2, etc."""
        m = re.search(r"\b[bB]\s*\.\s*(\d+)\b", str(block_id))
        return int(m.group(1)) if m else 1

    def is_header_row(it: dict) -> bool:
        """
        Detecta filas 'cabecera' que aparecen dentro de items por arte del parse:
        - nro: No. / N° / Nº
        - centro: Centro de estudios / Nombre de la Entidad ó Empresa
        - cap: Capacitación / Nombre del Proyecto
        - fecha_inicio: "Fecha de Inicio"
        """
        nro = norm(it.get("nro", "")).strip().lower()
        centro = norm(it.get("centro", "")).strip().lower()
        cap = norm(it.get("capacitacion", "")).strip().lower()
        fi = norm(it.get("fecha_inicio", "")).strip().lower()

        if nro in ("no.", "n°", "nº", "nro", "nro."):
            return True
        if centro in ("centro de estudios", "nombre de la entidad ó empresa"):
            return True
        if cap in ("capacitacion", "capacitación", "nombre del proyecto"):
            return True
        if "fecha de inicio" in fi:
            return True
        return False

    def format_course_line(x: dict) -> str:
        """Línea humana: Centro | Curso | FechaIni - FechaFin | Nh"""
        centro = norm(x.get("centro", ""))
        cap = norm(x.get("capacitacion", ""))
        fi = norm(x.get("fecha_inicio", ""))
        ff = norm(x.get("fecha_fin", ""))
        horas = x.get("horas", "")

        out = []
        if centro:
            out.append(centro)
        if cap:
            out.append(cap)

        fechas = " - ".join([p for p in [fi, ff] if p]).strip()
        if fechas:
            out.append(fechas)

        h = str(horas).strip() if horas is not None else ""
        if h and h not in ("0", "0.0"):
            if not h.lower().endswith("h"):
                h = f"{h}h"
            out.append(h)

        return " | ".join(out).strip()

    # ============================================================
    # 3) HALLAR SLOT LIBRE
    # ============================================================
    slot = next_free_slot(
        ws,
        max_slots=max_slots,
        slot_start_col=slot_start_col,
        slot_step_cols=slot_step_cols,
        header_row=header_row,
    )
    if slot is None:
        return False, "NO_HAY_SLOT"

    base_col, score_col = slot_columns(slot, slot_start_col, slot_step_cols)

    # Formateo del slot (fondos, bordes, widths, wrap base, etc.)
    apply_slot_format(ws, base_col, score_col, cfg, styles)

    # ============================================================
    # 4) PRINT ESTRUCTURADO: RESUMEN DEL INPUT (DATA)
    # ============================================================
    # (No imprimimos TODO el dict gigante, imprimimos un resumen útil)
    ec_dbg = data.get("estudios_complementarios") or {}
    blocks_dbg = (ec_dbg.get("blocks") or [])
    debug_payload = {
        "sheet": ws.title,
        "slot": slot,
        "base_col": base_col,
        "score_col": score_col,
        "cfg": {
            "header_row": header_row,
            "fa_row": fa_row,
            "ec_rows": ec_rows,
            "max_slots": max_slots,
            "slot_start_col": slot_start_col,
            "slot_step_cols": slot_step_cols,
        },
        "datos_personales": {
            "nombre_full": data.get("nombre_full", ""),
            "dni": data.get("dni", ""),
            "celular": data.get("celular", ""),
            "email": data.get("email", ""),
            "source_label": data.get("source_label", ""),
            "source_file": data.get("source_file", ""),
        },
        "formacion_academica": {
            "formacion_resumen_len": len(norm(data.get("formacion_resumen") or "")),
            "formacion_resumen_preview": (norm(data.get("formacion_resumen") or "")[:120] + "…")
                                         if len(norm(data.get("formacion_resumen") or "")) > 120
                                         else norm(data.get("formacion_resumen") or ""),
        },
        "estudios_complementarios": {
            "blocks_count": len(blocks_dbg),
            "blocks_ids": [b.get("id") for b in blocks_dbg],
            "total_horas": ec_dbg.get("total_horas", 0),
        }
    }
    print("[STAGE1 DEBUG] INPUT SUMMARY:\n" + _safe_json(debug_payload))

    # ============================================================
    # 5) SECCIÓN 1: DATOS PERSONALES (CABECERA)
    # ============================================================
    # Armamos el texto visible de cabecera:
    display_name = norm(data.get("nombre_full", ""))
    if not display_name:
        dni_fb = norm(data.get("dni", ""))
        src_fb = norm(data.get("source_label", "")) or norm(data.get("source_file", ""))
        display_name = dni_fb or src_fb or "SIN_NOMBRE"

    dni = norm(data.get("dni", ""))
    cel = norm(data.get("celular", ""))
    email = norm(data.get("email", ""))

    parts = [display_name]
    if dni:
        parts.append(f"DNI: {dni}")
    if cel:
        parts.append(f"Cel: {cel}")
    if email:
        parts.append(f"Email: {email}")

    header_text = " | ".join(parts).strip()

    print("[STAGE1 DEBUG] DATOS PERSONALES (CABECERA):\n" + _safe_json({
        "row": header_row,
        "col": base_col,
        "header_text": header_text
    }))

    c_header = write_value_safe(header_row, base_col, header_text, label="CABECERA")
    apply_wrap(c_header, vertical="center")

    # ============================================================
    # 6) SECCIÓN 2: FORMACIÓN ACADÉMICA (FA)
    # ============================================================
    fa = norm(data.get("formacion_resumen") or "")
    if not fa:
        fa = "SIN DATOS DE FORMACIÓN (revisar tabla 47–56)"

    print("[STAGE1 DEBUG] FORMACIÓN ACADÉMICA:\n" + _safe_json({
        "row": fa_row,
        "col": base_col,
        "fa_len": len(fa),
        "fa_preview": (fa[:160] + "…") if len(fa) > 160 else fa,
    }))

    c_fa = write_value_safe(fa_row, base_col, fa, label="FORMACION_ACADEMICA")
    apply_wrap(c_fa, vertical="top")

    # ============================================================
    # 7) SECCIÓN 3: ESTUDIOS COMPLEMENTARIOS (EC)
    # ============================================================
    ec = data.get("estudios_complementarios") or {}
    blocks = ec.get("blocks") or []

    # base: si no usas ec_rows, se usa row_ec_base
    ROW_BASE_EC = int(cfg.get("row_ec_base", ec_rows[0] if ec_rows else 8))

    print("[STAGE1 DEBUG] ESTUDIOS COMPLEMENTARIOS (CONFIG):\n" + _safe_json({
        "ROW_BASE_EC": ROW_BASE_EC,
        "ec_rows": ec_rows,
        "blocks_count": len(blocks),
        "blocks_ids": [b.get("id") for b in blocks],
    }))

    # 7.1 Limpieza segura SOLO de las filas EC definidas
    #     (esto evita limpiar accidentalmente FA u otras filas del template)
    if ec_rows:
        for rr in ec_rows:
            clear_value_safe(rr, base_col, label=f"EC_CLEAR_r{rr}")
    else:
        for rr in range(ROW_BASE_EC, ROW_BASE_EC + 10):
            clear_value_safe(rr, base_col, label=f"EC_CLEAR_r{rr}")

    # 7.2 Escritura por bloque b.1..b.N -> a su fila correspondiente
    for b in blocks:
        bid = b.get("id", "")
        idx = block_index(bid)  # 1..N

        # decide fila destino según ec_rows
        if ec_rows and 1 <= idx <= len(ec_rows):
            rr = ec_rows[idx - 1]
        else:
            rr = ROW_BASE_EC + (idx - 1)

        items = b.get("items") or []

        # DEDUP + FILTRO DE CABECERAS + FILTRO DE VACÍOS
        seen = set()
        lines = []
        for it in items:
            if is_header_row(it):
                continue
            if not (norm(it.get("centro", "")) or norm(it.get("capacitacion", ""))):
                continue

            key = (
                norm(it.get("centro", "")).lower(),
                norm(it.get("capacitacion", "")).lower(),
                norm(it.get("fecha_inicio", "")),
                norm(it.get("fecha_fin", "")),
                str(it.get("horas", "")).strip()
            )
            if key in seen:
                continue
            seen.add(key)

            lines.append(format_course_line(it))

        resumen = "\n".join([x for x in lines if x]).strip()
        if not resumen:
            resumen = "(sin cursos declarados)"

        print("[STAGE1 DEBUG] EC WRITE BLOCK:\n" + _safe_json({
            "block_id": bid,
            "block_idx": idx,
            "dest_row": rr,
            "dest_col": base_col,
            "items_in_block": len(items),
            "lines_written": len(lines),
            "resumen_preview": (resumen[:180] + "…") if len(resumen) > 180 else resumen,
        }))

        c_ec = write_value_safe(rr, base_col, resumen, label=f"EC_{bid}")
        apply_wrap(c_ec, vertical="top")

    # ============================================================
    # 8) LOG FINAL
    # ============================================================
    log_append(
        debug_log,
        f"[WRITE_STAGE1] sheet='{ws.title}' slot={slot} base_col={base_col} "
        f"name='{display_name}' dni='{dni}' fa_len={len(fa)} blocks={len(blocks)}",
        also_print=False
    )

    return True, f"SLOT_{slot}"

def write_postulante_experiencia_general(ws, data: dict, cfg: dict, styles: dict, debug_log):
    """
    Writer desacoplado de EDI:
    - Lee data["exp_general_resumen"] o data["exp_general"]["resumen"]
    - Escribe en la plantilla usando cfg["exp_general_row"] y slot columns.
    """

    max_slots = int(cfg.get("max_postulantes_por_hoja", 20))
    slot_start_col = int(cfg.get("slot_start_col", 6))
    slot_step_cols = int(cfg.get("slot_step_cols", 2))
    header_row = int(cfg.get("slot_header_row", 3))

    # fila destino (detectada por task_00)
    eg_row = int(cfg.get("exp_general_row", 12))

    slot = next_free_slot(
        ws,
        max_slots=max_slots,
        slot_start_col=slot_start_col,
        slot_step_cols=slot_step_cols,
        header_row=header_row,
    )
    if slot is None:
        return False, "NO_HAY_SLOT"

    base_col, score_col = slot_columns(slot, slot_start_col, slot_step_cols)

    # Texto a escribir
    eg = data.get("exp_general") or {}
    texto = norm(data.get("exp_general_resumen") or eg.get("resumen") or "")
    if not texto:
        texto = "(sin experiencia general declarada)"

    # Limpia solo esa celda antes de escribir
    clear_value_safe(ws, eg_row, base_col)

    cell = write_value_safe(ws, eg_row, base_col, texto)
    cell.alignment = (cell.alignment.copy(wrap_text=True, vertical="top")
                      if cell.alignment else Alignment(wrap_text=True, vertical="top"))

    # Print estructurado (debug)
    if cfg.get("debug", False):
        print(f"[WRITE EG] sheet='{ws.title}' slot={slot} base_col={base_col} eg_row={eg_row} len={len(texto)} dias={data.get('exp_general_dias', 0)}")

    log_append(
        debug_log,
        f"[WRITE_EG] sheet='{ws.title}' slot={slot} base_col={base_col} eg_row={eg_row} len={len(texto)}",
        also_print=False
    )

    return True, f"SLOT_{slot}"


# -------------------------
# Main
# -------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", required=True, help="Ruta raíz de ProcesoSelección")
    ap.add_argument("--only-proc", default="", help="Nombre exacto de proceso para ejecutar solo uno (opcional)")
    ap.add_argument("--limit", type=int, default=0, help="Limitar postulantes (0=sin limite)")
    args = ap.parse_args()

    cfg = json.loads(Path("configs/config.json").read_text(encoding="utf-8"))
    root = Path(args.root)

    use_ocr = bool(cfg.get("pdf", {}).get("use_ocr", False))
    processed_mode = cfg.get("processed_mode", "copy")
    if processed_mode not in ("copy", "move"):
        processed_mode = "copy"

    styles = build_slot_styles(cfg)

    print(f"[task_30] root={root} use_ocr={use_ocr} processed_mode={processed_mode}")

    procesos = [p for p in root.iterdir() if p.is_dir()]
    procesos.sort(key=lambda p: p.name.lower())
    print(f"[task_30] procesos detectados={len(procesos)}")

    ok_proc = 0
    skip_proc = 0

    for proc_dir in procesos:
        proceso = proc_dir.name
        if args.only_proc and proceso != args.only_proc:
            continue

        in_dir = proc_dir / IN_FOLDER_NAME
        out_dir = proc_dir / OUT_FOLDER_NAME

        print(f"\n[task_30] --- PROCESO: {proceso} ---")
        print(f"  in_dir = {in_dir}")
        print(f"  out_dir= {out_dir}")

        if not in_dir.exists():
            print("  [SKIP] No existe 009. EDI RECIBIDA")
            skip_proc += 1
            continue

        if not out_dir.exists():
            print("  [SKIP] No existe 011. INSTALACIÓN DE COMITÉ")
            skip_proc += 1
            continue

        debug_log = out_dir / "debug_task_30.log"
        ensure_dir(out_dir / PROCESADOS_SUBFOLDER)
        log_append(debug_log, f"== TASK 30 | PROCESO: {proceso} ==")

        tpl = find_process_template(out_dir)
        if not tpl:
            log_append(debug_log, f"[SKIP] No encuentro plantilla '{PROCESS_OUTPUT_TEMPLATE_PREFIX}*.xlsx' en {out_dir}")
            print(f"  [SKIP] No encuentro plantilla '{PROCESS_OUTPUT_TEMPLATE_PREFIX}*.xlsx' en 011")
            skip_proc += 1
            continue

        log_append(debug_log, f"[CFG] template={tpl}")
        print(f"  template = {tpl.name}")

        files = choose_one_file_per_postulante_folder(in_dir)
        if args.limit and args.limit > 0:
            files = files[: args.limit]

        print(f"  archivos elegidos = {len(files)}")
        log_append(debug_log, f"[INFO] archivos elegidos={len(files)}")

        if not files:
            print("  [SKIP] No hay archivos elegibles")
            skip_proc += 1
            continue

        wb = load_workbook(tpl)
        sheet_idx = 1
        ws = get_or_create_eval_sheet(wb, EVAL_SHEET_TEMPLATE, sheet_idx)

        rows_log = []
        rows_flat = []

        for fp in files:
            try:
                if fp.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
                    print("Antes de entrar a DATA: parse_eoi_excel")
                    data = parse_eoi_excel(fp, debug=False)
                    ftype = "EXCEL"
                else:
                    data = parse_eoi_pdf(fp, use_ocr=use_ocr)
                    ftype = "PDF"

                data["source_label"] = fp.parent.name

                print(data)
                #log_append(debug_log, f"[CHECK] wrote cell F6? value_now='{ws.cell(6, base_col).value}'", also_print=True)

                ok, where = write_postulante_stage1(ws, data, cfg, styles, debug_log)

                if not ok and where == "NO_HAY_SLOT":
                    sheet_idx += 1
                    ws = get_or_create_eval_sheet(wb, EVAL_SHEET_TEMPLATE, sheet_idx)
                    log_append(debug_log, f"[INFO] Nueva hoja: {ws.title}", also_print=False)
                    ok, where = write_postulante_stage1(ws, data, cfg, styles, debug_log)

                if not ok:
                    rows_log.append([ts(), str(fp), ftype, "ERROR", where])
                    continue

                # copiar/mover a procesados
                proc_out = out_dir / PROCESADOS_SUBFOLDER
                dest = proc_out / fp.name
                if dest.exists():
                    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    dest = proc_out / f"{fp.stem}_{stamp}{fp.suffix}"

                if processed_mode == "move":
                    shutil.move(str(fp), str(dest))
                else:
                    shutil.copy2(str(fp), str(dest))

                rows_log.append([ts(), str(fp), ftype, "OK", f"{ws.title}:{where}"])
                rows_flat.append([proceso, str(fp), ftype, ws.title, where,
                                  data.get("dni",""), data.get("nombre_full",""),
                                  data.get("celular",""), data.get("email","")])

            except Exception as e:
                rows_log.append([ts(), str(fp), fp.suffix.lower(), "ERROR", repr(e)])

        out_xlsx = out_dir / f"Cuadro_Evaluacion_{safe_filename(proceso)}.xlsx"
        #wb.save(out_xlsx)
        out_saved = safe_save_workbook(wb, out_xlsx, debug_log)
        print(f"  ✅ guardado: {out_saved}")


        write_csv(out_dir / "task30_log.csv", ["fecha", "archivo", "tipo", "estado", "detalle"], rows_log)
        write_csv(out_dir / "task30_consolidado.csv",
                  ["proceso","archivo_origen","tipo","hoja","slot","dni","nombre","celular","email"],
                  rows_flat)

        print(f"  ✅ guardado: {out_xlsx}")
        log_append(debug_log, f"[DONE] Guardado: {out_xlsx}")

        ok_proc += 1

    print(f"\n[task_30] OK_PROCESOS={ok_proc} SKIP_PROCESOS={skip_proc}")


if __name__ == "__main__":
    main()
