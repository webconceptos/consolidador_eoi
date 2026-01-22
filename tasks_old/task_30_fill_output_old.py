# tasks/task_30_fill_output.py
# -*- coding: utf-8 -*-
"""
Task 30: fill_output

Toma la plantilla local "Revision Preliminar ... .xlsx" por proceso (en 011),
lee 011/consolidado.jsonl y llena el cuadro por slots en la hoja "Evaluación CV"
creando hojas adicionales cuando sea necesario.

Salidas en <Proceso>/011. INSTALACIÓN DE COMITÉ/
  - Cuadro_Evaluacion_<proceso>.xlsx
  - fill_log.csv
  - debug_fill_output.log
"""

import argparse
import csv
import json
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple

from openpyxl import load_workbook

IN_FOLDER_NAME = "009. EDI RECIBIDA"
OUT_FOLDER_NAME = "011. INSTALACIÓN DE COMITÉ"

DEFAULT_EVAL_SHEET = "Evaluación CV"
CONFIG_LAYOUT = "config_layout.json"
IN_JSONL = "consolidado.jsonl"

OUT_XLSX_PREFIX = "Cuadro_Evaluacion_"
OUT_FILL_LOG = "fill_log.csv"
OUT_DEBUG_LOG = "debug_fill_output.log"


# -------------------------
# Helpers
# -------------------------
def ts():
    return datetime.now().isoformat(timespec="seconds")


def norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def safe_filename(s: str, max_len: int = 120) -> str:
    s = norm(s)
    s = re.sub(r"[^A-Za-z0-9._-]+", "_", s)
    return s[:max_len] if len(s) > max_len else s


def ensure_dir(p: Path):
    p.mkdir(parents=True, exist_ok=True)


def log_append(path: Path, msg: str):
    ensure_dir(path.parent)
    with path.open("a", encoding="utf-8") as f:
        f.write(f"[{ts()}] {msg}\n")


def load_json(path: Path, default=None):
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def read_jsonl(path: Path) -> List[Dict[str, Any]]:
    items = []
    if not path.exists():
        return items
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            items.append(json.loads(line))
    return items


def write_csv(path: Path, header: List[str], rows: List[List[Any]]):
    ensure_dir(path.parent)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)


# -------------------------
# Slot logic
# -------------------------
def slot_columns(slot_index: int, slot_start_col: int = 6, slot_step_cols: int = 2):
    base_col = slot_start_col + slot_index * slot_step_cols
    score_col = base_col + 1
    return base_col, score_col


def next_free_slot(ws, max_slots: int, slot_start_col: int = 6, slot_step_cols: int = 2, header_row: int = 3):
    for i in range(max_slots):
        base_col, _ = slot_columns(i, slot_start_col, slot_step_cols)
        v = ws.cell(row=header_row, column=base_col).value
        if v is None or str(v).strip() == "" or "NOMBRE DEL CONSULTOR" in str(v).upper():
            return i
    return None


def get_or_create_eval_sheet(wb, template_name: str, sheet_index: int):
    """
    Hoja 1 = plantilla original.
    Hoja 2..N = copia de la plantilla, renombrada "Evaluación CV (2)", etc.
    """
    if sheet_index == 1:
        return wb[template_name]

    title = f"{template_name} ({sheet_index})"
    if title in wb.sheetnames:
        return wb[title]

    new_ws = wb.copy_worksheet(wb[template_name])
    new_ws.title = title
    return new_ws


# -------------------------
# Write postulante (carga simple, sin evaluación fina aún)
# -------------------------
def write_postulante(ws, data: dict, cfg_global: dict, layout: dict) -> Tuple[bool, str]:
    max_slots = int(layout.get("max_postulantes_por_hoja") or cfg_global.get("max_postulantes_por_hoja", 20))
    slot_start_col = int(layout.get("slot_start_col", 6))
    slot_step_cols = int(layout.get("slot_step_cols", 2))
    header_row = int(layout.get("slot_header_row", 3))

    slot = next_free_slot(ws, max_slots, slot_start_col, slot_step_cols, header_row)
    if slot is None:
        return False, "NO_HAY_SLOT"

    base_col, score_col = slot_columns(slot, slot_start_col, slot_step_cols)

    # header: nombre
    nombre = norm(data.get("nombre_full", "")) or norm(data.get("dni", "")) or "SIN_NOMBRE"
    ws.cell(row=header_row, column=base_col).value = nombre

    # filas “default” (ajustables luego por layout/rules)
    # Formación: fila 6
    titulo = norm(data.get("titulo", ""))
    bachiller = norm(data.get("bachiller", ""))
    egresado = norm(data.get("egresado", ""))
    #ws.cell(row=6, column=base_col).value = f"Título: {titulo} | Bachiller: {bachiller} | Egresado: {egresado}"

    fa = data.get("formacion_academica", {}) or {}
    items = fa.get("items", []) or []

    # construir resumen compacto
    resumen = []
    if data.get("titulo"):
        resumen.append(f"{data.get('titulo')}")
    if data.get("bachiller"):
        resumen.append(f"{data.get('bachiller')}")
    if data.get("egresado"):
        resumen.append(f"{data.get('egresado')}")

    # si no hay nada, igual deja huella
    if not resumen and items:
        # muestra lo primero lleno
        it0 = items[0]
        resumen.append(f"{it0.get('nivel','')} - {it0.get('especialidad','')}")

    ws.cell(row=6, column=base_col).value = " | ".join([x for x in resumen if norm(x)])[:1200]


    # Cursos: fila 8 (texto) + fila 11 (conteo simple)
    cursos = data.get("cursos", []) or []
    cursos_txt = " | ".join([norm(x) for x in cursos if norm(x)])[:1200]
    ws.cell(row=8, column=base_col).value = cursos_txt
    ws.cell(row=11, column=score_col).value = 1 if cursos_txt else 0

    # Experiencia general/específica: filas 13 y 17
    gen_days = int(data.get("exp_general_dias", 0) or 0)
    spec_days = int(data.get("exp_especifica_dias", 0) or 0)
    ws.cell(row=13, column=base_col).value = f"{gen_days} días"
    ws.cell(row=17, column=base_col).value = f"{spec_days} días"

    # Deseables: filas 20/21
    ws.cell(row=20, column=score_col).value = 1 if bool(data.get("java_ok", False)) else 0
    ws.cell(row=21, column=score_col).value = 1 if bool(data.get("oracle_ok", False)) else 0

    # Pie: fila 22
    ws.cell(row=22, column=base_col).value = (
        f"DNI: {norm(data.get('dni',''))} | "
        f"Email: {norm(data.get('email',''))} | "
        f"Cel: {norm(data.get('celular',''))}"
    )

    return True, f"SLOT_{slot}"


# -------------------------
# Resolve template
# -------------------------
def find_template_in_011(out_dir: Path) -> Optional[Path]:
    """
    Busca "Revision Preliminar*.xlsx/.xlsm/.xls" en 011.
    """
    exts = (".xlsx", ".xlsm", ".xls")
    prefix = "revision preliminar"
    candidates = []
    if not out_dir.exists():
        return None
    for p in out_dir.iterdir():
        if not p.is_file():
            continue
        if p.name.startswith("~$"):
            continue
        if p.suffix.lower() not in exts:
            continue
        if not p.name.lower().startswith(prefix):
            continue
        candidates.append(p)
    if not candidates:
        return None
    # el más reciente
    candidates.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return candidates[0]


# -------------------------
# Main
# -------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", type=str, default="", help="Carpeta raíz con procesos")
    ap.add_argument("--only-proc", type=str, default="", help="Procesar solo procesos cuyo nombre contenga este texto")
    ap.add_argument("--limit", type=int, default=0, help="Limitar N postulantes por proceso (0=sin límite)")
    args = ap.parse_args()

    cfg_global = load_json(Path("configs/config.json"), default={}) or {}
    root = Path(args.root) if args.root.strip() else Path(cfg_global.get("input_root", ""))

    if not str(root).strip():
        raise SystemExit("Debes pasar --root o definir input_root en configs/config.json")
    if not root.exists():
        raise SystemExit(f"No existe root: {root}")

    only_filter = norm(args.only_proc).lower()

    procesos = [p for p in root.iterdir() if p.is_dir()]
    procesos.sort(key=lambda x: x.name.lower())

    print(f"[fill_output] root = {root}")
    print(f"[fill_output] procesos detectados = {len(procesos)}")

    ok_proc = 0
    skip_proc = 0

    for proc_dir in procesos:
        proceso = proc_dir.name
        if only_filter and only_filter not in proceso.lower():
            continue

        out_dir = proc_dir / OUT_FOLDER_NAME
        if not out_dir.exists():
            skip_proc += 1
            print(f"  - SKIP: {proceso} (no existe 011)")
            continue

        layout = load_json(out_dir / CONFIG_LAYOUT, default={}) or {}

        template_path = None
        # preferir layout.template_path si existe
        tp = layout.get("template_path")
        if tp:
            tp2 = Path(tp)
            if tp2.exists():
                template_path = tp2
        if template_path is None:
            template_path = find_template_in_011(out_dir)

        if template_path is None or not template_path.exists():
            skip_proc += 1
            print(f"  - SKIP: {proceso} (no encuentro plantilla 'Revision Preliminar...' en 011)")
            continue

        jsonl_path = out_dir / IN_JSONL
        if not jsonl_path.exists():
            skip_proc += 1
            print(f"  - SKIP: {proceso} (no existe {IN_JSONL}; ejecuta Task 20)")
            continue

        items = read_jsonl(jsonl_path)
        if args.limit and args.limit > 0:
            items = items[: args.limit]

        # abrir plantilla
        wb = load_workbook(template_path)

        # hoja base
        eval_sheet = layout.get("eval_sheet_name") or DEFAULT_EVAL_SHEET
        if eval_sheet not in wb.sheetnames:
            eval_sheet = wb.sheetnames[0]  # fallback
        sheet_idx = 1
        ws = wb[eval_sheet]

        debug_log = out_dir / OUT_DEBUG_LOG
        if debug_log.exists():
            debug_log.unlink(missing_ok=True)
        log_append(debug_log, f"== PROCESO: {proceso} ==")
        log_append(debug_log, f"template: {template_path}")
        log_append(debug_log, f"eval_sheet_base: {eval_sheet}")
        log_append(debug_log, f"postulantes: {len(items)}")

        fill_log_rows: List[List[Any]] = []
        fill_header = ["fecha", "proceso", "dni", "nombre", "hoja", "slot", "estado", "detalle"]

        for i, data in enumerate(items, start=1):
            dni = norm(data.get("dni", ""))
            nombre = norm(data.get("nombre_full", ""))
            log_append(debug_log, f"[{i}/{len(items)}] START dni={dni} nombre={nombre}")

            ok, where = write_postulante(ws, data, cfg_global, layout)

            if not ok and where == "NO_HAY_SLOT":
                sheet_idx += 1
                ws = get_or_create_eval_sheet(wb, eval_sheet, sheet_idx)
                log_append(debug_log, f"[INFO] Nueva hoja: {ws.title}")
                ok, where = write_postulante(ws, data, cfg_global, layout)

            if ok:
                fill_log_rows.append([ts(), proceso, dni, nombre, ws.title, where, "OK", ""])
                log_append(debug_log, f"[OK] {ws.title}:{where}")
            else:
                fill_log_rows.append([ts(), proceso, dni, nombre, ws.title, where, "ERROR", where])
                log_append(debug_log, f"[ERROR] {where}")

        # guardar salida final
        out_xlsx = out_dir / f"{OUT_XLSX_PREFIX}{safe_filename(proceso)}.xlsx"
        wb.save(out_xlsx)

        write_csv(out_dir / OUT_FILL_LOG, fill_header, fill_log_rows)
        log_append(debug_log, f"[DONE] saved: {out_xlsx}")
        log_append(debug_log, f"[DONE] fill_log: {OUT_FILL_LOG} rows={len(fill_log_rows)}")

        ok_proc += 1
        print(f"  - OK: {proceso} -> {out_xlsx.name} | postulantes={len(items)}")

    print("")
    print(f"[fill_output] resumen: OK_PROCESOS={ok_proc} SKIP_PROCESOS={skip_proc}")


if __name__ == "__main__":
    main()
